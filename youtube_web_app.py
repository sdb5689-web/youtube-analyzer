# ================================================================
# 🎬 YouTube 분석 웹앱 v1.0 (Streamlit)
#
# 설치: pip install streamlit requests youtube-transcript-api
#       openpyxl gspread google-auth pytrends
# 실행: streamlit run youtube_web_app.py
# ================================================================
# ================================================================
# 필수 패키지 자동 설치 (Streamlit Cloud / 서버 환경 대응)
# ================================================================
import subprocess as _sp, sys as _sys

def _ensure(pkg, import_name=None):
    """설치되지 않은 패키지를 자동 설치"""
    try:
        __import__(import_name or pkg)
    except ImportError:
        _sp.run(
            [_sys.executable, "-m", "pip", "install", pkg, "--quiet"],
            check=False
        )

_ensure("pytrends")
_ensure("streamlit-sortables", "streamlit_sortables")
_ensure("gspread")
_ensure("google-auth",   "google.oauth2")
_ensure("openpyxl")
_ensure("youtube-transcript-api", "youtube_transcript_api")
_ensure("openai")
_ensure("lxml")
# ================================================================

import streamlit as st
try:
    from streamlit_sortables import sort_items as _sort_items
    _HAS_SORTABLES = True
except ImportError:
    _HAS_SORTABLES = False
import requests, json, re, time, os, io
import urllib.parse
from collections import Counter, defaultdict
from datetime import datetime

# ─── 선택적 임포트 ────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GSHEET = True
except ImportError:
    HAS_GSHEET = False

# ── Whisper (OpenAI STT) ──────────────────────
try:
    import openai as _openai_lib
    HAS_WHISPER = True
except ImportError:
    HAS_WHISPER = False

# ── yt-dlp (오디오 다운로드) ──────────────────
try:
    import yt_dlp as _yt_dlp_lib
    HAS_YTDLP = True
except ImportError:
    HAS_YTDLP = False

# ================================================================
# 페이지 설정
# ================================================================
st.set_page_config(
    page_title="🎬 YouTube 분석 도구",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── 커스텀 CSS ───────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700&display=swap');

/* ══════════════════════════════════════════════════════════════
   CSS 변수 (디자인 토큰)
══════════════════════════════════════════════════════════════ */
:root {
  /* 브랜드 컬러 */
  --c-primary:      #2563eb;
  --c-primary-dark: #1d4ed8;
  --c-primary-deep: #1e40af;
  --c-primary-soft: #eff6ff;
  --c-primary-mid:  #bfdbfe;

  /* 중성색 */
  --c-bg:       #f8faff;
  --c-surface:  #ffffff;
  --c-border:   #e4e8f4;
  --c-border-hi:#c7d2fe;

  /* 텍스트 */
  --c-txt:      #1e293b;
  --c-txt-sub:  #475569;
  --c-txt-mute: #94a3b8;

  /* 상태색 */
  --c-danger:   #dc2626;
  --c-warn:     #d97706;
  --c-success:  #059669;

  /* 반경 */
  --r-sm:  8px;
  --r-md:  11px;
  --r-lg:  14px;
  --r-xl:  18px;

  /* 그림자 */
  --sh-xs: 0 1px 3px rgba(0,0,0,.06);
  --sh-sm: 0 2px 10px rgba(0,0,0,.07);
  --sh-md: 0 4px 18px rgba(0,0,0,.10);
  --sh-lg: 0 8px 32px rgba(0,0,0,.13);

  /* 전환 */
  --tr: all .18s ease;
}

/* ══════════════════════════════════════════════════════════════
   글로벌 리셋
══════════════════════════════════════════════════════════════ */
*, *::before, *::after { box-sizing: border-box; }

.stApp {
  font-family: "Noto Sans KR", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
  background: var(--c-bg) !important;
}

/* ── 메인 컨테이너 ── */
.main .block-container {
  padding: 1.4rem 1.8rem 2.5rem 1.8rem !important;
  max-width: 1280px !important;
}

/* ── 전역 스크롤바 ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; border-radius: 4px; }
::-webkit-scrollbar-thumb { background: #c7d2e8; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

/* ══════════════════════════════════════════════════════════════
   메인 헤더
══════════════════════════════════════════════════════════════ */
.main-header {
  background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 55%, #2563eb 100%);
  padding: 26px 34px;
  border-radius: var(--r-xl);
  color: white;
  margin-bottom: 26px;
  box-shadow: 0 8px 32px rgba(37,99,235,.28);
  position: relative;
  overflow: hidden;
}
.main-header::before {
  content: '';
  position: absolute;
  top: -60%; right: -8%;
  width: 320px; height: 320px;
  background: rgba(255,255,255,.06);
  border-radius: 50%;
  pointer-events: none;
}
.main-header::after {
  content: '';
  position: absolute;
  bottom: -40%; left: -5%;
  width: 200px; height: 200px;
  background: rgba(255,255,255,.04);
  border-radius: 50%;
  pointer-events: none;
}
.main-header h1 {
  margin: 0;
  font-size: 1.95rem;
  font-weight: 700;
  letter-spacing: -.03em;
}
.main-header p {
  margin: 7px 0 0 0;
  opacity: .88;
  font-size: 0.96rem;
}

/* ══════════════════════════════════════════════════════════════
   메트릭 카드
══════════════════════════════════════════════════════════════ */
.metric-card {
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: var(--r-lg);
  padding: 20px 22px;
  text-align: center;
  box-shadow: var(--sh-sm);
  transition: var(--tr);
}
.metric-card:hover {
  transform: translateY(-3px);
  box-shadow: var(--sh-md);
  border-color: var(--c-border-hi);
}
.metric-card .value {
  font-size: 2.0rem;
  font-weight: 700;
  color: var(--c-primary);
  line-height: 1.15;
}
.metric-card .label {
  font-size: 0.82rem;
  color: var(--c-txt-sub);
  margin-top: 6px;
  font-weight: 500;
}

/* ══════════════════════════════════════════════════════════════
   비디오 카드
══════════════════════════════════════════════════════════════ */
.video-card {
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: var(--r-lg);
  padding: 18px 20px;
  margin-bottom: 14px;
  box-shadow: var(--sh-xs);
  transition: var(--tr);
}
.video-card:hover {
  transform: translateY(-2px);
  box-shadow: var(--sh-md);
  border-color: var(--c-border-hi);
}
.video-title {
  font-size: 1.0rem;
  font-weight: 700;
  color: var(--c-txt);
  line-height: 1.45;
}
.video-meta {
  color: var(--c-txt-sub);
  font-size: 0.85rem;
  margin-top: 7px;
  line-height: 1.5;
}

/* ══════════════════════════════════════════════════════════════
   배지
══════════════════════════════════════════════════════════════ */
.badge-hot    { background: var(--c-primary);  color:#fff; padding:3px 10px; border-radius:20px; font-size:.76rem; font-weight:600; display:inline-block; }
.badge-good   { background: var(--c-warn);     color:#fff; padding:3px 10px; border-radius:20px; font-size:.76rem; font-weight:600; display:inline-block; }
.badge-new    { background: var(--c-success);  color:#fff; padding:3px 10px; border-radius:20px; font-size:.76rem; font-weight:600; display:inline-block; }
.badge-norm   { background: var(--c-txt-mute); color:#fff; padding:3px 10px; border-radius:20px; font-size:.76rem; font-weight:600; display:inline-block; }
.badge-shorts { background: var(--c-danger);   color:#fff; padding:3px 10px; border-radius:20px; font-size:.76rem; font-weight:600; display:inline-block; }

/* ── 통계 행 ── */
.stat-row  { display:flex; gap:10px; flex-wrap:wrap; margin-top:10px; }
.stat-item {
  background: #f1f5f9;
  border-radius: var(--r-sm);
  padding: 4px 12px;
  font-size: .83rem;
  color: var(--c-txt-sub);
  font-weight: 500;
}

/* ── 키워드 태그 ── */
.keyword-tag {
  display: inline-block;
  background: var(--c-primary-soft);
  color: var(--c-primary-dark);
  border-radius: 20px;
  padding: 4px 14px;
  margin: 3px;
  font-size: .82rem;
  border: 1px solid var(--c-primary-mid);
  font-weight: 500;
  transition: background .15s;
}
.keyword-tag:hover { background: #dbeafe; }

/* ── 섹션 제목 ── */
.section-title {
  font-size: 1.06rem;
  font-weight: 700;
  color: var(--c-primary-deep);
  border-left: 4px solid var(--c-primary);
  padding-left: 12px;
  margin: 22px 0 13px 0;
}


/* ── 요약 텍스트 박스 ── */
.summary-text {
  font-size: .88rem !important;
  line-height: 1.68 !important;
  color: #475569 !important;
  background: #f8faff !important;
  border-left: 3px solid #bfdbfe !important;
  border-radius: 0 8px 8px 0 !important;
  padding: 10px 14px !important;
  margin: 4px 0 10px 0 !important;
  word-break: keep-all !important;
  white-space: pre-wrap !important;
}

/* ── 전역 Markdown blockquote 크기 제한 ── */
.stMarkdown blockquote,
.stMarkdown blockquote p,
[data-testid="stMarkdownContainer"] blockquote,
[data-testid="stMarkdownContainer"] blockquote p {
  font-size: .88rem !important;
  line-height: 1.68 !important;
  color: #475569 !important;
}

/* ── 대본 박스 ── */
.transcript-box {
  background: #f8fafc;
  border: 1px solid #e2e8f0;
  border-radius: var(--r-md);
  padding: 14px;
  font-size: .86rem;
  max-height: 220px;
  overflow-y: auto;
  line-height: 1.72;
  color: #334155;
}

/* ── 뷰 바 ── */
.view-bar {
  height: 4px;
  border-radius: 3px;
  background: linear-gradient(90deg, #f97316, #ea580c);
  margin-top: 3px;
}

/* ── 랭크 배지 ── */
.rank-badge {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  background: #ef4444;
  color: white;
  border-radius: 50%;
  width: 21px; height: 21px;
  font-size: .70rem;
  font-weight: 700;
  margin-right: 5px;
  flex-shrink: 0;
}
.rank-badge.gold   { background: linear-gradient(135deg,#f59e0b,#d97706); }
.rank-badge.silver { background: linear-gradient(135deg,#94a3b8,#64748b); }
.rank-badge.bronze { background: linear-gradient(135deg,#b45309,#92400e); }

/* ══════════════════════════════════════════════════════════════
   Streamlit 메인 영역 오버라이드
══════════════════════════════════════════════════════════════ */
/* Expander */
div[data-testid="stExpander"] {
  border-radius: var(--r-md) !important;
  border: 1px solid var(--c-border) !important;
  box-shadow: var(--sh-xs) !important;
  background: var(--c-surface) !important;
  margin-bottom: 8px !important;
  overflow: hidden !important;
}
div[data-testid="stExpander"] summary {
  font-size: .90rem !important;
  font-weight: 600 !important;
  color: var(--c-txt) !important;
  padding: 10px 14px !important;
  background: var(--c-surface) !important;
  min-height: 38px !important;
}
div[data-testid="stExpander"] summary:hover {
  background: var(--c-primary-soft) !important;
}

/* 탭 */
.stTabs [data-baseweb="tab-list"] {
  gap: 4px !important;
  border-bottom: 2px solid var(--c-border) !important;
  background: transparent !important;
}
.stTabs [data-baseweb="tab"] {
  font-size: .92rem !important;
  font-weight: 600 !important;
  padding: 8px 18px !important;
  border-radius: var(--r-sm) var(--r-sm) 0 0 !important;
  color: var(--c-txt-sub) !important;
  border: none !important;
  background: transparent !important;
  transition: var(--tr) !important;
}
.stTabs [aria-selected="true"] {
  color: var(--c-primary) !important;
  background: var(--c-primary-soft) !important;
  border-bottom: 2px solid var(--c-primary) !important;
}

/* ── 메인 버튼 공통 ── */
.stButton button {
  border-radius: var(--r-md) !important;
  font-weight: 600 !important;
  font-size: .90rem !important;
  transition: var(--tr) !important;
  border: 1.5px solid var(--c-border) !important;
  background: var(--c-surface) !important;
  color: var(--c-txt-sub) !important;
  padding: 6px 16px !important;
  min-height: 36px !important;
  box-shadow: var(--sh-xs) !important;
}
.stButton button:hover {
  border-color: var(--c-primary-mid) !important;
  color: var(--c-primary) !important;
  background: var(--c-primary-soft) !important;
  box-shadow: 0 3px 12px rgba(37,99,235,.12) !important;
}

/* ── 메인 입력 필드 ── */
.stTextInput input, .stTextArea textarea {
  border-radius: var(--r-md) !important;
  border: 1.5px solid var(--c-border) !important;
  font-size: .92rem !important;
  padding: 8px 12px !important;
  transition: var(--tr) !important;
  background: var(--c-surface) !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
  border-color: var(--c-primary) !important;
  box-shadow: 0 0 0 3px rgba(37,99,235,.12) !important;
}

/* ── 셀렉트박스 ── */
.stSelectbox > div > div {
  border-radius: var(--r-md) !important;
  border: 1.5px solid var(--c-border) !important;
  font-size: .90rem !important;
}

/* ── 핫 서브토픽 버튼 ── */
.hot-topic-btn .stButton button {
  background: linear-gradient(135deg,#f97316,#ea580c) !important;
  color: white !important;
  border: none !important;
  border-radius: 16px !important;
  font-size: .84rem !important;
  padding: 5px 12px !important;
  font-weight: 600 !important;
  box-shadow: 0 2px 8px rgba(249,115,22,.30) !important;
}
.hot-topic-btn .stButton button:hover {
  background: linear-gradient(135deg,#ea580c,#c2410c) !important;
  box-shadow: 0 4px 14px rgba(249,115,22,.45) !important;
  transform: translateY(-1px) !important;
}

/* ══════════════════════════════════════════════════════════════
   Material Icons 텍스트 렌더링 수정 (expander 화살표만 숨김)
══════════════════════════════════════════════════════════════ */

/* Expander 화살표 텍스트만 숨김 (토글 버튼 제외) */
[data-testid="stExpander"] summary span[data-testid="stIconMaterial"],
details summary [data-testid="stIconMaterial"] {
  font-size: 0 !important;
  line-height: 0 !important;
  overflow: hidden !important;
  max-width: 0 !important;
  opacity: 0 !important;
  position: absolute !important;
}

/* Material Symbols 폰트 설정 (expander 내부만, 토글 버튼 제외) */
[data-testid="stExpander"] span[data-testid="stIconMaterial"] {
  font-family: "Material Symbols Rounded", "Material Icons" !important;
}

/* ══════════════════════════════════════════════════════════════
   사이드바 토글 버튼 (열림 << / 닫힘 >>)
   Streamlit 1.55+ 실제 DOM: stSidebarCollapseButton + stSidebarCollapsedControl
══════════════════════════════════════════════════════════════ */

/* ── 열린 상태: 닫기 버튼 (<<)
   DOM: stSidebarHeader > stSidebarCollapseButton > button[kind=headerNoPadding]
   JS에서 showSidebarCollapse=false 일 때 visibility:hidden → !important 로 강제 표시 ── */
[data-testid="stSidebarCollapseButton"] {
  display: inline-flex !important;
  visibility: visible !important;
  opacity: 1 !important;
  overflow: visible !important;
  align-items: center !important;
  margin-left: 4px !important;
}
[data-testid="stSidebarCollapseButton"] button,
[data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"] {
  opacity: 1 !important;
  visibility: visible !important;
  display: inline-flex !important;
  align-items: center !important;
  justify-content: center !important;
  background: rgba(37,99,235,.10) !important;
  border: 1.5px solid rgba(37,99,235,.25) !important;
  border-radius: 8px !important;
  padding: 5px 8px !important;
  cursor: pointer !important;
  color: #2563eb !important;
  min-width: 32px !important;
  min-height: 32px !important;
  transition: background .15s ease, box-shadow .15s ease !important;
  position: relative !important;
  z-index: 9999 !important;
}
[data-testid="stSidebarCollapseButton"] button:hover,
[data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"]:hover {
  background: rgba(37,99,235,.20) !important;
  box-shadow: 0 2px 8px rgba(37,99,235,.25) !important;
}
/* 닫기 버튼 내부 아이콘 (SVG/span) */
[data-testid="stSidebarCollapseButton"] button svg,
[data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"] svg {
  opacity: 1 !important;
  visibility: visible !important;
  display: block !important;
  width: 20px !important;
  height: 20px !important;
  color: inherit !important;
  fill: currentColor !important;
}
[data-testid="stSidebarCollapseButton"] button span,
[data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"] span {
  opacity: 1 !important;
  visibility: visible !important;
  display: inline-block !important;
  font-size: 20px !important;
  line-height: 1 !important;
}

/* ── stSidebarHeader 레이아웃 (버튼 오른쪽 정렬) ── */
[data-testid="stSidebarHeader"] {
  display: flex !important;
  align-items: center !important;
  justify-content: flex-end !important;
  padding: 4px 8px !important;
  min-height: 36px !important;
}

/* ── 닫힌 상태: 열기 버튼 (>>)
   DOM: div[data-testid="stSidebarCollapsedControl"] > button ── */
[data-testid="stSidebarCollapsedControl"] {
  position: fixed !important;
  top: 50% !important;
  left: 0 !important;
  transform: translateY(-50%) !important;
  z-index: 99999 !important;
  display: flex !important;
  align-items: center !important;
  visibility: visible !important;
  opacity: 1 !important;
}
[data-testid="stSidebarCollapsedControl"] button {
  opacity: 1 !important;
  visibility: visible !important;
  display: inline-flex !important;
  align-items: center !important;
  justify-content: center !important;
  background: #ffffff !important;
  border: 1.5px solid #bfdbfe !important;
  border-left: none !important;
  border-radius: 0 10px 10px 0 !important;
  padding: 14px 8px !important;
  box-shadow: 3px 0 14px rgba(37,99,235,.20) !important;
  cursor: pointer !important;
  color: #2563eb !important;
  min-width: 30px !important;
  min-height: 50px !important;
  transition: background .15s ease, box-shadow .15s ease !important;
  z-index: 99999 !important;
}
[data-testid="stSidebarCollapsedControl"] button:hover {
  background: #eff6ff !important;
  box-shadow: 3px 0 22px rgba(37,99,235,.30) !important;
  color: #1d4ed8 !important;
}
/* 열기 버튼 내부 아이콘 */
[data-testid="stSidebarCollapsedControl"] button svg {
  opacity: 1 !important;
  visibility: visible !important;
  display: block !important;
  width: 20px !important;
  height: 20px !important;
  color: inherit !important;
  fill: currentColor !important;
}
[data-testid="stSidebarCollapsedControl"] button span {
  opacity: 1 !important;
  visibility: visible !important;
  display: inline-block !important;
  font-size: 20px !important;
  line-height: 1 !important;
}

/* ── 아이콘 폰트 공통 설정 ── */
[data-testid="stSidebarCollapseButton"] span[data-testid="stIconMaterial"],
[data-testid="stSidebarCollapsedControl"] span[data-testid="stIconMaterial"],
[data-testid="stSidebarCollapseButton"] .e14lo0b10,
[data-testid="stSidebarCollapsedControl"] .e14lo0b10 {
  font-family: "Material Symbols Rounded", "Material Icons", sans-serif !important;
  font-style: normal !important;
  font-weight: 400 !important;
  font-feature-settings: 'liga' !important;
  -webkit-font-smoothing: antialiased !important;
  display: inline-block !important;
  visibility: visible !important;
  opacity: 1 !important;
  font-size: 20px !important;
  line-height: 1 !important;
  color: inherit !important;
  overflow: visible !important;
  width: auto !important;
  height: auto !important;
  max-width: none !important;
  position: static !important;
}

/* ══════════════════════════════════════════════════════════════
   사이드바 — 모던 클린 디자인
══════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
  overflow-y: auto !important;
  background: #f5f8ff !important;
  border-right: 1px solid #dde5f5 !important;
}
[data-testid="stSidebar"] > div:first-child {
  padding: 0.8rem 0.9rem 1.4rem 0.9rem !important;
}
[data-testid="stSidebar"] * {
  font-family: "Noto Sans KR", -apple-system, BlinkMacSystemFont, sans-serif !important;
  font-size: .85rem !important;
  letter-spacing: -.01em;
}
[data-testid="stSidebar"] hr {
  margin: 9px 0 !important;
  border: none !important;
  border-top: 1px solid #dde5f5 !important;
}

/* ── 사이드바 헤딩 ── */
[data-testid="stSidebar"] h2 {
  font-size: .78rem !important;
  font-weight: 700 !important;
  color: var(--c-txt-mute) !important;
  text-transform: uppercase !important;
  letter-spacing: .07em !important;
  margin: 14px 0 5px 0 !important;
}
[data-testid="stSidebar"] h3 {
  font-size: .74rem !important;
  font-weight: 700 !important;
  color: var(--c-txt-mute) !important;
  text-transform: uppercase !important;
  letter-spacing: .06em !important;
  margin: 9px 0 4px 0 !important;
}

/* ── 사이드바 텍스트 ── */
[data-testid="stSidebar"] p {
  margin: 0 !important;
  line-height: 1.48 !important;
  font-size: .86rem !important;
  color: var(--c-txt-sub) !important;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
[data-testid="stSidebar"] [data-baseweb="form-control-label"],
[data-testid="stSidebar"] .stRadio > div > label,
[data-testid="stSidebar"] .stCheckbox > div > label,
[data-testid="stSidebar"] [class*="label"],
[data-testid="stSidebar"] [class*="Label"] {
  font-size: .83rem !important;
  font-weight: 600 !important;
  color: var(--c-txt-sub) !important;
  margin-bottom: 4px !important;
}
[data-testid="stSidebar"] .stCaption,
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"],
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
  font-size: .68rem !important;
  line-height: 1.5 !important;
  color: var(--c-txt-mute) !important;
}

/* ── 사이드바 입력 ── */
[data-testid="stSidebar"] .stTextInput,
[data-testid="stSidebar"] .stTextArea,
[data-testid="stSidebar"] .stSelectbox,
[data-testid="stSidebar"] .stSlider,
[data-testid="stSidebar"] .stCheckbox,
[data-testid="stSidebar"] .stRadio {
  margin-bottom: 4px !important;
}
[data-testid="stSidebar"] .stTextInput input {
  font-size: .86rem !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #fff !important;
  padding: 7px 12px !important;
  color: var(--c-txt) !important;
  box-shadow: var(--sh-xs) !important;
  transition: var(--tr) !important;
}
[data-testid="stSidebar"] .stTextInput input:focus {
  border-color: var(--c-primary) !important;
  box-shadow: 0 0 0 3px rgba(37,99,235,.12) !important;
}
[data-testid="stSidebar"] .stTextArea textarea {
  font-size: .86rem !important;
  min-height: 60px !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #fff !important;
  color: var(--c-txt) !important;
  box-shadow: var(--sh-xs) !important;
}
[data-testid="stSidebar"] .stSelectbox > div > div {
  font-size: .86rem !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #fff !important;
  box-shadow: var(--sh-xs) !important;
}
[data-testid="stSidebar"] .stRadio label,
[data-testid="stSidebar"] .stCheckbox label {
  font-size: .83rem !important;
  color: var(--c-txt-sub) !important;
  font-weight: 500 !important;
}

/* ── element-container 간격 ── */
[data-testid="stSidebar"] .element-container {
  margin-bottom: 3px !important;
  padding-bottom: 0 !important;
}

/* ── 사이드바 일반 버튼 ── */
[data-testid="stSidebar"] .stButton {
  margin-top: 2px !important;
  margin-bottom: 2px !important;
}
[data-testid="stSidebar"] .stButton button {
  font-size: .83rem !important;
  font-weight: 600 !important;
  padding: 5px 12px !important;
  min-height: 32px !important;
  height: auto !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #ffffff !important;
  color: #334155 !important;
  transition: var(--tr) !important;
  white-space: nowrap !important;
  box-shadow: var(--sh-xs) !important;
}
[data-testid="stSidebar"] .stButton button:hover {
  background: var(--c-primary-soft) !important;
  border-color: var(--c-primary-mid) !important;
  color: var(--c-primary) !important;
  box-shadow: 0 2px 10px rgba(37,99,235,.13) !important;
}

/* ── 검색 시작 버튼 (Primary) ── */
[data-testid="stSidebar"] .stButton button[kind="primary"],
[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] {
  background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
  color: #ffffff !important;
  border: none !important;
  font-size: 1.10rem !important;
  font-weight: 700 !important;
  padding: 11px 18px !important;
  min-height: 44px !important;
  border-radius: var(--r-md) !important;
  box-shadow: 0 4px 16px rgba(37,99,235,.38) !important;
  letter-spacing: .01em !important;
}
[data-testid="stSidebar"] .stButton button[kind="primary"] p,
[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] p,
[data-testid="stSidebar"] .stButton button[kind="primary"] span,
[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] span {
  color: #ffffff !important;
  font-size: 1.10rem !important;
  font-weight: 700 !important;
}
[data-testid="stSidebar"] .stButton button[kind="primary"]:hover,
[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"]:hover {
  background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
  box-shadow: 0 6px 22px rgba(37,99,235,.48) !important;
  transform: translateY(-1px) !important;
}

/* ── 아이콘 소형 버튼 ── */
[data-testid="stSidebar"] .stButton button[kind="secondary"] {
  font-size: .83rem !important;
  padding: 3px 6px !important;
  min-height: 28px !important;
}

/* ── 다크모드 토글 버튼 ── */
[data-testid="stSidebar"] .dark-mode-btn .stButton button {
  font-size: .78rem !important;
  padding: 3px 10px !important;
  min-height: 26px !important;
  border-radius: 20px !important;
  font-weight: 700 !important;
}

/* ── 즐겨찾기 폴더 ── */
[data-testid="stSidebar"] .fav-folder-header {
  background: var(--c-primary-soft);
  border: 1px solid var(--c-primary-mid);
  border-radius: var(--r-md);
  padding: 7px 12px;
  margin: 6px 0 3px 0;
}
[data-testid="stSidebar"] .fav-folder-badge {
  font-size: .71rem;
  background: #dbeafe;
  color: var(--c-primary-dark);
  padding: 2px 7px;
  border-radius: 20px;
  font-weight: 600;
}

/* ── 최근 검색 기록 ── */
[data-testid="stSidebar"] .hist-btn .stButton button {
  font-size: .83rem !important;
  padding: 4px 11px !important;
  min-height: 28px !important;
  border-radius: 14px !important;
  background: var(--c-primary-soft) !important;
  border-color: var(--c-primary-mid) !important;
  color: var(--c-primary-dark) !important;
  font-weight: 500 !important;
  text-align: left !important;
  justify-content: flex-start !important;
  box-shadow: none !important;
}
[data-testid="stSidebar"] .hist-btn .stButton button:hover {
  background: #dbeafe !important;
  border-color: #60a5fa !important;
}

/* ── 관련영상 link_button ── */
[data-testid="stSidebar"] .stLinkButton a {
  font-size: .83rem !important;
  padding: 4px 6px !important;
  min-height: 28px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #fff8f8 !important;
  color: var(--c-danger) !important;
  font-weight: 700 !important;
  text-decoration: none !important;
  transition: var(--tr) !important;
}
[data-testid="stSidebar"] .stLinkButton a:hover {
  background: #fee2e2 !important;
  border-color: #f87171 !important;
}

/* ── 선택박스 소형 ── */
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] {
  min-height: 30px !important;
  font-size: .84rem !important;
}

/* ── 사이드바 Expander ── */
[data-testid="stSidebar"] [data-testid="stExpander"] {
  margin-bottom: 5px !important;
  border-radius: var(--r-md) !important;
  border: 1.5px solid #dde5f5 !important;
  background: #ffffff !important;
  overflow: hidden !important;
  box-shadow: var(--sh-xs) !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary {
  padding: 7px 13px !important;
  font-size: .85rem !important;
  font-weight: 700 !important;
  min-height: 34px !important;
  color: #334155 !important;
  background: #ffffff !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover {
  background: var(--c-primary-soft) !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] > div:last-child {
  background: #f8faff !important;
  padding: 8px 11px !important;
}

/* ── Secrets 로드 현황 expander 내부 글자 크기 축소 ── */
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stCaptionContainer"],
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stCaptionContainer"] p,
[data-testid="stSidebar"] [data-testid="stExpander"] .stCaption,
[data-testid="stSidebar"] [data-testid="stExpander"] .stCaption p,
[data-testid="stSidebar"] [data-testid="stExpander"] small {
  font-size: .68rem !important;
  line-height: 1.5 !important;
  color: var(--c-txt-mute) !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stAlert"] p,
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stAlert"] {
  font-size: .68rem !important;
  line-height: 1.5 !important;
  padding: 6px 10px !important;
}
/* success/info 박스 (Google Sheets 자동 로드 완료 등) */
[data-testid="stSidebar"] [data-testid="stAlert"][data-baseweb="notification"] p,
[data-testid="stSidebar"] [data-testid="stAlert"][data-baseweb="notification"] {
  font-size: .68rem !important;
  line-height: 1.5 !important;
}
/* st.success, st.info 공통 */
[data-testid="stSidebar"] .stSuccess p,
[data-testid="stSidebar"] .stInfo p,
[data-testid="stSidebar"] .stSuccess,
[data-testid="stSidebar"] .stInfo {
  font-size: .68rem !important;
  line-height: 1.5 !important;
  padding: 5px 10px !important;
}



/* ── 사이드바 multiselect ── */
[data-testid="stSidebar"] .stMultiSelect > div > div {
    border-radius: var(--r-md) !important;
    border: 1.5px solid #dde5f5 !important;
    background: #fff !important;
    font-size: .85rem !important;
    box-shadow: var(--sh-xs) !important;
}
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
    background: var(--c-primary-soft) !important;
    border: 1px solid var(--c-primary-mid) !important;
    border-radius: 14px !important;
    color: var(--c-primary-dark) !important;
    font-size: .78rem !important;
    font-weight: 600 !important;
    padding: 2px 8px !important;
}
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] span {
    color: var(--c-primary-dark) !important;
}
[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] button {
    color: var(--c-primary) !important;
    background: transparent !important;
    border: none !important;
    min-height: unset !important;
    box-shadow: none !important;
    padding: 0 2px !important;
}
/* 다크모드 multiselect */


/* ── 액션 버튼 색상 — 목록 지우기(danger) / 구글시트 내보내기(success) ── */
[data-testid="stSidebar"] .btn-danger .stButton button {
  background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%) !important;
  color: #ffffff !important;
  border: none !important;
  font-size: .65rem !important;
  font-weight: 700 !important;
  padding: 5px 6px !important;
  min-height: 30px !important;
  height: auto !important;
  border-radius: 8px !important;
  box-shadow: 0 2px 8px rgba(239,68,68,.30) !important;
  transition: all .15s ease !important;
  letter-spacing: 0 !important;
  white-space: normal !important;
  overflow: visible !important;
  line-height: 1.3 !important;
  width: 100% !important;
}
[data-testid="stSidebar"] .btn-danger .stButton button:hover {
  background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%) !important;
  box-shadow: 0 4px 14px rgba(220,38,38,.40) !important;
  transform: translateY(-1px) !important;
}
/* StyledEllipsizedDiv (Streamlit 내부 white-space:nowrap 강제 오버라이드) */
[data-testid="stSidebar"] .btn-danger .stButton button > div,
[data-testid="stSidebar"] .btn-danger .stButton button > div > div {
  white-space: normal !important;
  overflow: visible !important;
  text-overflow: unset !important;
  display: block !important;
  width: 100% !important;
}
[data-testid="stSidebar"] .btn-danger .stButton button p,
[data-testid="stSidebar"] .btn-danger .stButton button span {
  color: #ffffff !important;
  font-size: .65rem !important;
  font-weight: 700 !important;
  white-space: normal !important;
  overflow: visible !important;
  display: inline !important;
}

[data-testid="stSidebar"] .btn-success .stButton button {
  background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important;
  color: #ffffff !important;
  border: none !important;
  font-size: .65rem !important;
  font-weight: 700 !important;
  padding: 5px 6px !important;
  min-height: 30px !important;
  height: auto !important;
  border-radius: 8px !important;
  box-shadow: 0 2px 8px rgba(22,163,74,.30) !important;
  transition: all .15s ease !important;
  letter-spacing: 0 !important;
  white-space: normal !important;
  overflow: visible !important;
  line-height: 1.3 !important;
  width: 100% !important;
}
[data-testid="stSidebar"] .btn-success .stButton button:hover {
  background: linear-gradient(135deg, #15803d 0%, #166534 100%) !important;
  box-shadow: 0 4px 14px rgba(21,128,61,.40) !important;
  transform: translateY(-1px) !important;
}
/* StyledEllipsizedDiv 오버라이드 */
[data-testid="stSidebar"] .btn-success .stButton button > div,
[data-testid="stSidebar"] .btn-success .stButton button > div > div {
  white-space: normal !important;
  overflow: visible !important;
  text-overflow: unset !important;
  display: block !important;
  width: 100% !important;
}
[data-testid="stSidebar"] .btn-success .stButton button p,
[data-testid="stSidebar"] .btn-success .stButton button span {
  color: #ffffff !important;
  font-size: .65rem !important;
  font-weight: 700 !important;
  white-space: normal !important;
  overflow: visible !important;
  display: inline !important;
}
/* disabled 상태 */
[data-testid="stSidebar"] .btn-success .stButton button:disabled,
[data-testid="stSidebar"] .btn-success .stButton button[disabled] {
  background: linear-gradient(135deg, #86efac 0%, #6ee7b7 100%) !important;
  box-shadow: none !important;
  opacity: .6 !important;
  transform: none !important;
  cursor: not-allowed !important;
}

/* ── 다크모드 액션 버튼 ── */
/* ── 정렬 우선순위 버튼 (↑↓✕) ── */
[data-testid="stSidebar"] .stButton button[kind="secondary"]:has(+ *),
[data-testid="stSidebar"] .element-container:has(button) .stButton button {
    min-height: 28px !important;
    padding: 2px 5px !important;
    font-size: .78rem !important;
}
/* 우선순위 번호 배지 행 여백 */
[data-testid="stSidebar"] .sort-priority-row {
    margin: 2px 0 !important;
}
/* ══════════════════════════════════════════════════════════════
   반응형 레이아웃
══════════════════════════════════════════════════════════════ */
/* 넓은 데스크탑 */
@media (min-width: 1400px) {
  .main .block-container { max-width: 1440px !important; }
}

/* 태블릿 */
@media (max-width: 1024px) {
  .main .block-container { padding: 1rem 1.1rem 2rem 1.1rem !important; }
  .main-header { padding: 20px 26px; }
  .main-header h1 { font-size: 1.65rem; }
}

/* 모바일 */
@media (max-width: 768px) {
  .main .block-container { padding: .8rem .8rem 2rem .8rem !important; }
  .main-header { padding: 16px 18px; border-radius: var(--r-lg); margin-bottom: 16px; }
  .main-header h1 { font-size: 1.3rem !important; }
  .main-header p  { font-size: .86rem !important; }
  .video-card { padding: 14px 15px; border-radius: var(--r-md); }
  .video-title { font-size: .96rem; }
  .metric-card { padding: 14px 15px; }
  .metric-card .value { font-size: 1.55rem; }
  .stTabs [data-baseweb="tab"] { font-size: .83rem !important; padding: 6px 11px !important; }
  [data-testid="stSidebar"] { min-width: 90vw !important; max-width: 90vw !important; }
  .stat-row { gap: 7px; }
  .stat-item { font-size: .79rem; padding: 3px 9px; }
  .section-title { font-size: .98rem; }
}

/* 소형 모바일 */
@media (max-width: 480px) {
  .main-header h1 { font-size: 1.1rem !important; }
  .main .block-container { padding: .5rem .5rem 2rem .5rem !important; }
  .metric-card .value { font-size: 1.35rem; }
}

</style>

""", unsafe_allow_html=True)



# ── 다크모드 차트 팔레트 헬퍼 ────────────────────────────────────
def _chart_palette(dark: bool = False) -> dict:
    """다크/라이트 모드에 따른 차트 색상 팔레트 반환"""
    if dark:
        return {
            "bg":          "#1e2038",   # 카드 배경
            "card_border": "#2e3157",
            "grid":        "#2d3060",   # 그리드 라인
            "grid_text":   "#4a5080",   # 그리드 수치 텍스트
            "x_label":     "#5a6090",   # X축 날짜 레이블
            "rise":        "#ff6b6b",   # 상승 (밝은 레드)
            "fall":        "#69db7c",   # 하락 (밝은 그린)
            "flat":        "#ffa94d",   # 보합 (밝은 오렌지)
            "bar_default": "#3d5a80",   # 기본 막대
            "circle_stroke":"#1e2038", # 원 테두리
            "summary_bg":  "#252640",
            "summary_text":"#8890b0",
            "summary_val": "#c5c8e0",
            "sparkline_bg":"#181929",
            "sparkline_border":"#2e3157",
            "sparkline_label":"#6a70a0",
            "area_opacity_hi": "0.30",
            "area_opacity_lo": "0.03",
        }
    else:
        return {
            "bg":          "#ffffff",
            "card_border": "#bfdbfe",
            "grid":        "#eeeeee",
            "grid_text":   "#cccccc",
            "x_label":     "#aaaaaa",
            "rise":        "#E53935",
            "fall":        "#4CAF50",
            "flat":        "#FF9800",
            "bar_default": "#90CAF9",
            "circle_stroke":"#ffffff",
            "summary_bg":  "#f5f5f5",
            "summary_text":"#888888",
            "summary_val": "#333333",
            "sparkline_bg":"#f7f8ff",
            "sparkline_border":"#e8ecff",
            "sparkline_label":"#aaaaaa",
            "area_opacity_hi": "0.18",
            "area_opacity_lo": "0.02",
        }


# ── 다크모드 CSS 동적 적용 ──────────────────────────────────
if st.session_state.get("dark_mode", False):
    st.markdown("""<style>
    /* ════ 다크모드 전체 테마 ════ */
    .stApp, .main, [data-testid="stAppViewContainer"] {
        background-color: #0c0e18 !important;
        color: #e8eaf6 !important;
    }
    .main .block-container { background-color: #0c0e18 !important; }

    /* ── 사이드바 다크 ── */
    [data-testid="stSidebar"] {
        background: #13152a !important;
        border-right-color: #1e2140 !important;
    }
    [data-testid="stSidebar"] * { color: #c5c8e0 !important; }
    [data-testid="stSidebar"] .stTextInput input,
    [data-testid="stSidebar"] .stTextArea textarea,
    [data-testid="stSidebar"] .stSelectbox > div > div {
        background: #1e2038 !important;
        border-color: #2e3157 !important;
        color: #e0e2f5 !important;
    }
    [data-testid="stSidebar"] .stButton button {
        background: #1e2038 !important;
        border-color: #2e3157 !important;
        color: #c5c8e0 !important;
    }
    [data-testid="stSidebar"] .stButton button:hover {
        background: #252646 !important;
        border-color: #5c6ef5 !important;
        color: #aab0ff !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] {
        background: #1a1c30 !important;
        border-color: #2a2e52 !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary {
        background: #1a1c30 !important;
        color: #c5c8e0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] > div:last-child {
        background: #13152a !important;
    }
    /* 검색 시작 버튼 다크 */
    [data-testid="stSidebar"] .stButton button[kind="primary"],
    [data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] {
        color: #ffffff !important;
        background: linear-gradient(135deg,#4f7ef8,#3a5bd4) !important;
        border: none !important;
    }
    [data-testid="stSidebar"] .stButton button[kind="primary"] p,
    [data-testid="stSidebar"] .stButton button[kind="primary"] span,
    [data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] p,
    [data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] span {
        color: #ffffff !important;
    }
    /* 최근검색 */
    [data-testid="stSidebar"] .hist-btn .stButton button {
        background: #252640 !important;
        border-color: #3a3d5c !important;
        color: #8a90c8 !important;
    }
    /* link_button 다크 */
    [data-testid="stSidebar"] .stLinkButton a {
        background: #1e2038 !important;
        border-color: #3a3d5c !important;
        color: #ff8a80 !important;
    }
    [data-testid="stSidebar"] .stLinkButton a:hover {
        background: #2a2440 !important;
        border-color: #e53935 !important;
    }
    /* ── 메인 콘텐츠 다크 ── */
    .video-card {
        background: #15172a !important;
        border-color: #252840 !important;
        color: #e0e2f5 !important;
    }
    .video-card:hover { border-color: #3a3e68 !important; }
    .metric-card {
        background: #15172a !important;
        border-color: #252840 !important;
    }
    .metric-card .value { color: #7c9ef8 !important; }
    .metric-card .label { color: #8a90b0 !important; }
    .video-title  { color: #e8eaf6 !important; }
    .video-meta   { color: #8a90b0 !important; }
    .stat-item    { background: #1e2038 !important; color: #b0b8d0 !important; }
    .transcript-box {
        background: #15172a !important;
        border-color: #252840 !important;
        color: #c5c8e0 !important;
    }
    .section-title { color: #7c9ef8 !important; border-color: #5c6ef5 !important; }
    /* ── 탭 다크 ── */
    .stTabs [data-baseweb="tab-list"] { background: #13152a !important; border-color: #252840 !important; }
    .stTabs [data-baseweb="tab"]       { color: #8a90b0 !important; }
    .stTabs [aria-selected="true"]     { color: #7c9ef8 !important; background: #1e2038 !important; border-color: #5c6ef5 !important; }
    /* ── 입력 필드 메인 ── */
    .stTextInput input, .stTextArea textarea, .stSelectbox > div > div {
        background: #15172a !important;
        border-color: #252840 !important;
        color: #e0e2f5 !important;
    }
    /* ── Expander 메인 ── */
    div[data-testid="stExpander"] {
        border-color: #252840 !important;
        background: #15172a !important;
    }
    div[data-testid="stExpander"] summary {
        background: #15172a !important;
        color: #c5c8e0 !important;
    }
    div[data-testid="stExpander"] summary:hover { background: #1e2038 !important; }
    /* ── 기타 ── */
    .stCheckbox label, .stRadio label { color: #c5c8e0 !important; }
    hr { border-color: #252840 !important; }
    .dm-chart-wrap { background: #1e2038 !important; border-color: #2e3157 !important; }

    /* 다크모드 multiselect */
    [data-testid="stSidebar"] .stMultiSelect > div > div {
        background: #1e2038 !important;
        border-color: #2e3157 !important;
        color: #e0e2f5 !important;
    }
    [data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
        background: #252640 !important;
        border-color: #3a3d5c !important;
        color: #aab0ff !important;
    }
    [data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] span {
        color: #aab0ff !important;
    }
    /* ── 다크모드 토글 버튼 (>> <<) ── */
    /* 닫기 버튼(<<) 컨테이너 강제 표시 - 다크모드 */
    [data-testid="stSidebarCollapseButton"] {
        display: inline-flex !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    /* 닫기 버튼(<<) 스타일 - 다크모드 */
    [data-testid="stSidebarCollapseButton"] button,
    [data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"] {
        background: rgba(138,144,200,.15) !important;
        border: 1.5px solid rgba(138,144,200,.30) !important;
        color: #aab0ff !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebarCollapseButton"] button:hover,
    [data-testid="stSidebarCollapseButton"] [data-testid="stBaseButton-headerNoPadding"]:hover {
        background: rgba(138,144,200,.28) !important;
        color: #c5c8ff !important;
    }
    /* 아이콘 색상 - 다크모드 */
    [data-testid="stSidebarCollapseButton"] span,
    [data-testid="stSidebarCollapseButton"] svg,
    [data-testid="stSidebarCollapsedControl"] span[data-testid="stIconMaterial"] {
        color: #aab0ff !important;
        opacity: 1 !important;
        visibility: visible !important;
    }
    /* 열기 버튼(>>) 스타일 - 다크모드 */
    [data-testid="stSidebarCollapsedControl"] button {
        background: #1e2038 !important;
        border-color: #3a3d5c !important;
        color: #aab0ff !important;
        box-shadow: 3px 0 12px rgba(0,0,0,.40) !important;
    }
    [data-testid="stSidebarCollapsedControl"] button:hover {
        background: #252640 !important;
        color: #c5c8ff !important;
        box-shadow: 3px 0 20px rgba(79,126,248,.30) !important;
    }
        
    /* 다크모드 액션 버튼 */
    [data-testid="stSidebar"] .btn-danger .stButton button {
      background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%) !important;
      box-shadow: 0 2px 8px rgba(220,38,38,.40) !important;
    }
    [data-testid="stSidebar"] .btn-danger .stButton button:hover {
      background: linear-gradient(135deg, #b91c1c 0%, #991b1b 100%) !important;
      box-shadow: 0 4px 14px rgba(185,28,28,.50) !important;
    }
    [data-testid="stSidebar"] .btn-success .stButton button {
      background: linear-gradient(135deg, #15803d 0%, #166534 100%) !important;
      box-shadow: 0 2px 8px rgba(21,128,61,.40) !important;
    }
    [data-testid="stSidebar"] .btn-success .stButton button:hover {
      background: linear-gradient(135deg, #166534 0%, #14532d 100%) !important;
      box-shadow: 0 4px 14px rgba(20,83,45,.50) !important;
    }

    /* 다크모드 요약 텍스트 */
    .summary-text {
      background: #1e2038 !important;
      border-left-color: #4f6aaa !important;
      color: #c5c8e0 !important;
    }
    .stMarkdown blockquote,
    .stMarkdown blockquote p,
    [data-testid="stMarkdownContainer"] blockquote p {
      font-size: .88rem !important;
      color: #c5c8e0 !important;
    }
</style>""", unsafe_allow_html=True)

# ================================================================
# 유틸 함수
# ================================================================
STOPWORDS = set([
    "이","그","저","것","수","을","를","이","가","은","는","에","의","도","로","으로",
    "와","과","한","하는","있는","없는","하고","하면","그리고","하지만","때문에",
    "통해","위해","대한","같은","더","또한","이런","저런","그런","어떤","모든",
    "the","and","is","in","to","of","a","an","that","it","for","on","are","with",
    "이것","저것","그것","이렇게","그렇게","저렇게","때","곳","중","후","전","내","외",
    "씩","만","까지","부터","에서","들","할","될","있","없","않","못","안","다",
])

def fmt(n):
    try: n = int(n)
    except: return "0"
    if n >= 1_000_000_000: return f"{n/1e9:.1f}B"
    if n >= 1_000_000:     return f"{n/1e6:.1f}M"
    if n >= 1_000:         return f"{n/1e3:.0f}K"
    return str(n)

def parse_duration(dur_raw):
    m = re.findall(r'(\d+)([HMS])', dur_raw)
    d = {k: int(v) for v, k in m}
    h, mi, s = d.get('H',0), d.get('M',0), d.get('S',0)
    return f"{h}:{mi:02d}:{s:02d}" if h > 0 else f"{mi}:{s:02d}"

def parse_duration_seconds(dur_str):
    """'1:30' 또는 '0:45' 형식 → 초 단위 정수 반환"""
    try:
        parts = dur_str.split(':')
        if len(parts) == 3:   # h:mm:ss
            return int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])
        elif len(parts) == 2: # m:ss
            return int(parts[0])*60 + int(parts[1])
        return 0
    except:
        return 0

def is_shorts(video):
    """쇼츠 판별: 60초 이하 + #Shorts 태그 or 제목 포함"""
    sec = parse_duration_seconds(video.get("duration", "0:00"))
    tags = [t.lower() for t in video.get("tags", [])]
    title = video.get("title", "").lower()
    desc  = video.get("description", "").lower()
    shorts_keyword = ("shorts" in tags or "#shorts" in title
                      or "#shorts" in desc or "#쇼츠" in title
                      or "#쇼츠" in desc)
    return sec <= 60 or (sec <= 180 and shorts_keyword)

def extract_keywords(text, top_n=15):
    if not text or "자막 없음" in text or len(text) < 50:
        return []
    cleaned = re.sub(r'[^\w\s가-힣]', ' ', text.lower())
    words = cleaned.split()
    filtered = [
        w for w in words
        if len(w) >= 2 and not w.isdigit() and w not in STOPWORDS
        and not re.match(r'^\d+$', w)
    ]
    counter = Counter(filtered)
    return [word for word, _ in counter.most_common(top_n)]

def is_valid_transcript(tr: str) -> bool:
    """실제 사용 가능한 대본인지 판단 (자막 없음/오류 문자열 제외)"""
    if not tr or len(tr) < 20:
        return False
    BAD = ("자막 없음", "youtube-transcript-api 미설치",
           "[Whisper 오류]", "[Gemini 오류", "미설치", "다운로드 실패")
    return not any(tr.startswith(b) or b in tr[:40] for b in BAD)

def clean_transcript(text: str) -> str:
    """대본에서 타임스탬프/메타 태그를 제거하고 순수 대본 텍스트만 반환
    제거 대상:
      - 00:00 / 00:00:00 형식 타임스탬프
      - [🤖 Gemini 분석 (모델명)] 헤더 라인
      - [요약] / [주요 내용] 섹션 (대본 섹션만 추출)
      - ** 마크다운 볼드 기호
    """
    if not text:
        return text

    # ── [대본] 섹션만 추출 (Gemini 분석 결과인 경우) ──────────────
    # "[대본]" 이후 텍스트만 사용
    if "[대본]" in text:
        text = text.split("[대본]", 1)[1].strip()
        # 혹시 뒤에 다른 섹션이 있으면 거기까지만
        for _sec in ["[요약]", "[주요 내용]", "[\U0001f916"]:
            if _sec in text:
                text = text.split(_sec, 1)[0].strip()

    # ── Gemini 헤더 라인 제거 ─────────────────────────────────────
    # "[🤖 Gemini 분석 (gemini-2.5-flash)]" 같은 첫 줄 제거
    lines = text.splitlines()
    clean_lines = []
    for line in lines:
        # 헤더 라인 스킵
        if re.match(r'^\[\U0001f916.*\]', line.strip()):
            continue
        # 타임스탬프 라인 패턴: 줄 시작이 HH:MM:SS 또는 MM:SS
        # 예) 00:05 텍스트... / 01:23:45 텍스트...
        line_stripped = re.sub(
            r'^\s*(\d{1,2}:\d{2}(:\d{2})?)\s*',  # 앞쪽 타임스탬프 제거
            '', line
        )
        # 줄 중간/끝 타임스탬프도 제거 (예: 텍스트 [00:30] 텍스트)
        line_stripped = re.sub(r'\[\d{1,2}:\d{2}(:\d{2})?\]', '', line_stripped)
        # 마크다운 볼드 기호 제거
        line_stripped = line_stripped.replace('**', '')
        # 빈 줄이 아니면 추가
        if line_stripped.strip():
            clean_lines.append(line_stripped.strip())

    # 연속 빈 줄 정리 후 합치기
    result = '\n'.join(clean_lines)
    result = re.sub(r'\n{3,}', '\n\n', result)
    return result.strip()

def summarize_text(text, max_chars=300):
    if not text or "자막 없음" in text or len(text) < 30:
        return "(요약 없음)"
    cleaned = re.sub(r'\s+', ' ', text).strip()
    if len(cleaned) <= max_chars:
        return cleaned
    cut = cleaned[:max_chars]
    for end_char in ['다.', '요.', '죠.', '네.', '.', '!', '?']:
        idx = cut.rfind(end_char)
        if idx > max_chars // 2:
            return cut[:idx+1] + "..."
    return cut + "..."

def build_channel_stats(all_videos):
    ch = defaultdict(lambda: {
        "videos": [], "totalView": 0, "totalLike": 0,
        "totalComment": 0, "subscriber": "비공개"
    })
    for v in all_videos:
        cn = v["channelTitle"]
        ch[cn]["videos"].append(v)
        ch[cn]["totalView"]    += v["viewCount"]
        ch[cn]["totalLike"]    += v["likeCount"]
        ch[cn]["totalComment"] += v["commentCount"]
        ch[cn]["subscriber"]    = v.get("subscriberLabel", "비공개")
    stats = []
    for name, data in ch.items():
        cnt = len(data["videos"])
        stats.append({
            "channel":      name,
            "videoCount":   cnt,
            "subscriber":   data["subscriber"],
            "totalView":    data["totalView"],
            "avgView":      data["totalView"] // cnt if cnt else 0,
            "totalLike":    data["totalLike"],
            "avgLike":      data["totalLike"] // cnt if cnt else 0,
            "totalComment": data["totalComment"],
            "avgComment":   data["totalComment"] // cnt if cnt else 0,
            "videos":       data["videos"],
        })
    stats.sort(key=lambda x: x["totalView"], reverse=True)
    return stats

def get_badge(rank, view_count):
    if rank == 1: return "🥇"
    if rank == 2: return "🥈"
    if rank == 3: return "🥉"
    if view_count >= 1_000_000: return "🔥"
    if view_count >= 100_000:   return "⭐"
    return "▶"

# ================================================================
# YouTube API 함수
# ================================================================
def search_youtube(api_key, keyword, max_r, order_api, video_type="전체", published_after=None, dur_filter=None):
    video_ids = []
    token = None
    # 쇼츠 선택 시 API 레벨에서 4분 미만으로 pre-filter (API quota 절약)
    fetch_extra = max_r  # 후처리 필터 감안해 더 많이 가져옴
    if video_type in ("쇼츠", "동영상"):
        fetch_extra = min(max_r * 3, 50)  # 필터 손실 보정
    while len(video_ids) < max_r:
        params = {
            "key": api_key, "q": keyword,
            "part": "id", "type": "video",
            "maxResults": min(50, fetch_extra - len(video_ids)),
            "order": order_api,
            "regionCode": "KR", "relevanceLanguage": "ko"
        }
        if video_type == "쇼츠":
            params["videoDuration"] = "short"   # 4분 미만 pre-filter
        if dur_filter:
            params["videoDuration"] = dur_filter  # any, short, medium, long
        if published_after:
            params["publishedAfter"] = published_after
        if token: params["pageToken"] = token
        try:
            r = requests.get(
                "https://www.googleapis.com/youtube/v3/search",
                params=params, timeout=10
            )
            d = r.json()
        except Exception as e:
            return None, f"인터넷 연결 오류: {e}"
        if "error" in d:
            return None, f"API 오류 [{d['error']['code']}]: {d['error']['message']}"
        for item in d.get("items", []):
            video_ids.append(item["id"]["videoId"])
        token = d.get("nextPageToken")
        if not token: break
    return video_ids, None


def get_hot_subtopics(api_key: str, main_keyword: str, top_n: int = 10):
    """
    대표 키워드 → 실시간 인기 서브 주제 TOP N 추출

    소스별 비율 고정 (top_n=10 기준):
      B) YouTube 자동완성 (트렌드 연관검색어)  → 60%  (6개)
      C) Google Trends 급상승 연관검색어        → 40%  (4개)
      A) YouTube 인기영상 제목                  →  0%  (미사용)

    각 소스 내에서는 자체 기준(순위·급상승률·조회수×최신성)으로 정렬.
    소스가 부족할 경우 YouTube 자동완성으로 보완.

    Returns:
        (list[dict], None) 또는 (None, error_str)
    """
    import json as _json
    import urllib.parse as _up
    from datetime import datetime as _dt, timezone as _tz

    # ── 소스별 슬롯 계산 ───────────────────────────────────────
    n_suggest = max(1, round(top_n * 0.6))          # 60%  → 6
    n_trends  = top_n - n_suggest                   # 40%  → 4
    n_video   = 0                                   # 미사용

    # 각 소스별 독립 버킷
    bucket_suggest = []   # dict 리스트
    bucket_trends  = []
    bucket_video   = []
    seen_keys = set()

    # ── 헬퍼 ───────────────────────────────────────────────────
    def _fmt_views(n):
        if n >= 100_000_000: return f"{n//100_000_000}억회"
        if n >= 10_000:      return f"{n/10_000:.0f}만회"
        if n >= 1_000:       return f"{n//1_000}천회"
        return f"{n}회"

    def _days_ago(iso_str):
        try:
            pub  = _dt.fromisoformat(iso_str.replace("Z", "+00:00"))
            diff = _dt.now(_tz.utc) - pub
            d    = diff.days
            if d == 0:  return "오늘"
            if d == 1:  return "1일 전"
            if d < 7:   return f"{d}일 전"
            if d < 30:  return f"{d//7}주 전"
            if d < 365: return f"{d//30}개월 전"
            return f"{d//365}년 전"
        except Exception:
            return ""

    def _recency_factor(iso_str):
        try:
            pub  = _dt.fromisoformat(iso_str.replace("Z", "+00:00"))
            diff = _dt.now(_tz.utc) - pub
            d    = diff.days
            if d <= 7:   return 2.0
            if d <= 30:  return 1.5
            if d <= 90:  return 1.2
            if d <= 365: return 1.0
            return 0.6
        except Exception:
            return 1.0

    def _make_item(topic, label, source, score=0, views="", date="",
                   channel="", raw_views=0, trend_val=0, sug_rank=0,
                   sparkline=None):
        return {
            "topic":     topic.strip(),
            "score":     score,         # 소스 내 원시 점수 (정렬용)
            "label":     label,
            "source":    source,
            "views":     views,
            "raw_views": raw_views,
            "trend_val": trend_val,
            "sug_rank":  sug_rank,
            "date":      date,
            "channel":   channel,
            "sparkline": sparkline or [],
        }

    # ──────────────────────────────────────────────────────────────
    # B) YouTube 자동완성 — 트렌드 연관검색어 (60% 슬롯)
    # ──────────────────────────────────────────────────────────────
    try:
        _sug_url = (
            "https://suggestqueries.google.com/complete/search"
            f"?client=firefox&ds=yt"
            f"&q={_up.quote(main_keyword)}"
            f"&hl=ko&gl=KR"
        )
        _sr = requests.get(
            _sug_url, timeout=8,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        )
        if _sr.status_code == 200:
            _data = _json.loads(_sr.text)
            _sugs = _data[1] if len(_data) > 1 else []
            for _rank, _s in enumerate(_sugs[:20]):   # 여유분 20개 수집
                _topic = _s if isinstance(_s, str) else _s[0]
                if not _topic:
                    continue
                _key = _topic.strip().lower()
                if _key == main_keyword.strip().lower() or len(_key) < 2:
                    continue
                if _key not in seen_keys:
                    seen_keys.add(_key)
                    bucket_suggest.append(
                        _make_item(
                            _topic,
                            "🔴 실시간검색", "suggest",
                            score=20 - _rank,   # 1위=19, 20위=0
                            date="실시간",
                            sug_rank=_rank + 1,
                        )
                    )
    except Exception:
        pass

    # ──────────────────────────────────────────────────────────────
    # C) Google Trends — 급상승 연관검색어 (40% 슬롯)
    # ──────────────────────────────────────────────────────────────
    _trends_sparklines = {}
    _trends_available  = False
    try:
        from pytrends.request import TrendReq as _TR
        _pt = _TR(hl="ko-KR", tz=540, timeout=(8, 20))
        _pt.build_payload([main_keyword], cat=0, timeframe="now 7-d", geo="KR")
        # 메인 키워드 7일 스파크라인
        try:
            _iot = _pt.interest_over_time()
            if _iot is not None and not _iot.empty and main_keyword in _iot.columns:
                _vals   = _iot[main_keyword].tolist()
                _sl_raw = _vals[-7:] if len(_vals) >= 7 else _vals
                _sl_max = max(_sl_raw) if _sl_raw else 1
                _sl_min = min(_sl_raw) if _sl_raw else 0
                _rng    = _sl_max - _sl_min or 1
                _trends_sparklines[main_keyword] = [
                    round((_v - _sl_min) / _rng * 100) for _v in _sl_raw
                ]
        except Exception:
            pass
        _rel        = _pt.related_queries()
        _rising_df  = _rel.get(main_keyword, {}).get("rising")
        if _rising_df is not None and not _rising_df.empty:
            _trends_available = True
            for _, _row in _rising_df.head(n_trends * 3).iterrows():  # 여유분
                _q = str(_row.get("query", "")).strip()
                _v = int(_row.get("value", 0))
                if not _q:
                    continue
                _key = _q.lower()
                if _key in seen_keys or len(_key) < 2:
                    continue
                seen_keys.add(_key)
                # 급상승 스파크라인 시뮬레이션
                _sp = [max(0, min(100, 20 + int(_v/5)*_i//6 + (_i*8)))
                       for _i in range(7)]
                bucket_trends.append(
                    _make_item(
                        _q,
                        "📈 급상승트렌드", "trends",
                        score=_v,
                        date="급상승",
                        trend_val=_v,
                        sparkline=_sp,
                    )
                )
    except ImportError:
        pass
    except Exception:
        pass

    # ──────────────────────────────────────────────────────────────
    # A) YouTube API — 인기영상 제목 (미사용 — 슬롯 0)
    # ──────────────────────────────────────────────────────────────
    if api_key and n_video > 0:
        try:
            _r1 = requests.get(
                "https://www.googleapis.com/youtube/v3/search",
                params={
                    "key": api_key, "q": main_keyword,
                    "part": "id",   "type": "video",
                    "maxResults": 50,
                    "order": "relevance",
                    "regionCode": "KR",
                    "relevanceLanguage": "ko",
                    "publishedAfter": (
                        __import__('datetime').datetime.utcnow()
                        - __import__('datetime').timedelta(days=730)
                    ).strftime("%Y-%m-%dT%H:%M:%SZ"),
                },
                timeout=10
            )
            _d1 = _r1.json()
            if "error" not in _d1:
                _vids = [i["id"]["videoId"] for i in _d1.get("items", [])]
                if _vids:
                    _r2 = requests.get(
                        "https://www.googleapis.com/youtube/v3/videos",
                        params={
                            "key": api_key,
                            "id": ",".join(_vids),
                            "part": "snippet,statistics",
                        },
                        timeout=10
                    )
                    _items = _r2.json().get("items", [])
                    _scored = []
                    for _item in _items:
                        _title   = _item["snippet"].get("title", "").strip()
                        _pub     = _item["snippet"].get("publishedAt", "")
                        _channel = _item["snippet"].get("channelTitle", "")
                        _views   = int(_item.get("statistics", {}).get("viewCount", 0))
                        if not _title or _views < 1000:
                            continue
                        _sc = int(_views * _recency_factor(_pub))
                        _scored.append((_sc, _views, _pub, _title, _channel))
                    _scored.sort(key=lambda x: x[0], reverse=True)
                    for _sc, _views, _pub, _title, _ch in _scored[:n_video * 5]:  # 여유분
                        _key = _title.strip().lower()
                        if _key in seen_keys or len(_key) < 2:
                            continue
                        seen_keys.add(_key)
                        bucket_video.append(
                            _make_item(
                                _title,
                                "🔥 인기영상", "video",
                                score=_sc,
                                views=_fmt_views(_views),
                                date=_days_ago(_pub),
                                channel=_ch,
                                raw_views=_views,
                            )
                        )
        except Exception:
            pass

    # ──────────────────────────────────────────────────────────────
    # 소스별 정렬 + 슬롯 선택
    # ──────────────────────────────────────────────────────────────
    bucket_suggest.sort(key=lambda x: x["score"], reverse=True)
    bucket_trends.sort(key=lambda x:  x["score"], reverse=True)
    bucket_video.sort(key=lambda x:   x["score"], reverse=True)

    selected_suggest = bucket_suggest[:n_suggest]
    selected_trends  = bucket_trends[:n_trends]
    selected_video   = bucket_video[:n_video]

    # 부족분 보완: trends 부족 → suggest로, video 부족 → suggest로
    shortage = (n_suggest - len(selected_suggest)) +                (n_trends  - len(selected_trends))  +                (n_video   - len(selected_video))
    if shortage > 0:
        # 남은 suggest 여유분으로 보완 (suggest가 가장 많음)
        _extra_src = [b for b in bucket_suggest if b not in selected_suggest]
        _extra_src += [b for b in bucket_video  if b not in selected_video]
        _extra_src += [b for b in bucket_trends if b not in selected_trends]
        _extra_used = set()
        for _b in _extra_src:
            if shortage <= 0:
                break
            _key = _b["topic"].strip().lower()
            if _key not in _extra_used:
                _extra_used.add(_key)
                # trends가 비면 trends 슬롯에 채움
                if len(selected_trends) < n_trends:
                    selected_trends.append(_b)
                elif len(selected_suggest) < n_suggest:
                    selected_suggest.append(_b)
                shortage -= 1

    # ──────────────────────────────────────────────────────────────
    # 소스 내 정규화 점수 (0~100) 부여 → 전체 합치기
    # ──────────────────────────────────────────────────────────────
    def _norm_scores(items):
        """소스 내 score를 0~100으로 정규화"""
        if not items:
            return []
        max_s = max(it["score"] for it in items) or 1
        min_s = min(it["score"] for it in items)
        rng   = max_s - min_s or 1
        result = []
        for it in items:
            it = dict(it)  # 복사
            it["norm_score"] = round((it["score"] - min_s) / rng * 100)
            result.append(it)
        return result

    final = (
        _norm_scores(selected_suggest) +
        _norm_scores(selected_trends)  +
        _norm_scores(selected_video)
    )

    if not final:
        return None, (
            f"'{main_keyword}' 관련 서브 주제를 찾을 수 없습니다. "
            "API 키를 확인하거나 잠시 후 다시 시도하세요."
        )

    # ──────────────────────────────────────────────────────────────
    # 소스 우선순위 표시용 정렬
    #   표시 순서: suggest(1위→최하위) → trends(1위→최하위) → video(1위→최하위)
    #   각 소스 내에서는 norm_score 내림차순
    # ──────────────────────────────────────────────────────────────
    _order = {"suggest": 0, "trends": 1, "video": 2}
    final.sort(key=lambda x: (_order.get(x["source"], 9), -x["norm_score"]))

    # UI 바 너비 계산을 위해 score 필드에 norm_score 복사
    for it in final:
        it["score"] = it["norm_score"]

    # ──────────────────────────────────────────────────────────────
    # D) 스파크라인 없는 항목 시뮬레이션 보완
    # ──────────────────────────────────────────────────────────────
    import random as _rnd
    for _r in final:
        if not _r.get("sparkline"):
            _src = _r["source"]
            _ns  = _r.get("norm_score", 50)
            if _src == "video":
                _rnd.seed(hash(_r["topic"]) % 9999)
                _base = 55 + min(40, int(_ns / 100 * 40))
                _r["sparkline"] = [
                    max(5, min(100, _base + _rnd.randint(-12, 12)))
                    for _ in range(7)
                ]
                _r["sparkline"][-1] = min(100, _r["sparkline"][-1] + 5)
            elif _src == "suggest":
                _rank = _r.get("sug_rank", 5)
                _r["sparkline"] = [
                    max(5, 30 + (7 - _rank) * 5 +
                        int(_i * (100 - 30 - (7 - _rank) * 5) / 6))
                    for _i in range(7)
                ]
            else:
                _r["sparkline"] = [
                    max(5, min(100, 40 + int(_i * _ns / 6 / 2)))
                    for _i in range(7)
                ]

    return final[:top_n], None


def fetch_video_details(api_key, video_ids):
    videos = []
    for i in range(0, len(video_ids), 50):
        batch = ",".join(video_ids[i:i+50])
        r = requests.get(
            "https://www.googleapis.com/youtube/v3/videos",
            params={
                "key": api_key, "id": batch,
                "part": "snippet,statistics,contentDetails"
            },
            timeout=10
        )
        for item in r.json().get("items", []):
            sn = item.get("snippet", {})
            st = item.get("statistics", {})
            cd = item.get("contentDetails", {})
            vc = int(st.get("viewCount",   "0") if st.get("viewCount",   "0").isdigit() else 0)
            lc = int(st.get("likeCount",   "0") if st.get("likeCount",   "0").isdigit() else 0)
            cc = int(st.get("commentCount","0") if st.get("commentCount","0").isdigit() else 0)
            videos.append({
                "videoId":         item["id"],
                "channelId":       sn.get("channelId", ""),
                "title":           sn.get("title", ""),
                "channelTitle":    sn.get("channelTitle", ""),
                "description":     sn.get("description", ""),
                "publishedAt":     sn.get("publishedAt", "")[:10],
                "tags":            sn.get("tags", []),
                "thumbnail":       (sn.get("thumbnails",{}).get("maxres") or
                                    sn.get("thumbnails",{}).get("high") or {}).get("url",""),
                "duration":        parse_duration(cd.get("duration","PT0S")),
                "viewCount":       vc, "viewLabel":    fmt(vc)+"회",
                "likeCount":       lc, "likeLabel":    fmt(lc),
                "commentCount":    cc, "commentLabel": fmt(cc),
                "url":             f"https://www.youtube.com/watch?v={item['id']}",
                "subscriberLabel": "비공개",
                "transcript":      "",
                "keywords":        [],
                "summary":         "",
            })
    return videos



def get_related_videos(api_key, topic, top_n=3):
    """서브주제 키워드로 유튜브 관련 영상 TOP N 검색 (앱 내 검색 / 구글시트 내보내기 공용)"""
    try:
        params = {
            "key": api_key,
            "q": topic,
            "part": "snippet",
            "type": "video",
            "order": "viewCount",
            "maxResults": top_n * 2,
            "regionCode": "KR",
            "relevanceLanguage": "ko",
        }
        r = requests.get(
            "https://www.googleapis.com/youtube/v3/search",
            params=params, timeout=10
        )
        items = r.json().get("items", [])
        video_ids = [it["id"].get("videoId","") for it in items if it.get("id",{}).get("videoId")]
        if not video_ids:
            return []
        # 상세 정보 (조회수 포함)
        det_r = requests.get(
            "https://www.googleapis.com/youtube/v3/videos",
            params={"key": api_key, "id": ",".join(video_ids[:top_n*2]),
                    "part": "snippet,statistics"},
            timeout=10
        )
        videos = []
        for item in det_r.json().get("items", []):
            sn  = item.get("snippet", {})
            st2 = item.get("statistics", {})
            vc  = int(st2.get("viewCount","0")) if str(st2.get("viewCount","0")).isdigit() else 0
            vid = item["id"]
            videos.append({
                "videoId":   vid,
                "title":     sn.get("title",""),
                "channel":   sn.get("channelTitle",""),
                "views":     vc,
                "views_fmt": fmt(vc) + "회",
                "date":      sn.get("publishedAt","")[:10],
                "thumbnail": (sn.get("thumbnails",{}).get("high") or
                              sn.get("thumbnails",{}).get("medium") or {}).get("url",""),
                "url":       f"https://www.youtube.com/watch?v={vid}",
            })
        videos.sort(key=lambda x: x["views"], reverse=True)
        return videos[:top_n]
    except Exception:
        return []


def fetch_subscribers(api_key, videos):
    cache = {}
    ch_ids = list(set(v["channelId"] for v in videos if v["channelId"]))
    for i in range(0, len(ch_ids), 50):
        batch = ",".join(ch_ids[i:i+50])
        r = requests.get(
            "https://www.googleapis.com/youtube/v3/channels",
            params={"key": api_key, "id": batch, "part": "statistics"},
            timeout=10
        )
        for item in r.json().get("items", []):
            sub = item.get("statistics", {}).get("subscriberCount", "0")
            cache[item["id"]] = fmt(int(sub)) + "명" if sub.isdigit() else "비공개"
    for v in videos:
        v["subscriberLabel"] = cache.get(v["channelId"], "비공개")
    return videos

def get_transcript(video_id):
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
        for lang in ['ko', 'en']:
            try:
                segs = YouTubeTranscriptApi.get_transcript(video_id, languages=[lang])
                return " ".join(
                    s.get('text','') if isinstance(s, dict) else s.text
                    for s in segs
                )
            except: pass
        try:
            tlist = YouTubeTranscriptApi.list_transcripts(video_id)
            t = None
            for lang in ['ko','en']:
                try: t = tlist.find_transcript([lang]); break
                except: pass
            if t is None:
                for x in tlist: t = x; break
            if t:
                texts = []
                for seg in t.fetch():
                    if isinstance(seg, dict):   texts.append(seg.get('text',''))
                    elif hasattr(seg, 'text'):  texts.append(seg.text)
                    else:                       texts.append(str(seg))
                return " ".join(texts)
        except: pass
        return "자막 없음"
    except ImportError:
        return "youtube-transcript-api 미설치"
    except Exception as e:
        return f"자막 없음 ({str(e)[:50]})"


# ================================================================
# Whisper STT : 오디오 다운로드 → OpenAI Whisper API 변환
# ================================================================
def whisper_transcribe(video_id: str, openai_api_key: str) -> str:
    """
    yt-dlp 로 오디오를 직접 다운로드한 뒤 OpenAI Whisper API 로 텍스트 변환.
    yt-dlp 직접 다운로드 방식 → YouTube 403 차단 우회.
    """
    import os, tempfile, subprocess, sys

    if not openai_api_key:
        return "[Whisper 오류] OpenAI API 키가 없습니다."

    # yt-dlp 설치 확인
    try:
        import yt_dlp
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "yt-dlp", "-q"])
            import yt_dlp
        except Exception as e:
            return f"[Whisper 오류] yt-dlp 설치 실패: {e}"

    # openai 설치 확인
    try:
        import openai
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openai", "-q"])
            import openai
        except Exception as e:
            return f"[Whisper 오류] openai 설치 실패: {e}"

    url = f"https://www.youtube.com/watch?v={video_id}"
    tmp_dir = tempfile.mkdtemp()

    # Whisper API 지원 확장자
    SUPPORTED_EXTS = ('.m4a', '.webm', '.mp4', '.mp3', '.mpeg',
                      '.mpga', '.wav', '.ogg', '.opus')
    MAX_BYTES = 25 * 1024 * 1024  # 25MB

    try:
        # ── Step 1: ffmpeg 존재 여부 확인 ──
        import shutil as _shutil
        has_ffmpeg = _shutil.which("ffmpeg") is not None

        # ── Step 2: yt-dlp 직접 다운로드 (403 우회 핵심) ──
        out_tmpl = os.path.join(tmp_dir, "audio.%(ext)s")

        if has_ffmpeg:
            # ffmpeg 있으면 mp3로 변환 (가장 호환성 좋음)
            ydl_opts = {
                "format": "bestaudio/best",
                "outtmpl": out_tmpl,
                "quiet": True,
                "no_warnings": True,
                "noplaylist": True,
                "postprocessors": [{
                    "key": "FFmpegExtractAudio",
                    "preferredcodec": "mp3",
                    "preferredquality": "64",
                }],
            }
        else:
            # ffmpeg 없으면 m4a/webm 직접 다운로드
            ydl_opts = {
                "format": "bestaudio[ext=m4a]/bestaudio[ext=webm]/bestaudio[ext=opus]/bestaudio/best",
                "outtmpl": out_tmpl,
                "quiet": True,
                "no_warnings": True,
                "noplaylist": True,
            }

        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])

        # ── Step 3: 다운로드된 파일 찾기 ──
        all_files = [
            os.path.join(tmp_dir, f)
            for f in os.listdir(tmp_dir)
            if os.path.isfile(os.path.join(tmp_dir, f))
        ]
        # 지원 확장자 우선
        audio_files = [f for f in all_files if f.lower().endswith(SUPPORTED_EXTS)]
        if not audio_files:
            audio_files = all_files  # fallback: 모든 파일

        if not audio_files:
            return "[Whisper 오류] 다운로드된 오디오 파일 없음"

        audio_path = max(audio_files, key=os.path.getsize)
        file_size  = os.path.getsize(audio_path)

        if file_size < 1000:
            return "[Whisper 오류] 오디오 파일이 너무 작음 (1KB 미만)"
        if file_size > MAX_BYTES:
            return "[Whisper 오류] 파일 크기 초과 (25MB). 25분 이하 영상만 지원합니다."

        # ── Step 4: OpenAI Whisper API 호출 ──
        openai.api_key = openai_api_key
        with open(audio_path, "rb") as f:
            response = openai.audio.transcriptions.create(
                model="whisper-1",
                file=f,
                response_format="text",
            )

        result_text = response if isinstance(response, str) else getattr(response, 'text', str(response))
        return result_text if result_text else "[Whisper] 변환 결과 없음"

    except yt_dlp.utils.DownloadError as e:
        err = str(e)
        if "Private video" in err or "members-only" in err:
            return "[Whisper 오류] 비공개/멤버십 영상은 다운로드 불가"
        if "Sign in" in err or "bot" in err.lower():
            return "[Whisper 오류] YouTube 봇 차단. 잠시 후 다시 시도하세요."
        return f"[Whisper 오류] 다운로드 실패: {err[:120]}"
    except Exception as e:
        return f"[Whisper 오류] {str(e)[:120]}"
    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 🤖 Gemini 영상 분석 함수
# ================================================================
def gemini_analyze_video(video_id: str, gemini_api_key: str) -> str:
    """
    Gemini API로 YouTube 영상 분석
    - REST API 직접 호출 우선 (SDK 버전 무관)
    - 새 SDK / 구 SDK 순으로 폴백
    - 자막 없는 영상도 분석 가능
    - Streamlit Cloud 정상 작동
    """
    if not gemini_api_key:
        return "[Gemini 오류] API 키가 없습니다."

    video_url = f"https://www.youtube.com/watch?v={video_id}"
    prompt = """이 YouTube 영상의 내용을 상세히 분석해주세요.

다음 형식으로 응답해주세요:

[요약]
(3~5문장으로 핵심 내용 요약)

[주요 내용]
(- 항목별로 영상에서 다루는 핵심 내용 5~10개)

[대본]
아래 규칙을 반드시 지켜서 작성하세요:
- 타임스탬프(00:00, 01:23 등 시간 표시)를 절대 포함하지 마세요
- 장면 구분 없이 말한 내용을 자연스러운 흐름의 연속 문장으로 작성하세요
- 영상에서 실제로 한 말을 최대한 그대로, 빠짐없이 한국어로 작성하세요
- 문단은 주제가 바뀔 때만 나누세요
"""
    # v1beta 모델 목록 (2026년 3월 기준 — YouTube URL 직접 분석은 v1beta 전용)
    # gemini-2.0-flash/lite는 2026년부터 Deprecated → gemini-2.5-flash 우선
    _rest_models = [
        ("v1beta", "gemini-2.5-flash"),          # 현재 stable 최신
        ("v1beta", "gemini-2.5-flash-lite"),     # 경량 최신
        ("v1beta", "gemini-2.0-flash"),          # 이전 세대 (deprecated)
        ("v1beta", "gemini-1.5-flash"),          # 구버전 폴백
        ("v1beta", "gemini-1.5-pro"),            # 구버전 폴백
    ]
    # SDK 폴백용 모델 목록
    _models = [
        "gemini-2.5-flash",
        "gemini-2.0-flash",
        "gemini-1.5-flash",
        "gemini-1.5-pro",
    ]
    last_err = ""
    _rest_errors = []   # 모든 REST 시도 결과 누적

    # ══════════════════════════════════════════════════════
    # 방법 1: REST API 직접 호출 (v1beta 전용, SDK 불필요)
    # YouTube URL file_data 분석은 v1beta에서만 지원됨
    # ══════════════════════════════════════════════════════
    try:
        import requests as _req

        # YouTube URL은 mime_type 불필요 (공식 문서 기준)
        # file_data를 텍스트보다 먼저 배치 (공식 문서 순서)
        _payload = {
            "contents": [{
                "parts": [
                    {"file_data": {"file_uri": video_url}},
                    {"text": prompt}
                ]
            }]
        }

        for _api_ver, _model in _rest_models:
            try:
                _url = (
                    f"https://generativelanguage.googleapis.com/{_api_ver}/models/"
                    f"{_model}:generateContent?key={gemini_api_key}"
                )
                _resp = _req.post(_url, json=_payload, timeout=120)

                if _resp.status_code == 200:
                    _data = _resp.json()
                    _text = (
                        _data.get("candidates", [{}])[0]
                            .get("content", {})
                            .get("parts", [{}])[0]
                            .get("text", "")
                    )
                    if _text.strip():
                        return f"[\U0001f916 Gemini 분석 ({_model})]\n{_text.strip()}"
                    last_err = f"[Gemini 오류] {_model}: 응답 비어있음 (비공개/지역제한 영상)"
                    continue

                elif _resp.status_code == 400:
                    try:
                        _err_body = _resp.json().get("error", {})
                        _detail   = _err_body.get("message", "")
                        _err_status = _err_body.get("status", "")
                    except Exception:
                        _detail = _resp.text[:120]
                        _err_status = ""
                    # API 키 무효 → 즉시 반환
                    if "API_KEY_INVALID" in _err_status or "not valid" in _detail.lower() or "api key" in _detail.lower():
                        return "[Gemini 오류] API 키 인증 실패 (REST 400). secrets.toml의 GEMINI_API_KEY를 확인하세요."
                    # 영상 미지원 → continue (break 제거, 다른 모델 시도)
                    if "not supported" in _detail.lower() or "file_data" in _detail.lower() or "unsupported" in _detail.lower():
                        _e = f"{_api_ver}/{_model}: [400 미지원] {_detail[:80]}"
                        _rest_errors.append(_e); last_err = f"[Gemini 오류] {_e}"
                        continue
                    _e = f"{_api_ver}/{_model}: [400] {_detail[:120]}"
                    _rest_errors.append(_e); last_err = f"[Gemini 오류-REST] {_e}"
                    continue

                elif _resp.status_code == 401 or _resp.status_code == 403:
                    return "[Gemini 오류] API 키 인증 실패. GEMINI_API_KEY를 확인하세요."

                elif _resp.status_code == 404:
                    try:
                        _err404 = _resp.json().get("error", {}).get("message", _resp.text[:100])
                    except Exception:
                        _err404 = _resp.text[:100]
                    _e = f"{_api_ver}/{_model}: [404] {_err404[:120]}"
                    _rest_errors.append(_e); last_err = f"[Gemini 오류-REST] {_e}"
                    continue

                elif _resp.status_code == 429:
                    return "[Gemini 오류] API 할당량 초과(429). 잠시 후 재시도하세요."

                else:
                    _e = f"{_api_ver}/{_model}: [HTTP {_resp.status_code}] {_resp.text[:80]}"
                    _rest_errors.append(_e); last_err = f"[Gemini 오류-REST] {_e}"
                    continue

            except Exception as _e:
                _emsg = f"{_api_ver}/{_model}: [예외] {str(_e)[:80]}"
                _rest_errors.append(_emsg); last_err = f"[Gemini 오류-REST] {_emsg}"
                continue

        # REST API 모든 모델 실패 → 누적 에러 전체 반환
        if _rest_errors:
            _err_summary = "\n".join([f"  • {e}" for e in _rest_errors])
            return f"[Gemini 오류-REST] v1beta 모든 모델 실패:\n{_err_summary}"
        if last_err:
            return last_err

    except ImportError:
        last_err = "[Gemini] requests 미설치 → SDK 방식 시도"

    # ══════════════════════════════════════════════════════
    # 방법 2: 새 SDK (google-genai) — 딕셔너리 방식 (버전 독립적)
    # types 객체 대신 plain dict 사용 → SDK 버전 무관하게 동작
    # ══════════════════════════════════════════════════════
    try:
        from google import genai as _genai_new

        client = _genai_new.Client(api_key=gemini_api_key)
        # YouTube URL 직접 분석 지원 모델 (최신순)
        _new_sdk_models = [
            "gemini-2.0-flash",
            "gemini-2.0-flash-lite",
            "gemini-1.5-flash",
            "gemini-1.5-pro",
        ]
        _new_sdk_last_err = None
        for _model_name in _new_sdk_models:
            try:
                # plain dict 방식: SDK 버전에 무관하게 동작
                response = client.models.generate_content(
                    model=_model_name,
                    contents=[
                        {
                            "role": "user",
                            "parts": [
                                {"file_data": {"file_uri": video_url}},
                                {"text": prompt}
                            ]
                        }
                    ]
                )
                result_text = response.text.strip() if response.text else ""
                if not result_text:
                    _new_sdk_last_err = f"[Gemini 오류] {_model_name}: 응답 비어있음"
                    last_err = _new_sdk_last_err
                    continue
                return f"[\U0001f916 Gemini 분석 ({_model_name})]\n{result_text}"
            except Exception as e:
                err = str(e)
                _new_sdk_last_err = f"[Gemini 오류-newSDK] {_model_name}: {err[:150]}"
                last_err = _new_sdk_last_err
                if any(k in err.lower() for k in ["not found", "404", "not support", "unknown model", "deprecated"]):
                    continue  # 다음 모델 시도
                if "API_KEY_INVALID" in err or ("invalid" in err.lower() and "key" in err.lower()):
                    return f"[Gemini 오류] API 키 인증 실패\n키를 확인하세요: https://aistudio.google.com/app/apikey\n상세: {err[:100]}"
                if "quota" in err.lower() or "429" in err or "RESOURCE_EXHAUSTED" in err:
                    return f"[Gemini 오류] API 할당량 초과\n잠시 후 재시도하거나 유료 플랜을 확인하세요.\n상세: {err[:80]}"
                if "private" in err.lower() or "unavailable" in err.lower():
                    return f"[Gemini 오류] 비공개 또는 접근 불가 영상"
                # 기타 오류: 다음 모델 시도
                continue
        # 새 SDK 모든 모델 실패 → 마지막 오류 반환 (구 SDK로 넘어가지 않음)
        if _new_sdk_last_err:
            return _new_sdk_last_err
    except ImportError:
        pass  # 새 SDK 미설치 → 구 SDK 시도

    # ══════════════════════════════════════════════════════
    # 방법 3: 구 SDK (google.generativeai) 폴백
    # ══════════════════════════════════════════════════════
    try:
        import google.generativeai as _genai_old
    except ImportError:
        return (
            "[Gemini 오류] SDK 미설치.\n"
            "터미널에서 실행: pip install google-genai"
        )

    for _model_name in _models:
        try:
            _genai_old.configure(api_key=gemini_api_key)
            model = _genai_old.GenerativeModel(_model_name)
            try:
                response = model.generate_content([
                    prompt,
                    _genai_old.protos.Part(
                        file_data=_genai_old.protos.FileData(
                            mime_type="video/mp4", file_uri=video_url
                        )
                    )
                ])
            except Exception:
                response = model.generate_content([
                    {"role": "user", "parts": [
                        {"text": prompt},
                        {"file_data": {"file_uri": video_url, "mime_type": "video/mp4"}}
                    ]}
                ])
            result_text = response.text.strip() if response.text else ""
            if not result_text:
                last_err = f"[Gemini 오류] {_model_name}: 응답 비어있음"
                continue
            return f"[🤖 Gemini 분석 ({_model_name})]\n{result_text}"
        except Exception as e:
            err = str(e)
            last_err = f"[Gemini 오류-oldSDK] {_model_name}: {err[:100]}"
            if any(k in err.lower() for k in ["not found", "404", "not support", "v1beta"]):
                continue
            if "API_KEY_INVALID" in err or ("invalid" in err.lower() and "key" in err.lower()):
                return f"[Gemini 오류] API 키 인증 실패\n상세: {err[:80]}"
            if "quota" in err.lower() or "429" in err:
                return f"[Gemini 오류] API 할당량 초과\n상세: {err[:80]}"
            continue

    return last_err or "[Gemini 오류] 사용 가능한 방법 없음. API 키를 확인하세요."


def get_transcript_with_whisper(video_id: str,
                                 openai_api_key: str = "") -> str:
    """
    1차: youtube-transcript-api 로 자막 시도
    2차: 자막 없으면 Whisper API 로 변환
    """
    result = get_transcript(video_id)
    # 자막을 가져오지 못한 경우 Whisper 시도
    if (not result or
        result.startswith("자막 없음") or
        result.startswith("youtube-transcript") or
        result.startswith("[Whisper")):
        if openai_api_key:
            whisper_result = whisper_transcribe(video_id, openai_api_key)
            if whisper_result and not whisper_result.startswith("[Whisper 오류]"):
                return "[🎙️ Whisper 변환]\n" + whisper_result
            else:
                return f"자막 없음 / {whisper_result}"
        else:
            return result
    return result

# ================================================================
# 엑셀 저장 (바이트 반환)
# ================================================================
def save_xlsx_bytes(all_results_by_keyword, channel_stats):
    if not HAS_XLSX:
        return None
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    RED   = "1B3A6B"
    DRED  = "0D2347"
    WHITE = "FFFFFF"
    LGRAY = "F5F5F5"
    DGRAY = "D0D0D0"

    def style_header(ws, headers, row=1):
        fill = PatternFill("solid", fgColor=RED)
        font = Font(bold=True, color=WHITE, size=10)
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(
                bottom=Side(style="thin", color=DRED),
                right=Side(style="thin", color=DRED)
            )

    def style_cell(ws, row, col, value, fill=None, bold=False, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(bold=bold, size=9)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        c.border = Border(
            bottom=Side(style="thin", color=DGRAY),
            right=Side(style="thin", color=DGRAY)
        )

    # ── 시트1: 영상 목록 ─────────────────────────────────────
    ws1 = wb.create_sheet("📋 영상 목록")
    headers1 = ["검색어","순위","제목","채널","구독자","조회수","좋아요","댓글",
                "재생시간","업로드일","태그","핵심키워드","요약","URL"]
    style_header(ws1, headers1)
    widths1 = [12,5,40,20,10,10,8,8,10,12,30,25,50,45]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 28

    row = 2
    for kw, videos in all_results_by_keyword.items():
        for rank_idx, v in enumerate(videos, 1):
            fill_c = LGRAY if row % 2 == 0 else None
            data = [
                kw, rank_idx, v["title"], v["channelTitle"],
                v["subscriberLabel"], v["viewLabel"], v["likeLabel"], v["commentLabel"],
                v["duration"], v["publishedAt"],
                " | ".join(v["tags"][:10]) if v["tags"] else "",
                " · ".join(v.get("keywords", [])[:8]),
                v.get("summary", ""),
                v["url"]
            ]
            for col_idx, val in enumerate(data, 1):
                style_cell(ws1, row, col_idx, val, fill=fill_c)
            ws1.row_dimensions[row].height = 18
            row += 1

    # ── 시트2: 채널 통계 ─────────────────────────────────────
    ws2 = wb.create_sheet("📊 채널 통계")
    headers2 = ["채널명","구독자","영상수","총조회수","평균조회수","총좋아요","평균좋아요","총댓글","대표영상"]
    style_header(ws2, headers2)
    widths2 = [25,10,7,12,12,12,10,10,40]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 28

    row = 2
    for cs in channel_stats:
        rep = cs["videos"][0]["url"] if cs["videos"] else ""
        data = [
            cs["channel"], cs["subscriber"], cs["videoCount"],
            fmt(cs["totalView"])+"회", fmt(cs["avgView"])+"회",
            fmt(cs["totalLike"]), fmt(cs["avgLike"]),
            fmt(cs["totalComment"]), rep
        ]
        fill_c = LGRAY if row % 2 == 0 else None
        for col_idx, val in enumerate(data, 1):
            style_cell(ws2, row, col_idx, val, fill=fill_c, align="center" if col_idx in [2,3,4,5,6,7,8] else "left")
        ws2.row_dimensions[row].height = 18
        row += 1

    # ── 시트3: 키워드 요약 ───────────────────────────────────
    ws3 = wb.create_sheet("🔑 키워드 요약")
    headers3 = ["검색어","영상수","평균조회수","최고조회수","공통키워드 TOP5"]
    style_header(ws3, headers3)
    for i, w in enumerate([18,7,12,12,50], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[1].height = 28

    row = 2
    for kw, videos in all_results_by_keyword.items():
        all_kws = []
        for v in videos:
            all_kws.extend(v.get("keywords", []))
        top5 = " · ".join([w for w, _ in Counter(all_kws).most_common(5)])
        avg_v = sum(v["viewCount"] for v in videos) // len(videos) if videos else 0
        max_v = max((v["viewCount"] for v in videos), default=0)
        data = [kw, len(videos), fmt(avg_v)+"회", fmt(max_v)+"회", top5]
        fill_c = LGRAY if row % 2 == 0 else None
        for col_idx, val in enumerate(data, 1):
            style_cell(ws3, row, col_idx, val, fill=fill_c)
        ws3.row_dimensions[row].height = 18
        row += 1

    # ── 시트4: 대본 전문 ─────────────────────────────────────
    ws4 = wb.create_sheet("📜 대본 전문")
    headers4 = ["제목","채널","URL","대본 전문"]
    style_header(ws4, headers4)
    for i, w in enumerate([40,20,45,120], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.row_dimensions[1].height = 28

    row = 2
    for videos in all_results_by_keyword.values():
        for v in videos:
            if v.get("transcript") and "자막 없음" not in v["transcript"] and len(v["transcript"]) > 20:
                data = [v["title"], v["channelTitle"], v["url"], v["transcript"]]
                fill_c = LGRAY if row % 2 == 0 else None
                for col_idx, val in enumerate(data, 1):
                    style_cell(ws4, row, col_idx, val, fill=fill_c)
                ws4.row_dimensions[row].height = 80
                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ================================================================
# 텍스트 빌드
# ================================================================
def build_txt(all_results_by_keyword, channel_stats, sort_label):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = ["="*68,
             f"  📊 YouTube 분석 결과  |  생성: {ts}",
             f"  정렬: {sort_label}",
             "="*68]
    for kw, videos in all_results_by_keyword.items():
        lines += [f"\n{'━'*68}",
                  f"  🔍 검색어: [{kw}]  ({len(videos)}개 영상)",
                  f"{'━'*68}"]
        for v in videos:
            lines += [
                f"\n{'═'*68}",
                f"  {v.get('badge','▶')}  #{v.get('rank',0)}위  |  {v['title']}",
                f"{'═'*68}",
                f"🔗 URL       : {v['url']}",
                f"📺 채널      : {v['channelTitle']}  (구독자 {v['subscriberLabel']})",
                f"⏱️  길이     : {v['duration']}",
                f"📅 업로드    : {v['publishedAt']}",
                "",
                f"📈 통계",
                f"   👁️  조회수 : {v['viewLabel']} ({v['viewCount']:,}회)",
                f"   👍 좋아요 : {v['likeLabel']} ({v['likeCount']:,}개)",
                f"   💬 댓글수 : {v['commentLabel']} ({v['commentCount']:,}개)",
                "",
                f"📝 영상 설명:",
                "─"*68,
                v['description'] if v['description'] else "(없음)",
                "",
                f"🏷️  태그 ({len(v['tags'])}개):",
                "  " + " | ".join(v['tags'][:20]) if v['tags'] else "  (없음)",
                "",
                f"🔑 핵심 키워드:",
                "  " + " · ".join(v.get('keywords', [])) if v.get('keywords') else "  (추출 불가)",
                "",
                f"📋 요약:",
                "  " + v.get('summary', '(없음)'),
                "",
                f"📜 대본 전문:",
                "─"*68,
                v['transcript'] if v['transcript'] else "자막 없음",
                f"\n{'─'*68}",
            ]
    lines += [f"\n{'━'*68}",
              f"  📊 채널별 통계  (총 {len(channel_stats)}개 채널)",
              f"{'━'*68}",
              f"  {'채널명':<22} {'구독자':>8} {'영상수':>6} {'총조회수':>10} {'평균조회수':>10} {'평균좋아요':>10}",
              "  " + "─"*65]
    for cs in channel_stats:
        lines.append(
            f"  {cs['channel']:<22} {cs['subscriber']:>8} "
            f"{cs['videoCount']:>6} {fmt(cs['totalView'])+' 회':>10} "
            f"{fmt(cs['avgView'])+' 회':>10} {fmt(cs['avgLike']):>10}"
        )
    lines += ["", "="*68]
    return "\n".join(lines)

# ================================================================
# JSON 빌드
# ================================================================
def build_json(all_results_by_keyword, channel_stats):
    output = {"keywords": {}, "channel_stats": [], "generated_at": datetime.now().isoformat()}
    for kw, videos in all_results_by_keyword.items():
        output["keywords"][kw] = [{
            k: v[k] for k in ["videoId","title","channelTitle","url","publishedAt",
                               "viewCount","likeCount","commentCount","duration",
                               "subscriberLabel","tags","keywords","summary","transcript"]
        } for v in videos]
    output["channel_stats"] = [{
        k: cs[k] for k in ["channel","subscriber","videoCount","totalView","avgView","totalLike","avgLike"]
    } for cs in channel_stats]
    return json.dumps(output, ensure_ascii=False, indent=2)

# ================================================================
# Google Sheets 업로드
# ================================================================

def export_subtopics_to_gsheet(topics, main_keyword, credentials_dict=None, existing_id=None, share_email=None, api_key=None):
    """
    서브 주제를 구글시트에 누적 추가.
    - 시트명: 서브주제_기록
    - 헤더 자동 추가 (최초 1회)
    - 자동 필터 설정 (날짜/키워드/출처 필터링 가능)
    - 열 너비 자동 조정
    """
    if not HAS_GSHEET:
        return False, "gspread 라이브러리 미설치 (pip install gspread google-auth)"
    if not credentials_dict:
        return False, "GCP credentials 없음"

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    try:
        try:
            gc = gspread.service_account_from_dict(credentials_dict, scopes=SCOPES)
        except AttributeError:
            creds = Credentials.from_service_account_info(credentials_dict, scopes=SCOPES)
            try:    gc = gspread.Client(auth=creds)
            except: gc = gspread.authorize(creds)
    except Exception as e:
        return False, f"Google 인증 실패: {e}"

    try:
        if existing_id:
            eid = existing_id.strip()
            if "spreadsheets/d/" in eid:
                eid = eid.split("spreadsheets/d/")[1].split("/")[0]
            sh = gc.open_by_key(eid)
        else:
            sh = gc.create(f"YouTube_서브주제_{datetime.now().strftime('%Y%m%d')}")
            if share_email:
                sh.share(share_email, perm_type='user', role='writer')
            else:
                sh.share('', perm_type='anyone', role='reader')
    except Exception as e:
        return False, f"스프레드시트 열기 실패: {e}"

    SHEET_NAME = "서브주제_기록"
    # ── 개선된 헤더 (날짜/키워드 필터 최적화) ──────────────────────
    HEADER = [
        "추출날짜",       # A: YYYY-MM-DD (날짜 필터용)
        "추출시간",       # B: HH:MM
        "메인키워드",     # C: 키워드 필터용
        "순위",           # D: 1-10
        "서브주제(제목)", # E: 주요 내용
        "조회수",         # F: 숫자 (정렬용)
        "출처유형",       # G: 🔥인기영상 / 🔴실시간 / 📈급상승
        "업로드날짜",     # H: 영상 업로드 날짜
        "채널명",         # I
        "스코어",         # J: 내부 정렬 점수
        "주간변화",       # K: ↑↓ 트렌드
    ]

    # 시트 가져오기 or 생성
    is_new_sheet = False
    try:
        ws = sh.worksheet(SHEET_NAME)
    except Exception:
        try:
            ws = sh.add_worksheet(title=SHEET_NAME, rows=5000, cols=len(HEADER))
            is_new_sheet = True
        except Exception as e:
            return False, f"시트 생성 실패: {e}"

    # 기존 데이터 확인
    try:
        existing_rows = ws.get_all_values()
    except Exception:
        existing_rows = []

    rows_to_append = []
    need_header = (not existing_rows) or (existing_rows[0] != HEADER)
    if need_header:
        rows_to_append.append(HEADER)

    now_dt   = datetime.now()
    now_date = now_dt.strftime("%Y-%m-%d")
    now_time = now_dt.strftime("%H:%M")

    for i, t in enumerate(topics):
        _src_map = {"video": "🔥인기영상", "suggest": "🔴실시간검색", "trends": "📈급상승"}
        _raw_v   = t.get("raw_views", 0)
        _views_str = (
            str(_raw_v) if _raw_v > 0
            else (f"+{t.get('trend_val',0)}%" if t.get("source") == "trends"
                  else (f"#{t.get('sug_rank',0)}위" if t.get("sug_rank",0) > 0 else ""))
        )
        _sp   = t.get("sparkline", [])
        _trend_str = ""
        if _sp and len(_sp) >= 2:
            _diff = _sp[-1] - _sp[0]
            _trend_str = f"▲{_diff}" if _diff > 5 else (f"▼{abs(_diff)}" if _diff < -5 else "━보합")
        rows_to_append.append([
            now_date,
            now_time,
            main_keyword,
            str(i + 1),
            t.get("topic", ""),
            _views_str,
            _src_map.get(t.get("source", ""), t.get("label", "")),
            t.get("date", ""),
            t.get("channel", ""),
            str(t.get("score", 0)),
            _trend_str,
        ])

    # 누적 append
    try:
        ws.append_rows(rows_to_append, value_input_option="USER_ENTERED")
    except Exception as e:
        return False, f"데이터 추가 실패: {e}"

    # ── 자동 필터 + 열 너비 설정 (Sheets API batchUpdate) ──────────
    try:
        total_data_rows = len(existing_rows) + len(rows_to_append)
        sid = ws.id
        requests_body = []

        # 1) 자동 필터 (전체 헤더 범위)
        requests_body.append({"setBasicFilter": {"filter": {
            "range": {
                "sheetId": sid,
                "startRowIndex": 0, "endRowIndex": total_data_rows,
                "startColumnIndex": 0, "endColumnIndex": len(HEADER)
            }
        }}})

        # 2) 열 너비 설정 (픽셀)
        col_widths = [110, 70, 110, 50, 280, 90, 110, 100, 140, 90, 80]
        for _ci, _cw in enumerate(col_widths[:len(HEADER)]):
            requests_body.append({"updateDimensionProperties": {
                "range": {"sheetId": sid, "dimension": "COLUMNS",
                            "startIndex": _ci, "endIndex": _ci + 1},
                "properties": {"pixelSize": _cw},
                "fields": "pixelSize"
            }})

        # 3) 헤더 행 굵게 + 배경색 (파란 계열)
        if need_header:
            requests_body.append({"repeatCell": {
                "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1,
                            "startColumnIndex": 0, "endColumnIndex": len(HEADER)},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.11, "green": 0.37, "blue": 0.64},
                    "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                                    "bold": True, "fontSize": 10},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE"
                }},
                "fields": "userEnteredFormat"
            }})
            # 헤더 행 고정 (freeze)
            requests_body.append({"updateSheetProperties": {
                "properties": {"sheetId": sid,
                                 "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount"
            }})

        sh.batch_update({"requests": requests_body})
    except Exception:
        pass  # 스타일 실패해도 데이터는 저장됨

    sheet_url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    total_data = max(0, len(existing_rows) - (1 if existing_rows else 0)) + len(topics)

    # ── 관련 영상 TOP3 시트 추가 (api_key 있을 때만) ────────────────────
    if api_key:
        REL_SHEET   = "관련영상_TOP3"
        REL_HEADER  = ["추출날짜", "메인키워드", "서브주제", "영상순위",
                       "영상제목", "채널명", "조회수", "업로드날짜", "유튜브URL"]
        try:
            try:    ws_rel = sh.worksheet(REL_SHEET)
            except: ws_rel = sh.add_worksheet(title=REL_SHEET, rows=3000, cols=len(REL_HEADER))
            rel_existing = ws_rel.get_all_values()
            rel_rows     = []
            if (not rel_existing) or (rel_existing[0] != REL_HEADER):
                rel_rows.append(REL_HEADER)
            for t in topics:
                _rvs = get_related_videos(api_key, t.get("topic",""), top_n=3)
                for _ri, _rv in enumerate(_rvs):
                    rel_rows.append([
                        now_date,
                        main_keyword,
                        t.get("topic",""),
                        str(_ri + 1),
                        _rv.get("title",""),
                        _rv.get("channel",""),
                        str(_rv.get("views",0)),
                        _rv.get("date",""),
                        _rv.get("url",""),
                    ])
            if rel_rows:
                ws_rel.append_rows(rel_rows, value_input_option="USER_ENTERED")
            # 헤더 스타일
            try:
                _rs_sid = ws_rel.id
                _rs_req = [{
                    "repeatCell": {
                        "range": {"sheetId": _rs_sid, "startRowIndex": 0, "endRowIndex": 1,
                                  "startColumnIndex": 0, "endColumnIndex": len(REL_HEADER)},
                        "cell": {"userEnteredFormat": {
                            "backgroundColor": {"red": 0.06, "green": 0.49, "blue": 0.31},
                            "textFormat": {"foregroundColor": {"red":1,"green":1,"blue":1},
                                           "bold": True, "fontSize": 10},
                            "horizontalAlignment": "CENTER"
                        }},
                        "fields": "userEnteredFormat"
                    }
                }, {
                    "updateSheetProperties": {
                        "properties": {"sheetId": _rs_sid,
                                       "gridProperties": {"frozenRowCount": 1}},
                        "fields": "gridProperties.frozenRowCount"
                    }
                }, {
                    "setBasicFilter": {"filter": {"range": {
                        "sheetId": _rs_sid,
                        "startRowIndex": 0,
                        "endRowIndex": len(rel_existing) + len(rel_rows),
                        "startColumnIndex": 0,
                        "endColumnIndex": len(REL_HEADER)
                    }}}
                }]
                _rel_col_w = [110, 120, 200, 55, 260, 130, 80, 100, 200]
                for _ci, _cw in enumerate(_rel_col_w[:len(REL_HEADER)]):
                    _rs_req.append({"updateDimensionProperties": {
                        "range": {"sheetId": _rs_sid, "dimension": "COLUMNS",
                                  "startIndex": _ci, "endIndex": _ci+1},
                        "properties": {"pixelSize": _cw},
                        "fields": "pixelSize"
                    }})
                sh.batch_update({"requests": _rs_req})
            except Exception:
                pass
        except Exception:
            pass  # 관련영상 시트 실패해도 메인 결과는 유지

    return True, f"{sheet_url}|||{total_data}"

def upload_to_gsheet(all_results_by_keyword, channel_stats, sort_label,
                     credentials_dict=None, spreadsheet_name=None,
                     share_email=None, existing_id=None):
    if not HAS_GSHEET:
        return False, "gspread/google-auth 라이브러리가 설치되지 않았습니다.\n`pip install gspread google-auth` 를 실행해주세요."
    if not credentials_dict:
        return False, "credentials.json 파일이 없습니다."

    # ✅ gspread v6 호환 스코프
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    try:
        # ✅ gspread v6 권장 방식: service_account_from_dict (가장 안정적)
        try:
            gc = gspread.service_account_from_dict(credentials_dict, scopes=SCOPES)
        except AttributeError:
            # gspread v5 이하 fallback
            creds = Credentials.from_service_account_info(credentials_dict, scopes=SCOPES)
            try:
                gc = gspread.Client(auth=creds)
            except Exception:
                gc = gspread.authorize(creds)
    except Exception as e:
        err_hint = str(e)
        return False, (
            f"Google 인증 실패: {err_hint}\n\n"
            "확인사항:\n"
            "1️⃣ credentials.json 이 서비스 계정(Service Account) 파일인지 확인\n"
            "2️⃣ Google Cloud Console → API 라이브러리에서 'Google Sheets API' 활성화 확인\n"
            "3️⃣ 'Google Drive API' 도 활성화 확인\n"
            "4️⃣ 서비스 계정의 이메일을 스프레드시트 편집자로 공유했는지 확인"
        )

    try:
        if existing_id:
            eid = existing_id.strip()
            if "spreadsheets/d/" in eid:
                eid = eid.split("spreadsheets/d/")[1].split("/")[0]
            sh = gc.open_by_key(eid)
            sheet_url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
        else:
            if not spreadsheet_name:
                slug = "_".join(list(all_results_by_keyword.keys())[:2])[:30]
                ts   = datetime.now().strftime("%m%d_%H%M")
                spreadsheet_name = f"YouTube분석_{slug}_{ts}"
            sh = gc.create(spreadsheet_name)
            sheet_url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
            if share_email:
                sh.share(share_email, perm_type='user', role='writer')
            else:
                sh.share('', perm_type='anyone', role='reader')
    except gspread.exceptions.APIError as e:
        msg = str(e)
        if "quota" in msg.lower() or "403" in msg:
            return False, (
                "❌ Google Drive 저장 용량 초과 (403 오류)\n\n"
                "해결 방법:\n"
                "1️⃣  drive.google.com 에서 불필요한 파일 삭제 후 재시도\n"
                "2️⃣  또는 기존 스프레드시트 ID를 아래 입력란에 넣어주세요"
            )
        return False, f"스프레드시트 생성 오류: {msg}"
    except Exception as e:
        return False, f"스프레드시트 생성 오류: {e}"

    # ── 시트 작성 헬퍼 ─────────────────────────────────────────
    def safe_write(ws, rows_data):
        """시트에 데이터 쓰기 — gspread v5/v6 완전 호환"""
        if not rows_data:
            return
        safe = [[str(c) if c is not None else "" for c in row]
                for row in rows_data]
        ws.clear()
        written = False
        # 방법 1: gspread v5 방식 (range_name, values)
        try:
            ws.update("A1", safe)
            written = True
        except Exception:
            pass
        # 방법 2: gspread v6 키워드 방식
        if not written:
            try:
                ws.update(range_name="A1", values=safe)
                written = True
            except Exception:
                pass
        # 방법 3: batch_update 방식
        if not written:
            try:
                ws.batch_update([{"range": "A1", "values": safe}])
                written = True
            except Exception:
                pass
        # 방법 4: 최후 수단 append
        if not written:
            for i in range(0, len(safe), 100):
                ws.append_rows(safe[i:i+100], value_input_option="RAW")

    def get_or_create_ws(title, rows=500, cols=20):
        """시트 생성/조회 (이모지 이름 실패 시 plain 이름 fallback)"""
        current_titles = [ws.title for ws in sh.worksheets()]
        def _try(t):
            if t in current_titles:
                ws = sh.worksheet(t)
                ws.clear()
                return ws
            return sh.add_worksheet(title=t, rows=rows, cols=cols)
        try:
            return _try(title)
        except Exception:
            import re as _re
            plain = _re.sub(r'[^\w\s가-힣]', '', title).strip() or f"Sheet{len(current_titles)}"
            return _try(plain)

    # ── 시트1: 영상 목록 ───────────────────────────────────────
    try:
        ws1   = get_or_create_ws("영상 목록", rows=1000, cols=15)
        h1    = ["검색어","순위","제목","채널","구독자","조회수","좋아요","댓글",
                 "재생시간","업로드일","태그","핵심키워드","요약","URL"]
        rows1 = [h1]
        for kw, videos in all_results_by_keyword.items():
            for rank_idx, v in enumerate(videos, 1):
                rows1.append([
                    kw, rank_idx, v["title"], v["channelTitle"],
                    v["subscriberLabel"], v["viewLabel"], v["likeLabel"], v["commentLabel"],
                    v["duration"], v["publishedAt"],
                    " | ".join(v["tags"][:10]) if v["tags"] else "",
                    " · ".join(v.get("keywords", [])[:8]),
                    v.get("summary", ""),
                    v["url"]
                ])
        safe_write(ws1, rows1)
    except Exception as e:
        return False, f"❌ 시트1(영상 목록) 쓰기 오류: {e}"

    # ── 시트2: 채널 통계 ───────────────────────────────────────
    try:
        ws2   = get_or_create_ws("채널 통계", rows=200, cols=10)
        h2    = ["채널명","구독자","영상수","총조회수","평균조회수","총좋아요","평균좋아요","총댓글","대표영상"]
        rows2 = [h2]
        for cs in channel_stats:
            rep = cs["videos"][0]["url"] if cs["videos"] else ""
            rows2.append([
                cs["channel"], cs["subscriber"], cs["videoCount"],
                fmt(cs["totalView"])+"회", fmt(cs["avgView"])+"회",
                fmt(cs["totalLike"]), fmt(cs["avgLike"]),
                fmt(cs["totalComment"]), rep
            ])
        safe_write(ws2, rows2)
    except Exception as e:
        return False, f"❌ 시트2(채널 통계) 쓰기 오류: {e}"

    # ── 시트3: 키워드 요약 ─────────────────────────────────────
    try:
        ws3   = get_or_create_ws("키워드 요약", rows=100, cols=6)
        h3    = ["검색어","영상수","평균조회수","최고조회수","공통키워드 TOP5"]
        rows3 = [h3]
        for kw, videos in all_results_by_keyword.items():
            all_kws = []
            for v in videos:
                all_kws.extend(v.get("keywords", []))
            top5  = " · ".join([w for w, _ in Counter(all_kws).most_common(5)])
            avg_v = sum(v["viewCount"] for v in videos) // len(videos) if videos else 0
            max_v = max((v["viewCount"] for v in videos), default=0)
            rows3.append([kw, len(videos), fmt(avg_v)+"회", fmt(max_v)+"회", top5])
        safe_write(ws3, rows3)
    except Exception as e:
        return False, f"❌ 시트3(키워드 요약) 쓰기 오류: {e}"

    # ── 시트4: 대본 전문 ───────────────────────────────────────
    try:
        ws4   = get_or_create_ws("대본 전문", rows=500, cols=5)
        h4    = ["제목","채널","URL","대본 출처","대본 전문"]
        rows4 = [h4]
        MAX_CELL = 49000
        for kw, videos in all_results_by_keyword.items():
            for v in videos:
                tr = v.get("transcript", "")
                if not is_valid_transcript(tr):
                    continue
                whisper_flag = "🎙️ Whisper" if tr.startswith("[🎙️") else "📝 자막"
                if len(tr) > MAX_CELL:
                    tr = tr[:MAX_CELL] + f"\n\n[⚠️ {MAX_CELL}자 초과로 잘림. 원본: {len(v['transcript'])}자]"
                rows4.append([v["title"], v["channelTitle"], v["url"], whisper_flag, tr])
        # 대본은 크므로 50행씩 나눠서 update
        ws4.clear()
        for i in range(0, len(rows4), 50):
            chunk = rows4[i:i+50]
            safe_rows = [[str(c) if c is not None else "" for c in row]
                         for row in chunk]
            start_cell = f"A{i + 1}"
            try:
                ws4.update(start_cell, safe_rows)
            except Exception:
                try:
                    ws4.update(range_name=start_cell, values=safe_rows)
                except Exception:
                    ws4.append_rows(safe_rows, value_input_option="RAW")
    except Exception as e:
        return False, f"❌ 시트4(대본 전문) 쓰기 오류: {e}"

    # ── 기본 Sheet1 제거 ───────────────────────────────────────
    for ws in sh.worksheets():
        if ws.title in ("Sheet1", "시트1"):
            try:
                sh.del_worksheet(ws)
            except Exception:
                pass

    return True, sheet_url

# ================================================================
# 메인 UI
# ================================================================
def main():
    # ── 헤더 ─────────────────────────────────────────────────
    st.markdown("""
    <div class="main-header">
        <h1>🎬 YouTube 분석 도구</h1>
        <p>YouTube 동영상을 검색·분석하고 엑셀 / 텍스트 / Google Sheets로 내보낼 수 있습니다</p>
    </div>
    """, unsafe_allow_html=True)

# ================================================================
    # secrets.toml 에서 설정값 한꺼번에 로드 (없으면 빈값)
    # ================================================================
    def _secret(key, default=""):
        try:
            v = st.secrets.get(key, None)
            if v is None or str(v).strip() == "" or str(v) == "None":
                return default
            return str(v).strip()
        except Exception:
            return default

    _s_api_key    = _secret("YOUTUBE_API_KEY")
    _s_keywords   = _secret("DEFAULT_KEYWORDS")
    _s_max_count  = _secret("DEFAULT_MAX_COUNT", "20")
    _s_sort       = _secret("DEFAULT_SORT", "조회수순")
    _s_email      = _secret("GSHEET_SHARE_EMAIL")
    _s_existing   = _secret("GSHEET_EXISTING_ID")
    _s_openai_key  = _secret("OPENAI_API_KEY")
    _s_gemini_key  = _secret("GEMINI_API_KEY")

    # ✅ Streamlit Cloud Secrets의 gcp_service_account 자동 로드
    _s_gcp_creds = None
    try:
        _gcp = st.secrets.get("gcp_service_account", {})
        if _gcp and _gcp.get("type") == "service_account":
            _s_gcp_creds = dict(_gcp)
    except Exception:
        _s_gcp_creds = None

    # ================================================================
    # 클립보드 복사 헬퍼 함수 (JS 기반)
    # ================================================================
    def copy_button(copy_text: str, btn_label: str, btn_key: str):
        """텍스트를 클립보드에 복사하는 버튼을 렌더링"""
        import base64
        encoded = base64.b64encode(copy_text.encode("utf-8")).decode("utf-8")
        html = f"""
<div style="margin-top:8px;">
  <button onclick="
    (function(){{
      var text = atob('{encoded}');
      navigator.clipboard.writeText(text).then(function(){{
        var btn = document.getElementById('cpbtn_{btn_key}');
        var orig = btn.innerText;
        btn.innerText = '✅ 복사됨!';
        btn.style.background = '#4CAF50';
        setTimeout(function(){{ btn.innerText = orig; btn.style.background = '#1a73e8'; }}, 2000);
      }}).catch(function(){{
        var ta = document.createElement('textarea');
        ta.value = text;
        document.body.appendChild(ta);
        ta.select();
        document.execCommand('copy');
        document.body.removeChild(ta);
        var btn = document.getElementById('cpbtn_{btn_key}');
        btn.innerText = '✅ 복사됨!';
        btn.style.background = '#4CAF50';
        setTimeout(function(){{ btn.innerText = '{btn_label}'; btn.style.background = '#1a73e8'; }}, 2000);
      }});
    }})();
  "
  id="cpbtn_{btn_key}"
  style="background:#1a73e8;color:white;border:none;border-radius:8px;
         padding:7px 18px;cursor:pointer;font-size:0.88rem;font-weight:600;
         width:100%;transition:background 0.2s;">
    {btn_label}
  </button>
</div>
"""
        st.markdown(html, unsafe_allow_html=True)

    # ================================================================
    # ★ 사이드바 강제 열기 (최초 세션 시작 시)
    #   브라우저 localStorage에 닫힌 상태가 저장돼 있어도 항상 열린 상태로 시작
    # ================================================================
    if "sidebar_initialized" not in st.session_state:
        st.session_state["sidebar_initialized"] = True
        import streamlit.components.v1 as _stc
        _stc.html("""
<script>
(function() {
    // 1단계: localStorage에서 사이드바 닫힘 기록 초기화
    //        Streamlit은 stSidebarCollapsed-{id} 키로 상태 저장
    try {
        var lsKeys = Object.keys(window.parent.localStorage);
        lsKeys.forEach(function(k) {
            if (k.indexOf("stSidebarCollapsed") !== -1) {
                window.parent.localStorage.setItem(k, "false");
            }
        });
    } catch(e) {}

    // 2단계: 버튼 직접 클릭 (최대 25회 재시도)
    var _tries = 0;
    var _iv = setInterval(function() {
        _tries++;

        // stExpandSidebarButton : Streamlit 1.55+ 실제 >> 버튼 testid
        var expandBtn = window.parent.document.querySelector(
            '[data-testid="stExpandSidebarButton"]'
        );
        if (expandBtn) {
            expandBtn.click();
            clearInterval(_iv);
            return;
        }

        // 대체: stSidebarCollapsedControl 내 버튼
        var collapsedCtrl = window.parent.document.querySelector(
            '[data-testid="stSidebarCollapsedControl"] button'
        );
        if (collapsedCtrl) {
            collapsedCtrl.click();
            clearInterval(_iv);
            return;
        }

        if (_tries >= 25) { clearInterval(_iv); }
    }, 100);
})();
</script>
""", height=0, width=0)

    # ================================================================
    # 사이드바
    # ================================================================
    with st.sidebar:
        st.markdown(
            "<div style='background:linear-gradient(135deg,#4f7ef8,#3a5bd4);"
            "border-radius:10px;padding:10px 14px;margin:0 0 10px 0;"
            "box-shadow:0 3px 10px rgba(79,126,248,0.20)'>"
            "<span style='font-size:1.05rem;font-weight:800;color:#fff;"
            "letter-spacing:-0.01em'>🎬 YouTube 분석 도구</span><br>"
            "<span style='font-size:0.68rem;color:rgba(255,255,255,0.80);font-weight:400'>"
            "검색 · 분석 · 내보내기</span></div>",
            unsafe_allow_html=True
        )

        # ── 다크모드 토글 버튼 ──────────────────────────────
        _dm_on = st.session_state.get("dark_mode", False)
        _dm_c1, _dm_c2 = st.columns([3, 2])
        with _dm_c1:
            st.markdown(
                "<span style='font-size:0.78rem;font-weight:600;color:" + ("#a0a8cc" if _dm_on else "#5c6480") + "'>"
                + ("🌙 다크모드 ON" if _dm_on else "☀️ 라이트모드") + "</span>",
                unsafe_allow_html=True
            )
        with _dm_c2:
            if st.button(
                "🌙 켜기" if not _dm_on else "☀️ 끄기",
                key="btn_dark_mode_toggle",
                use_container_width=True,
                help="다크모드 전환"
            ):
                st.session_state["dark_mode"] = not _dm_on
                st.rerun()

        # API 키: secrets.toml 에서 자동 로드
        api_key = st.text_input(
            "🔑 YouTube API 키",
            value=_s_api_key,
            type="password",
            placeholder="AIzaSy...",
            help="Google Cloud Console에서 발급한 YouTube Data API v3 키를 입력하세요."
        )
        if _s_api_key:
            st.markdown(
                "<div style='background:#ecfdf5;border:1px solid #86efac;"
                "border-radius:6px;padding:3px 8px;margin-top:2px'>"
                "<span style='font-size:0.66rem;color:#15803d;font-weight:600'>"
                "✅ secrets.toml 자동 로드</span></div>",
                unsafe_allow_html=True
            )

        st.markdown('<hr style="margin:6px 0">', unsafe_allow_html=True)

        # ── 최근 검색 기록 섹션 ─────────────────────────────────────
        _hist_now = st.session_state.get("search_history", [])
        if _hist_now:
            _dm_hist = st.session_state.get("dark_mode", False)
            _hist_bg      = "#1e2038" if _dm_hist else "#f8f9ff"
            _hist_border  = "#2e3157" if _dm_hist else "#e4e8ff"
            _hist_title_c = "#7c9ef8" if _dm_hist else "#4f7ef8"
            _hist_chip_bg = "#252640" if _dm_hist else "#eef2ff"
            _hist_chip_c  = "#a0aec0" if _dm_hist else "#5c72b0"
            _hist_chip_ho = "#2e3157" if _dm_hist else "#dde6ff"
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:6px;margin:6px 0 4px 0'>"
                f"<div style='width:3px;height:14px;background:linear-gradient(180deg,#f59e0b,#f97316);"
                f"border-radius:2px'></div>"
                f"<span style='font-size:0.79rem;font-weight:700;color:{_hist_title_c};"
                f"letter-spacing:0.03em'>🕐 최근 검색 기록</span>"
                f"</div>",
                unsafe_allow_html=True
            )
            for _hi, _hkw in enumerate(_hist_now):
                _hist_short = _hkw[:22] + ("…" if len(_hkw) > 22 else "")
                _hc1, _hc2 = st.columns([5, 1])
                with _hc1:
                    if st.button(
                        f"🕐 {_hist_short}",
                        key=f"hist_btn_{_hi}",
                        use_container_width=True,
                        help=f"'{_hkw}' 로 재검색"
                    ):
                        st.session_state["kw_input"] = _hkw
                        st.rerun()
                with _hc2:
                    if st.button("✕", key=f"hist_del_{_hi}", help="기록 삭제"):
                        _h2 = st.session_state.get("search_history", [])
                        _h2 = [h for h in _h2 if h != _hkw]
                        st.session_state["search_history"] = _h2
                        st.rerun()
            if st.button("🗑 기록 전체 삭제", key="btn_hist_clear", use_container_width=True):
                st.session_state["search_history"] = []
                st.rerun()
            st.markdown('<hr style="margin:6px 0">', unsafe_allow_html=True)

        st.markdown(
            "<div style='display:flex;align-items:center;gap:6px;margin:8px 0 4px 0'>"
            "<div style='width:3px;height:14px;background:linear-gradient(180deg,#4f7ef8,#a78bfa);"
            "border-radius:2px'></div>"
            "<span style='font-size:0.73rem;font-weight:700;color:#5c6480;"
            "text-transform:uppercase;letter-spacing:0.05em'>🔍 검색 옵션</span></div>",
            unsafe_allow_html=True
        )

        # ── session_state 초기화 (키워드 입력창 + 서브 주제) ──────────────
        if "hot_topics" not in st.session_state:
            st.session_state["hot_topics"]       = []
        if "hot_topics_kw" not in st.session_state:
            st.session_state["hot_topics_kw"]    = ""
        if "hot_topic_clicked" not in st.session_state:
            st.session_state["hot_topic_clicked"] = ""
        if "hot_exp_open" not in st.session_state:
            st.session_state["hot_exp_open"] = True
        if "favorites" not in st.session_state:
            st.session_state["favorites"] = []  # [{topic, keyword, date, views, label}]
        if "fav_exp_open" not in st.session_state:
            st.session_state["fav_exp_open"] = False
        if "fav_action" not in st.session_state:
            st.session_state["fav_action"] = ""
        if "subtopic_export_trigger" not in st.session_state:
            st.session_state["subtopic_export_trigger"] = False
        if "detail_chart_topic" not in st.session_state:
            st.session_state["detail_chart_topic"] = ""
        if "detail_chart_data" not in st.session_state:
            st.session_state["detail_chart_data"] = {}
        if "inline_search_topic" not in st.session_state:
            st.session_state["inline_search_topic"] = ""
        if "inline_search_results" not in st.session_state:
            st.session_state["inline_search_results"] = []
        if "fav_dashboard_open" not in st.session_state:
            st.session_state["fav_dashboard_open"] = False
        # 키워드 입력창 초기값 설정 (최초 1회만)
        if "kw_input" not in st.session_state:
            st.session_state["kw_input"] = _s_keywords or ""
        if "dark_mode" not in st.session_state:
            st.session_state["dark_mode"] = False
        if "fav_order" not in st.session_state:
            st.session_state["fav_order"] = []
        if "search_history" not in st.session_state:
            st.session_state["search_history"] = []  # 최근 검색어 리스트 (최대 5개)
        # 폴더 관련 session_state
        if "fav_folders" not in st.session_state:
            # {"기본 폴더": ["topic1", "topic2", ...], "폴더2": [...]}
            st.session_state["fav_folders"] = {"기본 폴더": []}
        if "fav_folder_open" not in st.session_state:
            # 각 폴더의 펼침 상태: {폴더명: bool}
            st.session_state["fav_folder_open"] = {}
        if "fav_folder_action" not in st.session_state:
            st.session_state["fav_folder_action"] = ""   # "NEW_FOLDER","RENAME:old:new","DEL_FOLDER:name","MOVE:topic:folder"
        if "fav_new_folder_mode" not in st.session_state:
            st.session_state["fav_new_folder_mode"] = False
        if "fav_rename_folder" not in st.session_state:
            st.session_state["fav_rename_folder"] = ""   # 현재 이름 변경 중인 폴더명
        if "fav_add_target_folder" not in st.session_state:
            st.session_state["fav_add_target_folder"] = "기본 폴더"  # ADD 시 저장할 폴더

        # ★ 즐겨찾기 액션 처리
        _fav_action = st.session_state.pop("fav_action", "")
        if _fav_action.startswith("ADD:"):
            _fav_topic_key = _fav_action[4:]
            _all_t = st.session_state.get("hot_topics", [])
            _fav_found = next((t for t in _all_t if t["topic"] == _fav_topic_key), None)
            if _fav_found:
                _favs = st.session_state["favorites"]
                if not any(f["topic"] == _fav_topic_key for f in _favs):
                    # 저장할 폴더 결정
                    _target_folder = st.session_state.get("fav_add_target_folder", "기본 폴더")
                    # fav_folders에 폴더 없으면 생성
                    _ff = st.session_state.setdefault("fav_folders", {"기본 폴더": []})
                    if _target_folder not in _ff:
                        _ff[_target_folder] = []
                        st.session_state["fav_folders"] = _ff
                    # favorites 리스트에 추가 (folder 필드 포함)
                    _favs.append({
                        "topic":     _fav_found["topic"],
                        "keyword":   st.session_state.get("hot_topics_kw", ""),
                        "saved":     datetime.now().strftime("%m/%d %H:%M"),
                        "saved_dt":  datetime.now().isoformat(),
                        "views":     _fav_found.get("views", ""),
                        "label":     _fav_found.get("label", ""),
                        "score":     _fav_found.get("score", 0),
                        "raw_views": _fav_found.get("raw_views", 0),
                        "sparkline": _fav_found.get("sparkline", []),
                        "folder":    _target_folder,
                    })
                    st.session_state["favorites"] = _favs
                    # fav_folders에도 topic 등록
                    _ff2 = st.session_state.get("fav_folders", {"기본 폴더": []})
                    if _target_folder not in _ff2:
                        _ff2[_target_folder] = []
                    if _fav_topic_key not in _ff2[_target_folder]:
                        _ff2[_target_folder].append(_fav_topic_key)
                    st.session_state["fav_folders"] = _ff2
        elif _fav_action.startswith("DEL:"):
            _del_topic = _fav_action[4:]
            # favorites 리스트에서 제거
            st.session_state["favorites"] = [
                f for f in st.session_state["favorites"] if f["topic"] != _del_topic
            ]
            # fav_folders에서도 제거
            _ff3 = st.session_state.get("fav_folders", {})
            for _fn in _ff3:
                _ff3[_fn] = [t for t in _ff3[_fn] if t != _del_topic]
            st.session_state["fav_folders"] = _ff3
        elif _fav_action.startswith("MOVE:"):
            # MOVE:topic:새폴더
            _mv_parts = _fav_action[5:].split(":", 1)
            if len(_mv_parts) == 2:
                _mv_topic, _mv_dst = _mv_parts
                # favorites에서 folder 필드 업데이트
                _ff4 = st.session_state.get("fav_folders", {"기본 폴더": []})
                if _mv_dst not in _ff4:
                    _ff4[_mv_dst] = []
                for _fitem in st.session_state.get("favorites", []):
                    if _fitem["topic"] == _mv_topic:
                        _old_folder = _fitem.get("folder", "기본 폴더")
                        _fitem["folder"] = _mv_dst
                        # 이전 폴더에서 제거, 새 폴더에 추가
                        if _old_folder in _ff4 and _mv_topic in _ff4[_old_folder]:
                            _ff4[_old_folder].remove(_mv_topic)
                        if _mv_topic not in _ff4[_mv_dst]:
                            _ff4[_mv_dst].append(_mv_topic)
                st.session_state["fav_folders"] = _ff4
        elif _fav_action.startswith("NEW_FOLDER:"):
            _nf_name = _fav_action[11:].strip()
            if _nf_name:
                _ff5 = st.session_state.get("fav_folders", {"기본 폴더": []})
                if _nf_name not in _ff5:
                    _ff5[_nf_name] = []
                st.session_state["fav_folders"] = _ff5
        elif _fav_action.startswith("RENAME_FOLDER:"):
            _rf_parts = _fav_action[14:].split(":", 1)
            if len(_rf_parts) == 2:
                _rf_old, _rf_new = _rf_parts
                _rf_new = _rf_new.strip()
                if _rf_new and _rf_old != _rf_new:
                    _ff6 = st.session_state.get("fav_folders", {})
                    # 새 폴더명으로 내용 이전
                    _ff6[_rf_new] = _ff6.pop(_rf_old, [])
                    st.session_state["fav_folders"] = _ff6
                    # favorites의 folder 필드 업데이트
                    for _fitem in st.session_state.get("favorites", []):
                        if _fitem.get("folder") == _rf_old:
                            _fitem["folder"] = _rf_new
        elif _fav_action.startswith("DEL_FOLDER:"):
            _df_name = _fav_action[11:]
            if _df_name != "기본 폴더":
                _ff7 = st.session_state.get("fav_folders", {})
                _orphans = _ff7.pop(_df_name, [])
                # 이 폴더 항목들을 기본 폴더로 이동
                _ff7.setdefault("기본 폴더", []).extend(_orphans)
                st.session_state["fav_folders"] = _ff7
                # favorites의 folder 필드도 업데이트
                for _fitem in st.session_state.get("favorites", []):
                    if _fitem.get("folder") == _df_name:
                        _fitem["folder"] = "기본 폴더"

        # ★ 서브 주제 클릭 처리 → text_area 렌더링 전에 실행해야 반영됨
        _clicked = st.session_state.pop("hot_topic_clicked", "")
        if _clicked:
            _cur_kws = [k.strip() for k in
                        st.session_state["kw_input"].replace("，", ",").split(",")
                        if k.strip()]
            if _clicked not in _cur_kws:
                # 클릭한 서브 주제를 앞에 삽입
                _cur_kws.insert(0, _clicked)
            st.session_state["kw_input"] = ", ".join(_cur_kws)

        # ── 키워드 입력창 (key="kw_input" → session_state 기반) ──────────
        keywords_input = st.text_area(
            "검색 키워드 (쉼표로 여러 개 입력)",
            key="kw_input",
            placeholder="예: 비타민D 효능, 50대 영양제",
            height=65
        )

        # ── 🔥 실시간 인기 서브 주제 추출 UI ─────────────────────────────
        _main_kw = keywords_input.split(",")[0].strip() if keywords_input else ""

        # ── 서브주제 추출 헤더 + 버튼 ──────────────────────────
        _col_hot1, _col_hot2 = st.columns([3, 1])
        with _col_hot1:
            if _main_kw:
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:5px;margin:4px 0 2px 0'>"
                    f"<div style='width:3px;height:13px;background:linear-gradient(180deg,#ff6b35,#f7931e);"
                    f"border-radius:2px'></div>"
                    f"<span style='font-size:0.78rem;font-weight:700;color:#e55a2b'>🔥 인기 서브 주제</span>"
                    f"<span style='font-size:0.68rem;color:#aaa;font-weight:400'>· {_main_kw}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    "<div style='font-size:0.70rem;color:#bbb;margin:4px 0 2px 0;"
                    "display:flex;align-items:center;gap:4px'>"
                    "<span>🔥</span><span>키워드 입력 후 서브 주제 추출</span></div>",
                    unsafe_allow_html=True
                )
        with _col_hot2:
            _hot_btn = st.button(
                "추출",
                key="btn_hot_topics",
                disabled=not (_main_kw and api_key),
                help="첫 번째 키워드로 실시간 인기 서브 주제 TOP 10을 추출합니다"
            )

        if _hot_btn and _main_kw and api_key:
            with st.spinner(f"🔥 '{_main_kw}' 관련 인기 서브 주제 분석 중..."):
                try:
                    _topics, _err = get_hot_subtopics(api_key, _main_kw, top_n=10)
                except Exception as _exc:
                    _topics, _err = None, f"예기치 않은 오류: {_exc}"
            if _err:
                st.error(f"서브 주제 추출 실패: {_err}")
            elif not _topics:
                st.warning("서브 주제를 찾지 못했습니다. API 키와 키워드를 확인해주세요.")
            else:
                st.session_state["hot_topics"]    = _topics
                st.session_state["hot_topics_kw"] = _main_kw
                st.rerun()

        # ── 추출된 서브 주제 표시 (세로 1열) ─────────────────────────────
        _topics_now = st.session_state.get("hot_topics", [])
        _topics_kw  = st.session_state.get("hot_topics_kw", "")
        if _topics_now:
            _exp_label = (
                f"🔥 '{_topics_kw}' 인기 서브 주제 "
                f"TOP {len(_topics_now)} — 클릭하여 열기/접기"
            )
            with st.expander(_exp_label, expanded=st.session_state.get('hot_exp_open', True)):
                _max_score = max(t["score"] for t in _topics_now) or 1
                # 출처 통계
                _src_count = {}
                for _t in _topics_now:
                    _src_count[_t.get("source","?")] = _src_count.get(_t.get("source","?"),0) + 1
                _src_labels = []
                if _src_count.get("video"):    _src_labels.append(f"🔥인기영상 {_src_count['video']}개")
                if _src_count.get("suggest"):  _src_labels.append(f"🔴실시간검색 {_src_count['suggest']}개")
                if _src_count.get("trends"):   _src_labels.append(f"📈급상승트렌드 {_src_count['trends']}개")

                # 출처 칩
                # ── pytrends 미설치 안내 ────────────────────────────
                try:
                    import pytrends as _pyt_check
                    _pytrends_ok = True
                except Exception:
                    _pytrends_ok = False
                if not _pytrends_ok:
                    import sys as _sys
                    _py_path = _sys.executable
                    st.warning(
                        "⚠️ **Google Trends(pytrends) 미설치** — Google Trends 결과가 빠져 있습니다.  \n"
                        "아래 명령어를 터미널에서 실행 후 앱을 재시작하세요:  \n"
                        f"```bash\n{_py_path} -m pip install pytrends\n```",
                        icon=None
                    )

                _chip_html = "<div style='display:flex;flex-wrap:wrap;gap:4px;margin:4px 0 8px 0'>"

                _chip_colors = {"video":"#fff1ee:#e55a2b", "suggest":"#fff0f0:#d32f2f", "trends":"#eff6ff:#1565c0"}
                for _src_k, _src_label in [("video","🔥 인기영상"), ("suggest","🔴 실시간"), ("trends","📈 급상승")]:
                    if _src_count.get(_src_k):
                        _cbg, _cfg = _chip_colors.get(_src_k, "#f5f5f5:#555").split(":")
                        _chip_html += (
                            f"<span style='background:{_cbg};color:{_cfg};font-size:0.62rem;"
                            f"font-weight:700;padding:2px 7px;border-radius:10px;"
                            f"border:1px solid {_cfg}22'>{_src_label} {_src_count[_src_k]}</span>"
                        )
                _chip_html += (
                    "<span style='background:#f0f2fb;color:#8a93a8;font-size:0.60rem;"
                    "padding:2px 7px;border-radius:10px;margin-left:auto'>"
                    "클릭 → 검색창 입력</span></div>"
                )
                st.markdown(_chip_html, unsafe_allow_html=True)

                # ── 세로 1열 카드 (인기도 숫자 + 바 한눈 비교) ─────────────
                _cur_kws_check = [k.strip() for k in keywords_input.replace("，", ",").split(",") if k.strip()]
                for _ti, _t in enumerate(_topics_now):
                    _already_added = _t["topic"] in _cur_kws_check
                    _bar_w   = max(6, int(_t["score"] / _max_score * 100))
                    _label_badge = _t.get("label", "🔥 인기")
                    _source  = _t.get("source", "video")

                    # ── 소스별 인기도 숫자 포매팅 ──────────────────────────
                    _raw_v   = _t.get("raw_views", 0)
                    _tval    = _t.get("trend_val", 0)
                    _srank   = _t.get("sug_rank", 0)

                    if _source == "video" and _raw_v > 0:
                        # 실제 조회수 → 가장 큰 숫자로 표시
                        if _raw_v >= 100_000_000:
                            _hot_num = f"{_raw_v/100_000_000:.1f}억"
                        elif _raw_v >= 10_000:
                            _hot_num = f"{_raw_v/10_000:.0f}만"
                        elif _raw_v >= 1_000:
                            _hot_num = f"{_raw_v/1_000:.0f}천"
                        else:
                            _hot_num = f"{_raw_v:,}"
                        _hot_unit  = "조회수"
                        _hot_color = "#e53935"
                    elif _source == "suggest" and _srank > 0:
                        # 실시간 검색 순위
                        _hot_num   = f"#{_srank}"
                        _hot_unit  = "실검순위"
                        _hot_color = "#d32f2f"
                    elif _source == "trends" and _tval > 0:
                        # Google Trends 급상승률
                        _hot_num   = f"+{_tval}%"
                        _hot_unit  = "급상승"
                        _hot_color = "#1565c0"
                    else:
                        _hot_num   = f"TOP{_ti+1}"
                        _hot_unit  = "인기"
                        _hot_color = "#555"

                    # ── 순위별 색상 (모던 블루 그라디언트) ─────────────────
                    _rank_color = (
                        "#4f7ef8" if _ti == 0 else
                        "#7c5af8" if _ti == 1 else
                        "#e55a2b" if _ti == 2 else
                        "#3b82f6" if _ti < 6 else
                        "#64748b" if _ti < 8 else "#94a3b8"
                    )
                    _bg_color  = "#f5f7ff" if _ti == 0 else "#ffffff"
                    _txt_color = "#2e7d32" if _already_added else "#2d3250"

                    # ── 날짜 / 채널 정보 ───────────────────────────────────
                    _date_str    = _t.get("date", "")
                    _channel_str = _t.get("channel", "")
                    _sub_parts   = []
                    if _date_str:    _sub_parts.append(f"📅 {_date_str}")
                    if _channel_str: _sub_parts.append(f"📺 {_channel_str}")
                    _sub_html = "  ·  ".join(_sub_parts)
                    _sub_row  = (f"<div style='font-size:0.63rem;color:#aaa;margin-bottom:4px'>{_sub_html}</div>"
                                 if _sub_html else "")

                    # ── 카드 HTML (모던 재설계) ───────────────────────────
                    _card_bg    = "#ffffff"
                    _card_bd    = "#e8ebf5" if not _already_added else "#bbf7d0"
                    _txt_color2 = "#2d3250" if not _already_added else "#15803d"
                    _html_card = (
                        f"<div style='background:{_card_bg};"
                        f"border:1.5px solid {_card_bd};"
                        f"border-radius:10px;padding:8px 10px 6px 10px;margin-bottom:5px;"
                        f"border-left:4px solid {_rank_color};"
                        f"box-shadow:0 1px 4px rgba(0,0,0,0.06);'>"
                        # 1행: 순위뱃지 + 라벨 + 인기도 숫자
                        f"<div style='display:flex;justify-content:space-between;"
                        f"align-items:center;margin-bottom:4px'>"
                        f"<div style='display:flex;align-items:center;gap:5px;flex:1;min-width:0'>"
                        f"<span style='background:{_rank_color};color:#fff;"
                        f"font-weight:800;font-size:0.64rem;min-width:20px;height:20px;"
                        f"border-radius:6px;display:flex;align-items:center;justify-content:center;"
                        f"flex-shrink:0'>#{_ti+1}</span>"
                        f"<span style='font-size:0.60rem;background:#f0f2fb;"
                        f"padding:1px 6px;border-radius:6px;color:#8a93a8;"
                        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                        f"max-width:80px'>{_label_badge}</span>"
                        f"</div>"
                        f"<div style='text-align:right;line-height:1.1;flex-shrink:0'>"
                        f"<div style='font-size:0.95rem;font-weight:800;"
                        f"color:{_hot_color};letter-spacing:-0.5px'>{_hot_num}</div>"
                        f"<div style='font-size:0.55rem;color:#b0b8cc'>{_hot_unit}</div>"
                        f"</div></div>"
                        # 2행: 제목
                        f"<div style='font-size:0.81rem;font-weight:600;color:{_txt_color2};"
                        f"line-height:1.4;margin-bottom:3px;"
                        f"overflow:hidden;display:-webkit-box;"
                        f"-webkit-line-clamp:2;-webkit-box-orient:vertical;'>"
                        f"{'✅ ' if _already_added else ''}{_t['topic']}"
                        f"</div>"
                    )
                    # 3행: 날짜·채널 (있을 때만 추가)
                    if _sub_html:
                        _html_card += (
                            f"<div style='font-size:0.63rem;color:#aaa;margin-bottom:4px'>"
                            f"{_sub_html}</div>"
                        )
                    # 4행: 인기도 바 + 퍼센트 수치
                    _html_card += (
                        f"<div style='display:flex;align-items:center;gap:5px'>"
                        f"<div style='flex:1;height:4px;border-radius:2px;background:#efefef'>"
                        f"<div style='height:4px;border-radius:2px;"
                        f"background:linear-gradient(90deg,{_rank_color},{_rank_color}55);"
                        f"width:{_bar_w}%'></div>"
                        f"</div>"
                        f"<span style='font-size:0.60rem;color:{_rank_color};"
                        f"font-weight:700;min-width:26px;text-align:right'>{_bar_w}%</span>"
                        f"</div>"
                        f"</div>"
                    )
                    # ── 확대 스파크라인 SVG (날짜별 수치 포함) ──────────
                    _sp_data = _t.get("sparkline", [])
                    if not _sp_data:
                        _sp_data = [50]*7
                    # 날짜 레이블 (오늘 기준 D-6 ~ D-0)
                    from datetime import datetime as _dtnow, timedelta as _td
                    _today = _dtnow.now()
                    _date_labels = [(_today - _td(days=6-i)).strftime("%m/%d") for i in range(7)]
                    _sp_padded = (_sp_data + [50]*7)[:7]
                    _W2, _H2 = 200, 56   # 더 크게
                    _sp_n2   = len(_sp_padded)
                    _sp_max2 = max(_sp_padded) or 1
                    _sp_min2 = min(_sp_padded)
                    _sp_rng2 = _sp_max2 - _sp_min2 or 1
                    _pts2    = []
                    for _si2, _sv2 in enumerate(_sp_padded):
                        _sx2 = int(_si2 / (_sp_n2 - 1) * (_W2 - 16)) + 8
                        _sy2 = int((1 - (_sv2 - _sp_min2) / _sp_rng2) * (_H2 - 20)) + 6
                        _pts2.append((_sx2, _sy2, _sv2))
                    _polyline2 = " ".join(f"{x},{y}" for x,y,_ in _pts2)
                    _sp_trend2 = _sp_padded[-1] - _sp_padded[0]
                    _sp_pal    = _chart_palette(st.session_state.get("dark_mode", False))
                    _sp_color2 = _sp_pal["rise"] if _sp_trend2 > 5 else (_sp_pal["fall"] if _sp_trend2 < -5 else _sp_pal["flat"])
                    _sp_arrow2 = "▲" if _sp_trend2 > 5 else ("▼" if _sp_trend2 < -5 else "━")
                    # 그라디언트 fill 영역 (area chart)
                    _fill_pts  = f"8,{_H2-2} " + " ".join(f"{x},{y}" for x,y,_ in _pts2) + f" {_pts2[-1][0]},{_H2-2}"
                    # 각 포인트: 원 + 수치 레이블
                    _circles_html = ""
                    for _ci, (_cx, _cy, _cv) in enumerate(_pts2):
                        _is_max = (_cv == _sp_max2)
                        _is_min = (_cv == _sp_min2)
                        _cr     = 3.5 if (_is_max or _is_min) else 2.5
                        _cfill  = _sp_pal["rise"] if _is_max else (_sp_pal["fall"] if _is_min else _sp_color2)
                        # 수치 레이블 (최고/최저/마지막 포인트만 표시)
                        if _is_max or _is_min or _ci == _sp_n2 - 1:
                            _lbl_y  = _cy - 5 if _cy > 14 else _cy + 12
                            _lbl_x  = max(8, min(_W2-16, _cx))
                            _circles_html += (
                                f"<circle cx='{_cx}' cy='{_cy}' r='{_cr}' fill='{_cfill}' stroke='white' stroke-width='1'/>"
                                f"<text x='{_lbl_x}' y='{_lbl_y}' text-anchor='middle' "
                                f"font-size='7' fill='{_cfill}' font-weight='bold'>{_cv}</text>"
                            )
                        else:
                            _circles_html += f"<circle cx='{_cx}' cy='{_cy}' r='{_cr}' fill='{_sp_color2}' stroke='white' stroke-width='1'/>"
                    # 날짜 레이블 (X축)
                    _xlabels_html = ""
                    for _xi, ((_xx, _xy, _), _dl) in enumerate(zip(_pts2, _date_labels)):
                        if _xi == 0 or _xi == 3 or _xi == 6:
                            _sp_xlabel_color = _sp_pal["x_label"]
                            _xlabels_html += f"<text x='{_xx}' y='{_H2+1}' text-anchor='middle' font-size='6.5' fill='{_sp_xlabel_color}'>{_dl}</text>"
                    # 그리드라인 (수평 3개)
                    _grid_html = ""
                    for _gi in range(1, 4):
                        _gy = int(_H2 * _gi / 4)
                        _sp_grid_color = _sp_pal["grid"]
                        _grid_html += f"<line x1='8' y1='{_gy}' x2='{_W2-8}' y2='{_gy}' stroke='{_sp_grid_color}' stroke-width='0.8'/>"
                    _svg_big = (
                        f"<svg width='{_W2}' height='{_H2+10}' style='overflow:visible'>"
                        f"<defs><linearGradient id='sg{_ti}' x1='0' y1='0' x2='0' y2='1'>"
                        f"<stop offset='0%' stop-color='{_sp_color2}' stop-opacity='0.18'/>"
                        f"<stop offset='100%' stop-color='{_sp_color2}' stop-opacity='0.02'/>"
                        f"</linearGradient></defs>"
                        f"{_grid_html}"
                        f"<polygon points='{_fill_pts}' fill='url(#sg{_ti})'/>"
                        f"<polyline points='{_polyline2}' fill='none' stroke='{_sp_color2}' stroke-width='2' stroke-linejoin='round'/>"
                        f"{_circles_html}"
                        f"{_xlabels_html}"
                        f"</svg>"
                        f"<span style='font-size:0.7rem;color:{_sp_color2};font-weight:800;margin-left:4px'>{_sp_arrow2}</span>"
                    )

                    # 스파크라인 행 (카드 하단 구분선 위에 삽입)
                    _sp_row_html = (
                        f"<div style='margin-top:6px;padding-top:5px;"
                        f"border-top:1px solid {_sp_pal['grid']};'>"
                        f"<div style='font-size:0.6rem;color:{_sp_pal['sparkline_label']};margin-bottom:2px'>📈 최근 7일 조회 추이</div>"
                        f"<div style='overflow:hidden'>{_svg_big}</div>"
                        f"</div>"
                    )
                    _html_card = _html_card[:-6] + _sp_row_html + "</div>" 

                    st.markdown(_html_card, unsafe_allow_html=True)

                    # ── 버튼 행 4열 균등: [＋추가 | 📊 | 🔍 | ★] ──────────────
                    _is_fav    = any(f["topic"] == _t["topic"] for f in st.session_state.get("favorites", []))
                    _is_inline = st.session_state.get("inline_search_topic","") == _t["topic"]

                    # 버튼 라벨 (최대한 짧게 — 사이드바 폭 대응)
                    _lbl_add  = "✓" if _already_added else "＋ 추가"
                    _lbl_srch = "🔴" if _is_inline else "🔍"
                    _lbl_fav  = "★" if _is_fav else "☆"
                    _help_add = "이미 검색창에 추가됨" if _already_added else f"'{_t['topic']}' 검색창에 입력"
                    _help_fav = "즐겨찾기 해제" if _is_fav else "즐겨찾기에 저장"
                    # 폴더 선택 (★ 버튼 옆 — 미즐겨찾기 상태일 때만)
                    _folder_list = list(st.session_state.get("fav_folders", {"기본 폴더": []}).keys())
                    if not _folder_list:
                        _folder_list = ["기본 폴더"]
                    _cur_target  = st.session_state.get("fav_add_target_folder", "기본 폴더")
                    if _cur_target not in _folder_list:
                        _cur_target = _folder_list[0]

                    _yt_url_q  = urllib.parse.quote(_t['topic'])
                    _yt_url    = f"https://www.youtube.com/results?search_query={_yt_url_q}"
                    _btn_col1, _btn_col2, _btn_col3, _btn_col4, _btn_col5 = st.columns([3, 1, 1, 1, 1])
                    with _btn_col1:
                        if st.button(
                            _lbl_add,
                            key=f"hot_topic_{_ti}",
                            help=_help_add,
                            use_container_width=True,
                            disabled=_already_added
                        ):
                            st.session_state["hot_topic_clicked"] = _t["topic"]
                            st.rerun()
                    with _btn_col2:
                        if st.button(
                            "📊",
                            key=f"detail_btn_{_ti}",
                            help="7일 상세 트렌드 그래프",
                            use_container_width=True
                        ):
                            st.session_state["detail_chart_topic"] = _t["topic"]
                            st.session_state["detail_chart_data"]  = _t
                            st.rerun()
                    with _btn_col3:
                        if st.button(
                            _lbl_srch,
                            key=f"inline_btn_{_ti}",
                            help="앱 내 유튜브 검색 결과 보기 (토글)",
                            use_container_width=True
                        ):
                            if _is_inline:
                                st.session_state["inline_search_topic"]   = ""
                                st.session_state["inline_search_results"] = []
                            else:
                                st.session_state["inline_search_topic"] = _t["topic"]
                                with st.spinner(f"🔍 '{_t['topic']}' 검색 중..."):
                                    st.session_state["inline_search_results"] = (
                                        get_related_videos(api_key, _t["topic"], top_n=5)
                                        if api_key else []
                                    )
                            st.rerun()
                    with _btn_col4:
                        st.link_button(
                            "📺",
                            url=_yt_url,
                            help=f"'{_t['topic']}' YouTube 검색 결과 새 탭에서 열기",
                            use_container_width=True
                        )
                    with _btn_col5:
                        if st.button(
                            _lbl_fav,
                            key=f"fav_btn_{_ti}",
                            help=_help_fav,
                            use_container_width=True,
                        ):
                            if _is_fav:
                                st.session_state["fav_action"] = f"DEL:{_t['topic']}"
                            else:
                                st.session_state["fav_action"] = f"ADD:{_t['topic']}"
                            st.rerun()
                    # ── 폴더 선택 (미즐겨찾기 상태 + 폴더가 2개 이상일 때만 표시) ──
                    if not _is_fav and len(_folder_list) > 1:
                        _sel_folder = st.selectbox(
                            "저장 폴더",
                            options=_folder_list,
                            index=_folder_list.index(_cur_target) if _cur_target in _folder_list else 0,
                            key=f"folder_sel_{_ti}",
                            label_visibility="collapsed",
                            help="★ 버튼 클릭 시 저장될 폴더를 선택하세요"
                        )
                        if _sel_folder != st.session_state.get("fav_add_target_folder", "기본 폴더"):
                            st.session_state["fav_add_target_folder"] = _sel_folder
                            st.rerun()

                    # ── 인라인 검색 결과 패널 (해당 카드 바로 아래) ──────────
                    if _is_inline:
                        _isr = st.session_state.get("inline_search_results", [])
                        _panel_html = (
                            f"<div style='background:#f0f6ff;"
                            f"border:1.5px solid #bfdbfe;"
                            f"border-radius:10px;padding:10px 12px;margin:4px 0 8px 0;"
                            f"border-left:4px solid #4f7ef8;'>"
                            f"<div style='display:flex;align-items:center;gap:5px;margin-bottom:8px'>"
                            f"<span style='font-size:0.68rem;font-weight:800;color:#1d4ed8'>🔍 관련 영상</span>"
                            f"<span style='font-size:0.63rem;color:#93c5fd;font-weight:400'>"
                            f"· {_t['topic'][:20]}{'...' if len(_t['topic'])>20 else ''}</span>"
                            f"<span style='margin-left:auto;font-size:0.58rem;background:#dbeafe;"
                            f"color:#1d4ed8;padding:1px 6px;border-radius:8px'>TOP {len(_isr)}</span>"
                            f"</div>"
                        )
                        if not _isr:
                            _panel_html += "<div style='color:#aaa;font-size:0.7rem'>API 키 필요 또는 결과 없음</div>"
                        for _ri, _rv in enumerate(_isr):
                            _rv_views = fmt(_rv.get("views",0)) + "회"
                            _rv_title = _rv.get("title","")[:32] + ("..." if len(_rv.get("title",""))>32 else "")
                            _rv_ch    = _rv.get("channel","")
                            _rv_date  = _rv.get("date","")
                            _rv_url   = _rv.get("url","")
                            _rv_thumb = _rv.get("thumbnail","")
                            _rank_c   = ["#E53935","#F57C00","#795548","#1976D2","#546E7A"][_ri]
                            _panel_html += (
                                f"<div style='display:flex;align-items:center;gap:8px;"
                                f"margin-bottom:7px;padding-bottom:7px;"
                                f"border-bottom:1px solid #bbdefb;'>"
                                f"<span style='font-weight:900;font-size:0.82rem;color:{_rank_c};"
                                f"min-width:18px'>#{_ri+1}</span>"
                            )
                            if _rv_thumb:
                                _panel_html += (
                                    f"<img src='{_rv_thumb}' width='60' height='34' "
                                    f"style='border-radius:4px;object-fit:cover;flex-shrink:0'/>"
                                )
                            _panel_html += (
                                f"<div style='flex:1;min-width:0'>"
                                f"<a href='{_rv_url}' target='_blank' style='font-size:0.75rem;"
                                f"font-weight:700;color:#1976D2;text-decoration:none;"
                                f"display:block;overflow:hidden;white-space:nowrap;"
                                f"text-overflow:ellipsis'>{_rv_title}</a>"
                                f"<div style='font-size:0.62rem;color:#888;margin-top:2px'>"
                                f"📺 {_rv_ch}  ·  👁 {_rv_views}  ·  📅 {_rv_date}</div>"
                                f"</div></div>"
                            )
                        _panel_html += "</div>"
                        st.markdown(_panel_html, unsafe_allow_html=True)

                # ── 하단 액션 버튼 행 ─────────────────────────────────────
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                _act_c1, _act_c2 = st.columns(2)
                with _act_c1:
                    st.markdown('<div class="btn-danger">', unsafe_allow_html=True)
                    if st.button("🗑️ 목록 지우기", key="btn_clear_hot", use_container_width=True):
                        st.session_state["hot_topics"]    = []
                        st.session_state["hot_topics_kw"] = ""
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with _act_c2:
                    _can_export = HAS_GSHEET and bool(_s_gcp_creds or False)
                    st.markdown('<div class="btn-success">', unsafe_allow_html=True)
                    if st.button(
                        "📊 시트 저장",
                        key="btn_export_subtopics",
                        use_container_width=True,
                        disabled=not _can_export,
                        help="GCP credentials 설정 필요" if not _can_export else f"'{_topics_kw}' 서브주제 10개를 구글시트에 누적 저장"
                    ):
                        st.session_state["subtopic_export_trigger"] = True
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

                # ── 구글시트 내보내기 실행 ────────────────────────────────
                if st.session_state.pop("subtopic_export_trigger", False):
                    with st.spinner("📊 구글시트에 서브주제 저장 중..."):
                        _ok, _msg = export_subtopics_to_gsheet(
                            _topics_now, _topics_kw,
                            credentials_dict=_s_gcp_creds,
                            existing_id=_s_existing,
                            share_email=_s_email,
                            api_key=api_key,
                        )
                    if _ok:
                        _url_part, _cnt = _msg.split("|||")
                        st.success(f"✅ 구글시트 저장 완료 (누적 {_cnt}행)")
                        st.markdown(f"[📋 시트 열기]({_url_part})", unsafe_allow_html=False)
                    else:
                        st.error(f"❌ 내보내기 실패: {_msg}")

        # ── 즐겨찾기 폴더 섹션 ─────────────────────────────────────────
        _favs_now   = st.session_state.get("favorites", [])
        _fav_folders_d = st.session_state.get("fav_folders", {"기본 폴더": []})
        _dm_fav     = st.session_state.get("dark_mode", False)

        # 폴더 액션 처리 (rename / new / del)
        _ff_action = st.session_state.pop("fav_folder_action", "")
        if _ff_action.startswith("NEW_FOLDER:"):
            _nf = _ff_action[11:].strip()
            if _nf and _nf not in _fav_folders_d:
                _fav_folders_d[_nf] = []
                st.session_state["fav_folders"] = _fav_folders_d
                st.session_state["fav_new_folder_mode"] = False
                st.rerun()
        elif _ff_action.startswith("RENAME_FOLDER:"):
            _rp = _ff_action[14:].split(":", 1)
            if len(_rp) == 2:
                _ro, _rn = _rp[0], _rp[1].strip()
                if _rn and _ro != _rn and _rn not in _fav_folders_d:
                    _fav_folders_d[_rn] = _fav_folders_d.pop(_ro, [])
                    st.session_state["fav_folders"] = _fav_folders_d
                    for _fi2 in st.session_state.get("favorites", []):
                        if _fi2.get("folder") == _ro:
                            _fi2["folder"] = _rn
                    st.session_state["fav_rename_folder"] = ""
                    st.rerun()
        elif _ff_action.startswith("DEL_FOLDER:"):
            _df2 = _ff_action[11:]
            if _df2 != "기본 폴더":
                _orphans2 = _fav_folders_d.pop(_df2, [])
                _fav_folders_d.setdefault("기본 폴더", []).extend(_orphans2)
                st.session_state["fav_folders"] = _fav_folders_d
                for _fi3 in st.session_state.get("favorites", []):
                    if _fi3.get("folder") == _df2:
                        _fi3["folder"] = "기본 폴더"
                st.rerun()

        # 즐겨찾기 있을 때 전체 컨테이너 헤더
        _total_fav = len(_favs_now)
        _all_folders = list(_fav_folders_d.keys())

        # ── 즐겨찾기 전체 expander (최상위) ──────────────────────────
        _fav_main_label = (
            f"⭐ 즐겨찾기 {_total_fav}개  ·  📁 {len(_all_folders)}폴더"
            if _total_fav > 0 else "⭐ 즐겨찾기 — 클릭하여 열기"
        )
        # palette
        _fmain_bg   = "#1e2038" if _dm_fav else "#f8f9ff"
        _fmain_bd   = "#2e3157" if _dm_fav else "#e0e5ff"
        _fhdr_c     = "#7c9ef8" if _dm_fav else "#4f7ef8"
        _ftxt_c     = "#c5c8e0" if _dm_fav else "#3a4166"
        _fbadge_bg  = "#252640" if _dm_fav else "#eef2ff"
        _fbadge_c   = "#a0aee0" if _dm_fav else "#4338ca"
        _folder_hdr_bg  = "#252640" if _dm_fav else "#f0f4ff"
        _folder_hdr_bd  = "#3a3d60" if _dm_fav else "#c7d2fe"
        _folder_hdr_c   = "#8090d0" if _dm_fav else "#4338ca"
        _card_bg        = "#1e2038" if _dm_fav else "#ffffff"
        _card_bd_added  = "#2a5040" if _dm_fav else "#86efac"
        _card_bd_normal = "#3a3d60" if _dm_fav else "#fde68a"
        _card_lc_added  = "#4ade80" if _dm_fav else "#16a34a"
        _card_lc_normal = "#fbbf24" if _dm_fav else "#d97706"
        _card_title_c   = "#d0d4f0" if _dm_fav else "#2d3250"
        _card_sub_c     = "#6070a0" if _dm_fav else "#b0b8cc"
        _badge_bg_kw    = "#2a2e50" if _dm_fav else "#fef9c3"
        _badge_c_kw     = "#a0a8d0" if _dm_fav else "#a16207"

        with st.expander(_fav_main_label, expanded=st.session_state.get("fav_exp_open", False)):

            # ── 상단 액션 바: 새 폴더 만들기 / 대시보드 ─────────────
            _act_c1, _act_c2 = st.columns([1, 1])
            with _act_c1:
                if st.button(
                    "📁 새 폴더",
                    key="btn_new_folder",
                    use_container_width=True,
                    help="새 폴더를 만들어 즐겨찾기를 분류하세요"
                ):
                    st.session_state["fav_new_folder_mode"] = not st.session_state.get("fav_new_folder_mode", False)
                    st.rerun()
            with _act_c2:
                _fdb_open = st.session_state.get("fav_dashboard_open", False)
                if st.button(
                    "📊 대시보드 닫기" if _fdb_open else "📊 대시보드",
                    key="btn_fav_dashboard",
                    use_container_width=True
                ):
                    st.session_state["fav_dashboard_open"] = not _fdb_open
                    st.rerun()

            # ── 새 폴더 입력 폼 ──────────────────────────────────────
            if st.session_state.get("fav_new_folder_mode", False):
                st.markdown(
                    f"<div style='background:{_fmain_bg};border:1px solid {_fmain_bd};"
                    f"border-radius:8px;padding:8px 10px;margin:4px 0'>"
                    f"<span style='font-size:0.68rem;color:{_fhdr_c};font-weight:700'>📁 새 폴더 이름 입력</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                _nf_inp_c, _nf_btn_c = st.columns([3, 1])
                with _nf_inp_c:
                    _new_fname = st.text_input(
                        "폴더명",
                        key="new_folder_name_input",
                        placeholder="예: 요리, 여행, 뷰티...",
                        label_visibility="collapsed"
                    )
                with _nf_btn_c:
                    if st.button("✚ 생성", key="btn_create_folder", use_container_width=True):
                        if _new_fname and _new_fname.strip():
                            st.session_state["fav_folder_action"] = f"NEW_FOLDER:{_new_fname.strip()}"
                            st.rerun()
                st.markdown("<hr style='margin:4px 0'>", unsafe_allow_html=True)

            # ── 📊 대시보드 뷰 ────────────────────────────────────────
            if st.session_state.get("fav_dashboard_open", False):
                # 정렬 옵션 (대시보드 내)
                _fsort_db = st.selectbox(
                    "정렬",
                    options=["저장순", "조회수순", "저장날짜↑"],
                    key="fav_sort_db",
                    label_visibility="collapsed"
                )
                _favs_db = list(_favs_now)
                if _fsort_db == "조회수순":
                    _favs_db.sort(key=lambda f: f.get("score", 0), reverse=True)
                elif _fsort_db == "저장날짜↑":
                    _favs_db.sort(key=lambda f: f.get("saved", ""), reverse=False)

                _fdb_items = [f for f in _favs_db if f.get("sparkline")]
                if not _fdb_items:
                    st.info("📈 스파크라인 데이터가 있는 즐겨찾기가 없습니다.")
                else:
                    st.markdown(
                        f"<div style='font-size:0.72rem;font-weight:800;color:{_fhdr_c};"
                        f"margin:4px 0 6px 0'>📊 조회수 추이 비교</div>",
                        unsafe_allow_html=True
                    )
                    _fdb_show    = _fdb_items[:6]
                    _fdb_all_max = max((max(f.get("sparkline", [50])) for f in _fdb_show), default=1)
                    _fdb_pairs   = [_fdb_show[i:i+2] for i in range(0, len(_fdb_show), 2)]
                    _fdb_pal     = _chart_palette(_dm_fav)
                    for _pair in _fdb_pairs:
                        _db_c = st.columns(len(_pair))
                        for _dci, _ditem in enumerate(_pair):
                            with _db_c[_dci]:
                                _dsp   = (_ditem.get("sparkline", [50]*7) + [50]*7)[:7]
                                _dmax  = max(_dsp) or 1
                                _dmin  = min(_dsp)
                                _drng  = _dmax - _dmin or 1
                                _dW, _dH = 140, 50
                                _dpts  = []
                                for _dsi, _dsv in enumerate(_dsp):
                                    _dsx = int(_dsi / 6 * (_dW - 12)) + 6
                                    _dsy = int((1 - (_dsv - _dmin) / _drng) * (_dH - 16)) + 6
                                    _dpts.append((_dsx, _dsy, _dsv))
                                _dpoly  = " ".join(f"{x},{y}" for x,y,_ in _dpts)
                                _dtrend = _dsp[-1] - _dsp[0]
                                _fdb_pal2 = _chart_palette(_dm_fav)
                                _dcol   = _fdb_pal2["rise"] if _dtrend > 5 else (_fdb_pal2["fall"] if _dtrend < -5 else _fdb_pal2["flat"])
                                _darrow = "▲" if _dtrend > 5 else ("▼" if _dtrend < -5 else "━")
                                _dfill  = f"6,{_dH} " + _dpoly + f" {_dpts[-1][0]},{_dH}"
                                _last_val  = _dsp[-1]
                                _first_val = _dsp[0]
                                _dtitle_full = _ditem.get("topic", "")
                                _dtitle  = _dtitle_full[:14] + ("…" if len(_dtitle_full) > 14 else "")
                                _dkw     = _ditem.get("keyword", "")[:8]
                                _dscore  = _ditem.get("score", 0)
                                _dbar_w  = max(4, int(_dscore / (_fdb_all_max or 1) * 100))
                                _dfolder_badge = _ditem.get("folder", "기본 폴더")[:6]
                                _dsvg = (
                                    f"<div style='background:{_fdb_pal2['sparkline_bg']};border:1px solid {_fdb_pal2['sparkline_border']};"
                                    f"border-radius:10px;padding:7px 8px 5px 8px;margin-bottom:6px;"
                                    f"border-top:3px solid {_dcol};'>"
                                    f"<div style='font-size:0.72rem;font-weight:800;color:{_fdb_pal2['summary_val']};"
                                    f"margin-bottom:1px;overflow:hidden;white-space:nowrap;"
                                    f"text-overflow:ellipsis' title='{_dtitle_full}'>{_dtitle}</div>"
                                    f"<div style='font-size:0.57rem;color:{_fdb_pal2['sparkline_label']};margin-bottom:4px'>"
                                    f"📁 {_dfolder_badge}  ·  {_ditem.get('saved', '')}</div>"
                                    f"<svg width='100%' viewBox='0 0 {_dW} {_dH+8}' style='overflow:visible'>"
                                    f"<defs><linearGradient id='dbg{_dci}_{id(_ditem)}' x1='0' y1='0' x2='0' y2='1'>"
                                    f"<stop offset='0%' stop-color='{_dcol}' stop-opacity='{_fdb_pal2['area_opacity_hi']}'/>"
                                    f"<stop offset='100%' stop-color='{_dcol}' stop-opacity='{_fdb_pal2['area_opacity_lo']}'/>"
                                    f"</linearGradient></defs>"
                                    f"<polygon points='{_dfill}' fill='url(#dbg{_dci}_{id(_ditem)})'/>"
                                    f"<polyline points='{_dpoly}' fill='none' stroke='{_dcol}' stroke-width='2' stroke-linejoin='round'/>"
                                    + "".join(
                                        "<circle cx='" + str(x) + "' cy='" + str(y) + "' r='" + ("3.5" if v in (_dmax, _dmin) else "2") + "' "
                                        "fill='" + (_fdb_pal2["rise"] if v == _dmax else (_fdb_pal2["fall"] if v == _dmin else _dcol)) + "' stroke='" + _fdb_pal2["bg"] + "' stroke-width='1'/>"
                                        for x, y, v in _dpts
                                    )
                                    + "<text x='" + str(_dpts[-1][0]) + "' y='" + str(max(6, _dpts[-1][1]-5)) + "' text-anchor='middle' "
                                    + "font-size='8' fill='" + _dcol + "' font-weight='bold'>" + str(_last_val) + "</text>"
                                    + "<text x='6' y='" + str(max(6, _dpts[0][1]-5)) + "' text-anchor='start' "
                                    + "font-size='7' fill='" + _fdb_pal2["x_label"] + "'>" + str(_first_val) + "</text>"
                                    f"</svg>"
                                    f"<div style='margin-top:4px;display:flex;align-items:center;gap:4px'>"
                                    + "<div style='flex:1;height:3px;border-radius:2px;background:" + _fdb_pal2["grid"] + "'>"
                                    f"<div style='height:3px;border-radius:2px;background:{_dcol};width:{_dbar_w}%'></div></div>"
                                    f"<span style='font-size:0.58rem;color:{_dcol};font-weight:700'>{_darrow}{abs(_dtrend)}</span>"
                                    f"</div>"
                                    f"</div>"
                                )
                                st.markdown(_dsvg, unsafe_allow_html=True)
                    if len(_fdb_items) > 6:
                        st.caption(f"※ 상위 6개만 표시 (전체 {len(_fdb_items)}개)")
                st.markdown("<hr style='margin:6px 0'>", unsafe_allow_html=True)

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 📁 폴더별 카드 렌더링
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # favorites를 folder 기준으로 그룹화
            _fav_by_folder = {}
            for _fitem in _favs_now:
                _fol = _fitem.get("folder", "기본 폴더")
                if _fol not in _fav_by_folder:
                    _fav_by_folder[_fol] = []
                _fav_by_folder[_fol].append(_fitem)
            # fav_folders에 있지만 아직 항목 없는 폴더도 표시
            for _fn in _fav_folders_d:
                if _fn not in _fav_by_folder:
                    _fav_by_folder[_fn] = []
            # 폴더 순서: fav_folders 키 순서 유지
            _ordered_folders = [f for f in _fav_folders_d if f in _fav_by_folder]

            _global_fi = 0  # 전체 카드 인덱스 (button key 충돌 방지)
            for _fol_name in _ordered_folders:
                _fol_items = _fav_by_folder[_fol_name]
                _fol_cnt   = len(_fol_items)
                # 폴더 헤더 HTML
                _fol_open_key = f"fol_open_{_fol_name}"
                _fol_is_open  = st.session_state.get(_fol_open_key, True)

                # 폴더 헤더 렌더링
                st.markdown(
                    f"<div style='background:{_folder_hdr_bg};border:1px solid {_folder_hdr_bd};"
                    f"border-radius:8px;padding:5px 10px;margin:5px 0 2px 0;"
                    f"display:flex;align-items:center;justify-content:space-between'>"
                    f"<span style='font-size:0.78rem;font-weight:800;color:{_folder_hdr_c}'>"
                    f"{'📂' if _fol_is_open else '📁'} {_fol_name}"
                    f"<span style='font-size:0.62rem;font-weight:500;margin-left:5px;"
                    f"background:{_fbadge_bg};color:{_fbadge_c};padding:1px 6px;"
                    f"border-radius:8px'>{_fol_cnt}개</span></span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                # 폴더 열기/접기 + 이름변경 + 삭제 버튼 행
                _fhdr_c1, _fhdr_c2, _fhdr_c3, _fhdr_c4 = st.columns([3, 1, 1, 1])
                with _fhdr_c1:
                    if st.button(
                        f"{'🔼 접기' if _fol_is_open else '🔽 펼치기'}",
                        key=f"fol_toggle_{_fol_name}",
                        use_container_width=True,
                        help=f"'{_fol_name}' 폴더 {'접기' if _fol_is_open else '펼치기'}"
                    ):
                        st.session_state[_fol_open_key] = not _fol_is_open
                        st.rerun()
                with _fhdr_c2:
                    if st.button(
                        "✏️",
                        key=f"fol_rename_btn_{_fol_name}",
                        use_container_width=True,
                        help=f"'{_fol_name}' 폴더 이름 변경"
                    ):
                        st.session_state["fav_rename_folder"] = (
                            "" if st.session_state.get("fav_rename_folder") == _fol_name else _fol_name
                        )
                        st.rerun()
                with _fhdr_c3:
                    # 폴더 이동 대상 폴더 선택 (다른 폴더로 전체이동 — 빈 폴더 삭제용)
                    pass  # 삭제 버튼
                with _fhdr_c4:
                    if _fol_name != "기본 폴더":
                        if st.button(
                            "🗑️",
                            key=f"fol_del_btn_{_fol_name}",
                            use_container_width=True,
                            help=f"'{_fol_name}' 폴더 삭제 (항목은 기본 폴더로 이동)"
                        ):
                            st.session_state["fav_folder_action"] = f"DEL_FOLDER:{_fol_name}"
                            st.rerun()

                # 이름 변경 인풋
                if st.session_state.get("fav_rename_folder") == _fol_name:
                    _rn_c1, _rn_c2 = st.columns([3, 1])
                    with _rn_c1:
                        _rename_val = st.text_input(
                            "새 이름",
                            value=_fol_name,
                            key=f"rename_input_{_fol_name}",
                            label_visibility="collapsed"
                        )
                    with _rn_c2:
                        if st.button("✔", key=f"rename_ok_{_fol_name}", use_container_width=True):
                            if _rename_val and _rename_val != _fol_name:
                                st.session_state["fav_folder_action"] = f"RENAME_FOLDER:{_fol_name}:{_rename_val}"
                                st.rerun()

                # 폴더 내 카드 (접힌 상태면 생략)
                if _fol_is_open:
                    if not _fol_items:
                        st.markdown(
                            f"<div style='font-size:0.72rem;color:{_card_sub_c};"
                            f"padding:6px 10px;font-style:italic'>빈 폴더</div>",
                            unsafe_allow_html=True
                        )
                    else:
                        # 정렬 선택 (폴더별)
                        _f_sortcol1, _f_sortcol2 = st.columns([3, 2])
                        with _f_sortcol1:
                            st.markdown(
                                f"<div style='font-size:0.72rem;color:{_card_sub_c};padding-top:3px'>"
                                f"★→검색창  🗑️→삭제  📂→이동</div>",
                                unsafe_allow_html=True
                            )
                        with _f_sortcol2:
                            _fol_sort = st.selectbox(
                                "정렬",
                                options=["저장순", "조회수순"],
                                key=f"fol_sort_{_fol_name}",
                                label_visibility="collapsed"
                            )
                        if _fol_sort == "조회수순":
                            _fol_items = sorted(_fol_items, key=lambda f: f.get("score", 0), reverse=True)

                        # 드래그&드롭 (저장순 + sortables 사용 가능 시)
                        if _fol_sort == "저장순" and _HAS_SORTABLES and len(_fol_items) > 1:
                            _drag_topics = [f.get("topic", f"항목{i}") for i, f in enumerate(_fol_items)]
                            st.markdown(
                                f"<div style='font-size:0.63rem;color:{_card_sub_c};margin-bottom:3px'>"
                                f"↕ 드래그하여 순서 변경</div>",
                                unsafe_allow_html=True
                            )
                            _sorted_fol = _sort_items(
                                items=_drag_topics,
                                direction="vertical",
                                key=f"fol_sortable_{_fol_name}"
                            )
                            if _sorted_fol != _drag_topics:
                                _omap = {t: i for i, t in enumerate(_sorted_fol)}
                                _fol_reordered = sorted(_fol_items, key=lambda f: _omap.get(f.get("topic", ""), 999))
                                # 전체 favorites 재정렬 반영
                                _other_favs = [f for f in st.session_state["favorites"] if f.get("folder") != _fol_name]
                                st.session_state["favorites"] = _other_favs + _fol_reordered
                                st.rerun()

                        for _fav in _fol_items:
                            _fav_kw    = _fav.get("keyword", "")
                            _fav_saved = _fav.get("saved", "")
                            _fav_views = _fav.get("views", "")
                            _fav_label_badge = _fav.get("label", "")
                            _fav_cur_kws = [k.strip() for k in keywords_input.replace("，", ",").split(",") if k.strip()]
                            _fav_added = _fav["topic"] in _fav_cur_kws

                            _bd_col = _card_bd_added  if _fav_added else _card_bd_normal
                            _lc_col = _card_lc_added  if _fav_added else _card_lc_normal

                            _fav_html = (
                                f"<div style='background:{_card_bg};"
                                f"border:1.5px solid {_bd_col};"
                                f"border-radius:10px;padding:7px 10px;margin-bottom:4px;"
                                f"border-left:3px solid {_lc_col};"
                                f"box-shadow:0 1px 3px rgba(0,0,0,0.08);'>"
                                f"<div style='display:flex;justify-content:space-between;"
                                f"align-items:center;margin-bottom:3px'>"
                                f"<span style='font-size:0.60rem;background:{_badge_bg_kw};"
                                f"color:{_badge_c_kw};padding:1px 7px;border-radius:8px;font-weight:700'>"
                                f"{_fav_label_badge or '⭐ 즐겨찾기'}</span>"
                                f"<span style='font-size:0.58rem;color:{_card_sub_c}'>"
                                f"{_fav_kw} · {_fav_saved}</span>"
                                f"</div>"
                                f"<div style='font-size:0.80rem;font-weight:600;"
                                f"color:{_card_title_c};line-height:1.4;margin-bottom:2px'>"
                                f"{'✓ ' if _fav_added else '⭐ '}{_fav['topic']}"
                                f"</div>"
                                + (f"<div style='font-size:0.62rem;color:{_card_sub_c};margin-top:2px'>{_fav_views}</div>" if _fav_views else "")
                                + f"</div>"
                            )
                            st.markdown(_fav_html, unsafe_allow_html=True)

                            # 버튼 행: [추가 | 이동 | 삭제]
                            _fb1, _fb2, _fb3 = st.columns([3, 2, 1])
                            with _fb1:
                                if st.button(
                                    "✓ 추가됨" if _fav_added else "＋ 검색창에 추가",
                                    key=f"fav_add_{_global_fi}",
                                    use_container_width=True,
                                    disabled=_fav_added
                                ):
                                    st.session_state["hot_topic_clicked"] = _fav["topic"]
                                    st.rerun()
                            with _fb2:
                                # 이동할 폴더 선택
                                _other_folders = [f for f in _all_folders if f != _fol_name]
                                if _other_folders:
                                    _move_dst = st.selectbox(
                                        "이동",
                                        options=["📂 이동..."] + _other_folders,
                                        key=f"fav_move_sel_{_global_fi}",
                                        label_visibility="collapsed"
                                    )
                                    if _move_dst != "📂 이동...":
                                        st.session_state["fav_action"] = f"MOVE:{_fav['topic']}:{_move_dst}"
                                        st.rerun()
                                else:
                                    st.markdown(
                                        f"<div style='font-size:0.62rem;color:{_card_sub_c};"
                                        f"padding-top:4px;text-align:center'>폴더 없음</div>",
                                        unsafe_allow_html=True
                                    )
                            with _fb3:
                                if st.button(
                                    "🗑️",
                                    key=f"fav_del_{_global_fi}",
                                    use_container_width=True,
                                    help="즐겨찾기 해제"
                                ):
                                    st.session_state["fav_action"] = f"DEL:{_fav['topic']}"
                                    st.rerun()
                            _global_fi += 1

                st.markdown("<hr style='margin:3px 0'>", unsafe_allow_html=True)

            # ── 전체 삭제 버튼 ────────────────────────────────────────
            if st.button("🗑️ 즐겨찾기 전체 삭제", key="btn_clear_favs", use_container_width=True):
                st.session_state["favorites"] = []
                _ff_reset = {k: [] for k in st.session_state.get("fav_folders", {"기본 폴더": []})}
                st.session_state["fav_folders"] = _ff_reset
                st.rerun()



        # ── 📊 상세 그래프 (전체폭 팝업) ──────────────────────────────────
        _detail_topic = st.session_state.get("detail_chart_topic", "")
        _detail_data  = st.session_state.get("detail_chart_data", {})
        if _detail_topic and _detail_data:
            from datetime import datetime as _dtnow2, timedelta as _td2
            _pal = _chart_palette(st.session_state.get("dark_mode", False))
            st.markdown(
                f"<div style='background:{_pal['bg']};"
                f"border:1.5px solid {_pal['card_border']};"
                f"border-radius:12px;padding:12px 14px;margin:8px 0 6px 0;"
                f"box-shadow:0 3px 12px rgba(79,126,248,0.12)'>"
                f"<div style='display:flex;align-items:center;gap:6px;margin-bottom:8px'>"
                f"<div style='width:4px;height:16px;"
                f"background:linear-gradient(180deg,#4f7ef8,#a78bfa);border-radius:2px'></div>"
                f"<span style='font-size:0.75rem;font-weight:800;color:#1d4ed8'>📊 7일 트렌드</span>"
                f"<span style='font-size:0.68rem;color:#64748b;font-weight:500'>{_detail_topic[:20]}{'...' if len(_detail_topic)>20 else ''}</span>"
                f"</div>",
                unsafe_allow_html=True
            )
            _dd_sp    = _detail_data.get("sparkline", [50]*7)
            _dd_sp    = (_dd_sp + [50]*7)[:7]
            _dd_today = _dtnow2.now()
            _dd_dates = [(_dd_today - _td2(days=6-i)).strftime("%m/%d(%a)") for i in range(7)]
            _dd_max   = max(_dd_sp) or 1
            _dd_min   = min(_dd_sp)
            _dd_rng   = _dd_max - _dd_min or 1
            # 전체폭 SVG 차트
            _DW, _DH  = 320, 100
            _dd_pts   = []
            for _di, _dv in enumerate(_dd_sp):
                _dx = int(_di / 6 * (_DW - 24)) + 12
                _dy = int((1 - (_dv - _dd_min) / _dd_rng) * (_DH - 24)) + 10
                _dd_pts.append((_dx, _dy, _dv))
            _dd_poly  = " ".join(f"{x},{y}" for x,y,_ in _dd_pts)
            _dd_fill  = f"12,{_DH} " + _dd_poly + f" {_dd_pts[-1][0]},{_DH}"
            _dd_trend = _dd_sp[-1] - _dd_sp[0]
            _dd_col   = _pal["rise"] if _dd_trend > 5 else (_pal["fall"] if _dd_trend < -5 else _pal["flat"])
            # 막대 차트 (바 형태)
            _bar_section = ""
            _bar_w_each  = int((_DW - 24) / 7 * 0.72)
            for _bi, ((_bx, _by, _bv), _bd) in enumerate(zip(_dd_pts, _dd_dates)):
                _bh     = int((_bv - _dd_min) / _dd_rng * (_DH - 24)) + 4
                _b_col  = _dd_col if _bv == _dd_max else _pal["bar_default"]
                _bar_x  = _bx - _bar_w_each // 2
                _bar_y  = _DH - _bh
                _bar_section += (
                    f"<rect x='{_bar_x}' y='{_bar_y}' width='{_bar_w_each}' height='{_bh}' "
                    f"rx='2' fill='{_b_col}' opacity='0.35'/>"
                    f"<text x='{_bx}' y='{_bar_y - 3}' text-anchor='middle' "
                    "font-size='8' fill='" + (_dd_col if _bv == _dd_max else _pal["grid_text"]) + "' font-weight='" + ("bold" if _bv == _dd_max else "normal") + "'>" + str(_bv) + "</text>"
                    "<text x='" + str(_bx) + "' y='" + str(_DH + 12) + "' text-anchor='middle' font-size='7.5' fill='" + _pal["x_label"] + "'>" + str(_bd) + "</text>"
                )
            # 그리드
            _dd_grid = ""
            for _gi in range(1, 5):
                _gy2 = int(_DH * _gi / 4)
                _gv2 = int(_dd_max - (_dd_max - _dd_min) * _gi / 4)
                _dd_grid += (
                    "<line x1='12' y1='" + str(_gy2) + "' x2='" + str(_DW-8) + "' y2='" + str(_gy2) + "' stroke='" + _pal["grid"] + "' stroke-width='1'/>"
                    "<text x='6' y='" + str(_gy2+3) + "' text-anchor='end' font-size='7' fill='" + _pal["grid_text"] + "'>" + str(_gv2) + "</text>"
                )
            _detail_svg = (
                f"<svg width='100%' viewBox='0 0 {_DW} {_DH+18}' style='overflow:visible;max-width:100%'>"
                f"<defs><linearGradient id='dg' x1='0' y1='0' x2='0' y2='1'>"
                "<stop offset='0%' stop-color='" + str(_dd_col) + "' stop-opacity='" + str(_pal["area_opacity_hi"]) + "'/>"
                "<stop offset='100%' stop-color='" + str(_dd_col) + "' stop-opacity='" + str(_pal["area_opacity_lo"]) + "'/>"
                f"</linearGradient></defs>"
                f"{_dd_grid}"
                f"{_bar_section}"
                f"<polygon points='{_dd_fill}' fill='url(#dg)'/>"
                f"<polyline points='{_dd_poly}' fill='none' stroke='{_dd_col}' stroke-width='2.5' stroke-linejoin='round'/>"
                + "".join(
                    "<circle cx='" + str(x) + "' cy='" + str(y) + "' r='" + ('4' if v==_dd_max else '3') + "' fill='" + str(_dd_col) + "' stroke='" + _pal["bg"] + "' stroke-width='1.5'/>"
                    for x,y,v in _dd_pts
                )
                + f"</svg>"
            )
            # 수치 요약 테이블
            _dd_change  = _dd_sp[-1] - _dd_sp[0]
            _dd_avg     = round(sum(_dd_sp) / len(_dd_sp), 1)
            _dd_arrow   = "▲" if _dd_trend > 5 else ("▼" if _dd_trend < -5 else "━")
            _dd_chcol   = "#E53935" if _dd_trend > 5 else ("#4CAF50" if _dd_trend < -5 else "#FF9800")
            _dd_chcol2  = _pal["rise"] if _dd_trend > 5 else (_pal["fall"] if _dd_trend < -5 else _pal["flat"])
            _summary_html = (
                f"<div style='display:flex;gap:8px;margin:8px 0 4px 0;flex-wrap:wrap'>"
                f"<div style='flex:1;min-width:60px;background:{_pal['summary_bg']};border-radius:8px;padding:6px 8px;text-align:center'>"
                f"<div style='font-size:0.62rem;color:{_pal['summary_text']}'>최고</div>"
                f"<div style='font-size:1.0rem;font-weight:900;color:{_pal['rise']}'>{_dd_max}</div></div>"
                f"<div style='flex:1;min-width:60px;background:{_pal['summary_bg']};border-radius:8px;padding:6px 8px;text-align:center'>"
                f"<div style='font-size:0.62rem;color:{_pal['summary_text']}'>최저</div>"
                f"<div style='font-size:1.0rem;font-weight:900;color:{_pal['fall']}'>{_dd_min}</div></div>"
                f"<div style='flex:1;min-width:60px;background:{_pal['summary_bg']};border-radius:8px;padding:6px 8px;text-align:center'>"
                f"<div style='font-size:0.62rem;color:{_pal['summary_text']}'>평균</div>"
                f"<div style='font-size:1.0rem;font-weight:900;color:#7c9ef8'>{_dd_avg}</div></div>"
                f"<div style='flex:1;min-width:60px;background:{_pal['summary_bg']};border-radius:8px;padding:6px 8px;text-align:center'>"
                f"<div style='font-size:0.62rem;color:{_pal['summary_text']}'>변화</div>"
                f"<div style='font-size:1.0rem;font-weight:900;color:{_dd_chcol2}'>{_dd_arrow}{abs(_dd_change)}</div></div>"
                f"</div>"
            )
            st.markdown(
                _summary_html + _detail_svg + "</div>",
                unsafe_allow_html=True
            )
            if st.button("✕ 상세 그래프 닫기", key="btn_close_detail", use_container_width=True):
                st.session_state["detail_chart_topic"] = ""
                st.session_state["detail_chart_data"]  = {}
                st.rerun()

        st.markdown('<hr style="margin:4px 0">', unsafe_allow_html=True)
        # ── 서브 주제 UI 끝 ────────────────────────────────────────────────


        _max_default = int(_s_max_count) if _s_max_count.isdigit() else 20
        _max_default = max(5, min(50, _max_default))
        max_count = st.slider(
            "키워드당 최대 검색 수",
            min_value=5, max_value=50, value=_max_default, step=5
        )

        # ── 정렬 방식 (복수 + 우선순위) ─────────────────────────
        _ALL_SORT_OPTS = ["조회수순","최신순","관련성순","평점순"]
        SORT_MAP = {"조회수순":"viewCount","최신순":"date","관련성순":"relevance","평점순":"rating"}
        SORT_ICON = {"조회수순":"👁","최신순":"🕐","관련성순":"🎯","평점순":"⭐"}

        # 세션에서 우선순위 리스트 복원
        if "sort_priority" not in st.session_state:
            _s_sort_norm = [_s_sort] if isinstance(_s_sort, str) and _s_sort in _ALL_SORT_OPTS else (
                [s for s in (_s_sort if isinstance(_s_sort, list) else []) if s in _ALL_SORT_OPTS] or ["조회수순"]
            )
            st.session_state["sort_priority"] = _s_sort_norm

        # 정렬 항목 추가 multiselect
        st.markdown(
            "<div style='font-size:.78rem;font-weight:700;color:#94a3b8;"
            "text-transform:uppercase;letter-spacing:.06em;margin:10px 0 4px 0'>"
            "📊 정렬 방식</div>", unsafe_allow_html=True
        )
        _cur_priority = st.session_state["sort_priority"]
        _available_add = [o for o in _ALL_SORT_OPTS if o not in _cur_priority]
        if _available_add:
            _to_add = st.selectbox(
                "정렬 추가",
                options=["— 추가 선택 —"] + _available_add,
                index=0, key="sort_add_sel",
                label_visibility="collapsed"
            )
            if _to_add != "— 추가 선택 —":
                st.session_state["sort_priority"].append(_to_add)
                st.rerun()

        # 우선순위 리스트 표시 + 위/아래/삭제 버튼
        _new_priority = list(st.session_state["sort_priority"])
        for _si, _sname in enumerate(_new_priority):
            _sc1, _sc2, _sc3, _sc4 = st.columns([4,1,1,1])
            with _sc1:
                _badge_c = {"조회수순":"#2563eb","최신순":"#059669","관련성순":"#d97706","평점순":"#7c3aed"}
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:6px;"
                    f"background:#f1f5f9;border-radius:8px;padding:4px 9px;margin:2px 0'>"
                    f"<span style='font-size:.72rem;font-weight:700;color:white;"
                    f"background:{_badge_c.get(_sname,'#64748b')};border-radius:50%;"
                    f"width:18px;height:18px;display:inline-flex;align-items:center;"
                    f"justify-content:center'>{_si+1}</span>"
                    f"<span style='font-size:.82rem;font-weight:600;color:#334155'>"
                    f"{SORT_ICON.get(_sname,'')} {_sname}</span></div>",
                    unsafe_allow_html=True
                )
            with _sc2:
                if _si > 0 and st.button("↑", key=f"sort_up_{_si}", help="우선순위 올리기"):
                    _new_priority[_si-1], _new_priority[_si] = _new_priority[_si], _new_priority[_si-1]
                    st.session_state["sort_priority"] = _new_priority
                    st.rerun()
            with _sc3:
                if _si < len(_new_priority)-1 and st.button("↓", key=f"sort_dn_{_si}", help="우선순위 내리기"):
                    _new_priority[_si], _new_priority[_si+1] = _new_priority[_si+1], _new_priority[_si]
                    st.session_state["sort_priority"] = _new_priority
                    st.rerun()
            with _sc4:
                if len(_new_priority) > 1 and st.button("✕", key=f"sort_del_{_si}", help="제거"):
                    _new_priority.pop(_si)
                    st.session_state["sort_priority"] = _new_priority
                    st.rerun()

        if not _new_priority:
            st.warning("⚠️ 정렬 방식을 1개 이상 선택하세요.")
            _new_priority = ["조회수순"]
            st.session_state["sort_priority"] = _new_priority

        sort_options    = _new_priority
        order_api_list  = [SORT_MAP[s] for s in sort_options]
        sort_option     = " + ".join(sort_options)

        # ── 업로드 기간 (복수 선택) ──────────────────────────
        st.markdown(
            "<div style='font-size:.78rem;font-weight:700;color:#94a3b8;"
            "text-transform:uppercase;letter-spacing:.06em;margin:10px 0 4px 0'>"
            "📅 업로드 기간</div>", unsafe_allow_html=True
        )
        _DATE_OPTS = ["전체","오늘","1주일","1개월","3개월","6개월","1년"]
        _date_sel = st.multiselect(
            "업로드 기간",
            options=_DATE_OPTS,
            default=["전체"],
            key="date_filter_multi",
            label_visibility="collapsed",
            help="복수 선택 시 각 기간 결과를 합산합니다."
        )
        if not _date_sel:
            _date_sel = ["전체"]
        # "전체" 포함 시 나머지 무시
        if "전체" in _date_sel:
            _date_sel = ["전체"]

        from datetime import datetime as _dt, timedelta as _td
        def _date_to_after(label):
            now = _dt.utcnow()
            _map = {"오늘":1,"1주일":7,"1개월":30,"3개월":90,"6개월":180,"1년":365}
            days = _map.get(label)
            if not days: return None
            return (now - _td(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
        _published_after_list = [_date_to_after(d) for d in _date_sel if d != "전체"]
        if not _published_after_list:
            _published_after_list = [None]

        # ── 영상 길이 (복수 선택) ────────────────────────────
        st.markdown(
            "<div style='font-size:.78rem;font-weight:700;color:#94a3b8;"
            "text-transform:uppercase;letter-spacing:.06em;margin:10px 0 4px 0'>"
            "⏱️ 영상 길이</div>", unsafe_allow_html=True
        )
        _DUR_OPTS = ["전체","단편 (4분 미만)","중편 (4-20분)","장편 (20분 초과)"]
        _DUR_API  = {"전체":None,"단편 (4분 미만)":"short","중편 (4-20분)":"medium","장편 (20분 초과)":"long"}
        _dur_sel = st.multiselect(
            "영상 길이",
            options=_DUR_OPTS,
            default=["전체"],
            key="dur_filter_multi",
            label_visibility="collapsed",
            help="복수 선택 시 각 길이 범위 결과를 합산합니다."
        )
        if not _dur_sel:
            _dur_sel = ["전체"]
        if "전체" in _dur_sel:
            _dur_sel = ["전체"]
        _dur_api_list = [_DUR_API[d] for d in _dur_sel]  # None이면 필터 없음

        video_type = st.radio(
            "📹 영상 종류",
            options=["전체", "동영상", "쇼츠"],
            index=0,
            horizontal=True,
            help="전체: 모든 영상 | 동영상: 60초 초과 일반 영상 | 쇼츠: 60초 이하 세로형 영상"
        )

        fetch_transcript = st.checkbox(
            "📜 자막(대본) 가져오기",
            value=True,
            help="체크 시 각 영상의 자막을 가져와 키워드 추출과 요약을 수행합니다. 영상이 많으면 시간이 걸립니다."
        )

        # ── 자막 없는 영상 처리 방식 설정 ─────────────────────────
        use_whisper          = False
        use_gemini           = False
        openai_api_key_input = ""
        gemini_api_key_input = ""

        if fetch_transcript:
            st.caption("📡 자막 없는 영상 처리 방식")
            _stt_mode = st.radio(
                "자막 없는 영상 처리",
                options=[
                    "사용 안 함",
                    "🤖 Gemini 분석",
                    "🎙️ Whisper 변환",
                    "🤖→🎙️ Gemini 우선 + Whisper 폴백"
                ],
                index=0,
                label_visibility="collapsed",
                help="Gemini: YouTube URL만 전달 → 클라우드·로컬 모두 OK\n"
                     "Whisper: 오디오 다운로드 → 로컬 PC 전용\n"
                     "Gemini+Whisper: Gemini 실패 시 Whisper로 자동 재시도"
            )

            use_gemini  = "Gemini"  in _stt_mode
            use_whisper = "Whisper" in _stt_mode

            # ── Gemini 설치 사전 체크 ──────────────────────────────────
            if use_gemini:
                # google-genai 설치 여부 확인 (YouTube URL 분석의 필수 라이브러리)
                try:
                    from google import genai as _chk_genai  # noqa
                    _genai_installed = True
                except ImportError:
                    _genai_installed = False

                if not _genai_installed:
                    st.error(
                        "📦 **google-genai 미설치 — Gemini 분석 불가**\n\n"
                        "YouTube URL 직접 분석은 새 SDK만 지원합니다.\n"
                        "**지금 바로 PowerShell에서 실행하세요:**"
                    )
                    st.code("pip install google-genai", language="bash")
                    st.caption("설치 후 앱을 재시작하면 자동으로 Gemini 분석이 활성화됩니다.")

            # ── Gemini 키 UI ─────────────────────────────────────────
            if use_gemini:
                if _s_gemini_key:
                    gemini_api_key_input = _s_gemini_key
                    _key_ok = _s_gemini_key.startswith("AIza") and len(_s_gemini_key) > 30
                    if _key_ok:
                        st.caption(f"✅ GEMINI_API_KEY 자동 로드됨 ({_s_gemini_key[:8]}...)")
                        st.caption("💡 Gemini 2.0 Flash · 무료 티어 가능 · REST API 직접 호용")
                    else:
                        st.caption(f"⚠️ GEMINI_API_KEY 형식 이상 ({_s_gemini_key[:8]}...)")
                        st.warning(
                            "🔑 **키 형식 오류**: Google AI Studio 키는 `AIzaSy...` 형식이어야 합니다.\n"
                            "YouTube API 키를 잘못 입력했을 수 있습니다.\n"
                            "🔗 https://aistudio.google.com/app/apikey 에서 Gemini 키 발급"
                        )
                else:
                    gemini_api_key_input = st.text_input(
                        "🔑 Gemini API Key",
                        value="",
                        type="password",
                        placeholder="AIzaSy...",
                        help="https://aistudio.google.com/app/apikey 에서 무료 발급\n"
                             "또는 secrets.toml에 GEMINI_API_KEY 추가"
                    )
                    if not gemini_api_key_input:
                        st.caption("⚠️ GEMINI_API_KEY를 입력하거나 secrets.toml에 추가하세요.")
                    else:
                        st.caption("✅ Gemini API 키 설정됨")

            # ── Whisper 키 UI ────────────────────────────────────────
            if use_whisper:
                import os as _os_w
                _is_cloud = (
                    _os_w.environ.get("STREAMLIT_SHARING_MODE") == "1"
                    or "streamlit.app" in _os_w.environ.get("HOSTNAME", "")
                    or _os_w.path.exists("/mount/src")
                )
                if _is_cloud:
                    st.warning("⚠️ Whisper는 Streamlit Cloud에서 IP 차단으로 동작하지 않습니다.\n"
                               "✅ 로컬 PC에서만 정상 작동합니다.")
                if _s_openai_key:
                    openai_api_key_input = _s_openai_key
                    st.caption("✅ OPENAI_API_KEY 자동 로드됨")
                    st.caption("💡 비용: ~$0.006/분 · 25분 이하 영상 권장")
                else:
                    openai_api_key_input = st.text_input(
                        "🔑 OpenAI API Key",
                        value="",
                        type="password",
                        placeholder="sk-...",
                        help="https://platform.openai.com/api-keys 에서 발급\n"
                             "또는 secrets.toml에 OPENAI_API_KEY 추가"
                    )
                    if not openai_api_key_input:
                        st.caption("⚠️ OPENAI_API_KEY를 입력하거나 secrets.toml에 추가하세요.")
                    else:
                        st.caption("✅ Whisper API 키 설정됨")

        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
        search_btn = st.button("🚀 검색 시작", use_container_width=True, type="primary")

        # ✅ FIX: Secrets 로드 현황 디버그 패널
        with st.expander("🔧 Secrets 로드 현황 (클릭)", expanded=False):
            st.caption(f"YOUTUBE_API_KEY: {'✅ 로드됨' if _s_api_key else '❌ 없음'}")
            st.caption(f"OPENAI_API_KEY: {'✅ 로드됨 (' + _s_openai_key[:8] + '...)' if _s_openai_key else '❌ 없음 → Whisper 사용 시 필요'}")
            st.caption(f"GEMINI_API_KEY: {'✅ 로드됨 (' + _s_gemini_key[:8] + '...)' if _s_gemini_key else '❌ 없음 → Gemini 사용 시 필요'}")
            st.caption(f"GSHEET_SHARE_EMAIL: {'✅ ' + _s_email if _s_email else '❌ 없음'}")
            st.caption(f"GSHEET_EXISTING_ID: {'✅ 설정됨' if _s_existing else '❌ 없음'}")
            st.caption(f"gcp_service_account: {'✅ 로드됨' if _s_gcp_creds else '❌ 없음'}")
            if not _s_gemini_key:
                st.info("💡 GEMINI_API_KEY 설정 방법:\n```\nGEMINI_API_KEY = \"AIzaSy...\"\n```\nhttps://aistudio.google.com/app/apikey 에서 무료 발급")
            if not _s_openai_key:
                st.info("💡 OPENAI_API_KEY 설정 방법:\n```\nOPENAI_API_KEY = \"sk-proj-...\"\n```")

        st.markdown("---")
        st.markdown(
            "<div style='display:flex;align-items:center;gap:6px;margin:8px 0 4px 0'>"
            "<div style='width:3px;height:14px;background:linear-gradient(180deg,#22c55e,#16a34a);"
            "border-radius:2px'></div>"
            "<span style='font-size:0.73rem;font-weight:700;color:#5c6480;"
            "text-transform:uppercase;letter-spacing:0.05em'>📊 Google Sheets</span></div>",
            unsafe_allow_html=True
        )

        use_gsheet = st.checkbox(
            "Google Sheets 자동 업로드",
            value=False,
            disabled=not HAS_GSHEET
        )
        if not HAS_GSHEET:
            st.caption("⚠️ `pip install gspread google-auth` 필요")

        # ── credentials.json 자동 탐색 ─────────────────────────
        import os as _os
        _cred_search_paths = [
            _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "credentials.json"),
            _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".streamlit", "credentials.json"),
            _os.path.join(_os.getcwd(), "credentials.json"),
            _os.path.join(_os.getcwd(), ".streamlit", "credentials.json"),
        ]
        _auto_creds_path = None
        for _p in _cred_search_paths:
            if _os.path.exists(_p):
                _auto_creds_path = _p
                break

        credentials_file = None
        share_email      = _s_email
        existing_id      = _s_existing

        if use_gsheet:
            # ── 우선순위: ①Secrets gcp_service_account → ②로컬 파일 → ③수동 업로드
            _all_auto = bool((_s_gcp_creds or _auto_creds_path) and _s_existing)

            if _s_gcp_creds:
                # ✅ Streamlit Cloud Secrets에서 자동 로드됨
                st.success("✅ Secrets에서 Google 인증 자동 로드 완료")
                if _s_email:
                    st.caption(f"📧 공유 이메일: {_s_email}")
                if _s_existing:
                    _eid_display = _s_existing[:20] + "..." if len(_s_existing) > 20 else _s_existing
                    st.caption(f"📊 시트 ID: {_eid_display}")
                _show_manual = st.checkbox("⚙️ 설정 수동 변경", value=False, key="gsheet_manual")
            elif _all_auto:
                # ✅ 로컬 파일에서 자동 로드됨
                st.success("✅ 모든 설정 자동 로드 완료")
                st.caption(f"🔑 credentials: {_os.path.basename(_auto_creds_path)}")
                if _s_email:
                    st.caption(f"📧 공유 이메일: {_s_email}")
                _eid_display = _s_existing[:20] + "..." if len(_s_existing) > 20 else _s_existing
                st.caption(f"📊 시트 ID: {_eid_display}")
                _show_manual = st.checkbox("⚙️ 설정 수동 변경", value=False, key="gsheet_manual")
            else:
                _show_manual = True

            if _show_manual if not _s_gcp_creds else st.session_state.get("gsheet_manual", False):
                # credentials.json
                if _s_gcp_creds:
                    st.info("✅ Secrets(gcp_service_account)에서 인증 정보 로드됨")
                elif _auto_creds_path:
                    st.info(f"✅ credentials.json 자동 감지: {_os.path.basename(_auto_creds_path)}")
                    if st.checkbox("다른 파일로 교체", value=False, key="replace_creds"):
                        credentials_file = st.file_uploader(
                            "credentials.json 업로드",
                            type=["json"],
                            help="Google Cloud 서비스 계정 JSON 키 파일"
                        )
                else:
                    st.warning("⚠️ credentials.json 없음\n\nStreamlit Cloud Secrets에 [gcp_service_account] 섹션을 추가하거나 파일을 업로드하세요")
                    credentials_file = st.file_uploader(
                        "credentials.json 업로드",
                        type=["json"],
                        help="Google Cloud 서비스 계정 JSON 키 파일"
                    )

                # 이메일
                share_email = st.text_input(
                    "📧 공유할 이메일",
                    value=_s_email,
                    placeholder="yourname@gmail.com",
                    help="secrets.toml의 GSHEET_SHARE_EMAIL에 저장하면 자동 입력됩니다."
                )

                # 스프레드시트 ID
                existing_id = st.text_input(
                    "📊 스프레드시트 ID",
                    value=_s_existing,
                    placeholder="1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms",
                    help="secrets.toml의 GSHEET_EXISTING_ID에 저장하면 자동 입력됩니다."
                )
                if not _s_existing:
                    st.caption("💡 스프레드시트 URL에서 /d/ 뒤의 문자열을 secrets.toml의 GSHEET_EXISTING_ID 에 저장하면 다음부터 자동 입력됩니다!")

    # ================================================================
    # 검색 실행
    # ================================================================
    if search_btn:
        if not api_key:
            st.error("❌ YouTube API 키를 입력해주세요.")
            st.stop()
        if not keywords_input.strip():
            st.error("❌ 검색 키워드를 입력해주세요.")
            st.stop()

        keywords = [kw.strip() for kw in keywords_input.replace("，",",").split(",") if kw.strip()]

        # ── 검색 기록 저장 (최근 5개 유지, 중복 제거) ──────────────
        _hist_kw = keywords_input.strip()
        _hist    = st.session_state.get("search_history", [])
        _hist    = [h for h in _hist if h != _hist_kw]   # 중복 제거
        _hist.insert(0, _hist_kw)                          # 맨 앞에 추가
        st.session_state["search_history"] = _hist[:5]    # 최대 5개


        # credentials.json 로드: ①Secrets gcp → ②수동 업로드 → ③자동 탐색 순서
        creds_dict = None
        if use_gsheet:
            # ✅ 우선순위 1: Streamlit Cloud Secrets gcp_service_account
            if _s_gcp_creds:
                creds_dict = _s_gcp_creds
            elif credentials_file:
                try:
                    creds_dict = json.load(credentials_file)
                except:
                    st.error("❌ 업로드한 credentials.json 파일을 읽을 수 없습니다.")
                    st.stop()
            elif _auto_creds_path:
                try:
                    with open(_auto_creds_path, "r", encoding="utf-8") as _f:
                        creds_dict = json.load(_f)
                except Exception as _e:
                    st.error(f"❌ credentials.json 자동 로드 실패: {_e}")
                    st.stop()
            else:
                st.error("❌ credentials.json 없음\n\n해결 방법:\n1. Streamlit Cloud → Settings → Secrets에 [gcp_service_account] 추가\n2. 또는 파일 직접 업로드")
                st.stop()

        # 진행 상태
        progress_bar = st.progress(0)
        status_text  = st.empty()

        all_results   = {}
        all_videos_flat = []
        total_steps   = len(keywords)
        _whisper_errors = []  # Whisper 오류 수집용

        for ki, kw in enumerate(keywords):
            status_text.info(f"🔍 [{kw}] 검색 중... ({ki+1}/{total_steps})")

            # 1) 검색 (정렬 우선순위 × 업로드기간 × 영상길이 조합 병합, 중복 제거)
            _seen_ids   = {}   # video_id → set(sort_source 레이블)
            # 정렬 우선순위 순서대로 순회 (1순위 결과가 앞에 오도록)
            for _oi, _oapi in enumerate(order_api_list):
                _sort_label_tag = sort_options[_oi]  # 예: "조회수순"
                for _pub_after in _published_after_list:
                    for _dur_api in _dur_api_list:
                        _ids_part, err = search_youtube(
                            api_key, kw, max_count, _oapi, video_type,
                            published_after=_pub_after, dur_filter=_dur_api
                        )
                        if err:
                            st.error(f"❌ 오류: {err}")
                            st.stop()
                        for _vid in (_ids_part or []):
                            if _vid not in _seen_ids:
                                _seen_ids[_vid] = set()
                            _seen_ids[_vid].add(_sort_label_tag)
            video_ids = list(_seen_ids.keys())
            # sort_sources 매핑 저장 (video 객체에 나중에 붙임)
            _vid_sort_sources = {vid: sorted(tags) for vid, tags in _seen_ids.items()}
            if not video_ids:
                st.warning(f"⚠️ [{kw}] 검색 결과가 없습니다.")
                all_results[kw] = []
                progress_bar.progress((ki+1)/total_steps)
                continue

            # 2) 상세 정보
            status_text.info(f"📊 [{kw}] 영상 상세 정보 수집 중...")
            videos = fetch_video_details(api_key, video_ids)

            # 2-1) 영상 종류 필터링 (쇼츠 / 동영상)
            if video_type == "쇼츠":
                videos = [v for v in videos if is_shorts(v)]
                if not videos:
                    st.warning(f"⚠️ [{kw}] 쇼츠 영상이 없습니다.")
                    all_results[kw] = []
                    progress_bar.progress((ki+1)/total_steps)
                    continue
            elif video_type == "동영상":
                videos = [v for v in videos if not is_shorts(v)]
                if not videos:
                    st.warning(f"⚠️ [{kw}] 일반 동영상이 없습니다.")
                    all_results[kw] = []
                    progress_bar.progress((ki+1)/total_steps)
                    continue
            # max_count 초과 제거 (pre-filter 감안해 더 많이 가져왔으므로)
            videos = videos[:max_count]

            # 3) 구독자
            status_text.info(f"👥 [{kw}] 구독자 수 수집 중...")
            videos = fetch_subscribers(api_key, videos)

            # 4) 자막 → Gemini / Whisper 폴백
            if fetch_transcript:
                for vi, v in enumerate(videos):
                    _mode_note = ""
                    if use_gemini and gemini_api_key_input:
                        _mode_note = " 🤖Gemini 대기중"
                    elif use_whisper and openai_api_key_input:
                        _mode_note = " 🎙️Whisper 대기중"
                    status_text.info(f"📜 [{kw}] 자막 수집 중... ({vi+1}/{len(videos)}) - {v['title'][:30]}...{_mode_note}")

                    # ① 먼저 유튜브 자막 시도
                    raw = get_transcript(v["videoId"])
                    _no_caption = (
                        not raw
                        or raw.startswith("자막 없음")
                        or raw.startswith("youtube-transcript")
                        or raw.startswith("[Whisper")
                    )

                    # ② 자막 없을 때 Gemini 시도
                    if _no_caption and use_gemini and gemini_api_key_input:
                        # ── 영상 길이 사전 체크 ──────────────────────────────────
                        # Gemini 무료 티어(Tier 1): 25분↑ 영상은 토큰 초과 가능성 높음
                        # 초당 약 300 토큰 × 25분(1500초) ≈ 450,000 토큰 → 제한 근접
                        _vid_sec = parse_duration_seconds(v.get("duration", "0:00"))
                        _GEMINI_MAX_SEC = 25 * 60  # 25분 = 1500초

                        if _vid_sec > _GEMINI_MAX_SEC:
                            # 25분 초과 → Gemini 스킵, Whisper로 자동 전환 시도
                            _skip_msg = (
                                f"[Gemini 스킵] {v['title'][:35]} "
                                f"— 영상 길이 {v.get('duration','?')} (25분 초과, 토큰 한도 위험) "
                                f"→ {'Whisper로 전환' if use_whisper and openai_api_key_input else '자막 없음 처리'}"
                            )
                            _whisper_errors.append(f"• {v['title'][:35]} [Gemini 스킵]: 영상 길이 {v.get('duration','?')} 25분 초과")
                            status_text.info(f"⏭️ [{kw}] Gemini 스킵 (25분 초과): {v['title'][:25]}...")
                            gemini_result = None
                        else:
                            status_text.info(
                                f"🤖 [{kw}] Gemini 분석 중... ({vi+1}/{len(videos)}) "
                                f"- {v['title'][:25]}..."
                            )
                            gemini_result = gemini_analyze_video(v["videoId"], gemini_api_key_input)

                        if gemini_result and not gemini_result.startswith("[Gemini 오류"):
                            raw = gemini_result
                            _no_caption = False
                        else:
                            err_msg = gemini_result or "[Gemini 오류] 알 수 없는 오류"
                            _whisper_errors.append(f"• {v['title'][:35]} [Gemini 실패]: {err_msg[:120]}")
                            # 항상 raw를 "자막 없음"으로 설정 (오류 텍스트가 transcript에 저장되지 않도록)
                            raw = "자막 없음 (Gemini 실패)"
                            # ✅ 개별 st.warning 제거 → 마지막에 요약 패널로 표시

                    # ③ 자막 없을 때(Gemini도 실패 or 미사용) Whisper 시도
                    if _no_caption and use_whisper and openai_api_key_input:
                        status_text.info(
                            f"🎙️ [{kw}] Whisper 변환 중... ({vi+1}/{len(videos)}) "
                            f"- {v['title'][:25]}... (수 분 소요될 수 있습니다)"
                        )
                        whisper_result = whisper_transcribe(v["videoId"], openai_api_key_input)
                        if whisper_result and not whisper_result.startswith("[Whisper 오류]"):
                            raw = f"[🎙️ Whisper 변환]\n{whisper_result}"
                        else:
                            err_msg = whisper_result or "[Whisper 오류] 알 수 없는 오류"
                            _whisper_errors.append(f"• {v['title'][:35]} [Whisper]: {err_msg}")
                            st.warning(f"🎙️ Whisper 변환 실패: {v['title'][:30]}\n→ {err_msg}")
                            raw = f"자막 없음 (Whisper 실패: {err_msg[:60]})"

                    # 타임스탬프 제거 후 저장 (오류/빈 값 제외)
                    if raw and is_valid_transcript(raw):
                        v["transcript"] = clean_transcript(raw)
                    else:
                        v["transcript"] = raw

                    v["keywords"] = extract_keywords(
                        v["transcript"] + " " + v["description"] + " " + " ".join(v["tags"])
                    )
                    v["summary"]  = summarize_text(
                        v["transcript"] if len(v.get("transcript","")) > 100 else v["description"]
                    )

            # 5) 배지 & 순위 & 정렬 출처 태그
            for rank_i, v in enumerate(videos, 1):
                v["rank"]  = rank_i
                v["badge"] = get_badge(rank_i, v["viewCount"])
                v["sort_sources"] = _vid_sort_sources.get(v["videoId"], [])

            all_results[kw] = videos
            all_videos_flat.extend(videos)
            progress_bar.progress((ki+1)/total_steps)

        status_text.success("✅ 분석 완료!")
        progress_bar.progress(1.0)

        # 채널 통계
        channel_stats = build_channel_stats(all_videos_flat)

        # ── 세션에 저장 ─────────────────────────────────────
        st.session_state["results"]       = all_results
        st.session_state["channel_stats"] = channel_stats
        st.session_state["sort_label"]    = sort_option
        st.session_state["filter_summary"] = {
            "sort":  sort_options,
            "date":  _date_sel,
            "dur":   _dur_sel,
            "vtype": video_type,
            "max":   max_count,
            "kws":   keywords,
        }
        st.session_state["creds_dict"]    = creds_dict
        st.session_state["share_email"]   = share_email
        st.session_state["existing_id"]   = existing_id
        st.session_state["use_gsheet"]    = use_gsheet
        st.session_state["whisper_errors"] = _whisper_errors

        # ✅ FIX: 구글시트 자동 업로드 (체크박스 ON 시 검색 완료 후 즉시 실행)
        if use_gsheet and creds_dict:
            with st.spinner("☁️ 구글 스프레드시트 자동 업로드 중..."):
                _auto_ok, _auto_result = upload_to_gsheet(
                    all_results, channel_stats, sort_option,
                    credentials_dict=creds_dict,
                    share_email=share_email,
                    existing_id=existing_id
                )
            if _auto_ok:
                st.success("✅ 구글 스프레드시트 자동 업로드 완료!")
                st.markdown(f"🔗 [스프레드시트 열기]({_auto_result})")
                st.session_state["gsheet_url"] = _auto_result
            else:
                st.error(f"❌ 자동 업로드 실패: {_auto_result}")
        elif use_gsheet and not creds_dict:
            st.error("❌ 자동 업로드 실패: credentials 없음. Secrets의 [gcp_service_account]를 확인하세요.")

    # ================================================================
    # 결과 표시
    # ================================================================
    if "results" not in st.session_state or not st.session_state["results"]:
        if not search_btn:
            st.info("👈 왼쪽 사이드바에서 API 키와 검색 키워드를 입력하고 **검색 시작** 버튼을 눌러주세요.")
            with st.expander("📖 사용 방법 & 준비 사항"):
                st.markdown("""
**① 필수 라이브러리 설치**
```bash
pip install streamlit requests youtube-transcript-api openpyxl gspread google-auth pytrends
```

> 💡 **pytrends**: Google Trends 연관검색어(급상승) 수집에 필요합니다.  
> 미설치 시 서브주제 검색에서 Google Trends 결과(30% 슬롯)가 YouTube 자동완성으로 대체됩니다.

**② YouTube API 키 준비**
1. [Google Cloud Console](https://console.cloud.google.com) 접속
2. YouTube Data API v3 활성화
3. 사용자 인증정보 → API 키 생성

**③ Google Sheets 업로드 (선택)**
1. Google Sheets API + Google Drive API 활성화
2. 서비스 계정 생성 → JSON 키 다운로드
3. 사이드바에서 credentials.json 업로드

**④ 앱 실행**
```bash
streamlit run youtube_web_app.py
```
                """)
        st.stop()

    all_results   = st.session_state["results"]
    channel_stats = st.session_state["channel_stats"]
    sort_label    = st.session_state.get("sort_label", "조회수순")

    # ✅ FIX: Whisper 오류 요약 표시
    _we = st.session_state.get("whisper_errors", [])
    if _we:
        _gemini_errs  = [e for e in _we if "Gemini" in e]
        _whisper_errs = [e for e in _we if "Whisper" in e]

        # ── Gemini 오류 패널 ──
        if _gemini_errs:
            with st.expander(f"🤖 Gemini 분석 실패 {len(_gemini_errs)}개 (클릭하여 원인 확인)", expanded=True):
                _first = _gemini_errs[0]
                if "v1beta" in _first or "not found for API version" in _first or "oldSDK" in _first or "newSDK" in _first or "REST" in _first:
                    st.error(f"⚙️ **Gemini 분석 실패 — 상세 오류:**\n```\n{_first[:300]}\n```")
                    st.info("💡 **확인사항**: 아래를 순서대로 점검하세요")
                    st.markdown(
                        "1. **google-genai 최신 버전으로 업그레이드**:\n"
                        "   ```\n   pip install --upgrade google-genai\n   ```\n"
                        "2. **앱 완전 재시작** (Ctrl+C 후 bat 파일 재실행)\n"
                        "3. **Gemini API 키 재확인**: https://aistudio.google.com/app/apikey\n"
                        "   - secrets.toml의 `GEMINI_API_KEY` 값이 AI Studio 키인지 확인\n"
                        "   - YouTube API 키와 혼동 금지 (둘 다 `AIzaSy`로 시작)"
                    )
                elif "Gemini 스킵" in _first or "25분 초과" in _first:
                    st.warning("⏭️ **영상 길이 초과 (25분 초과 스킵)**")
                    st.info(
                        "💡 Gemini 무료 티어(Tier 1)는 25분 초과 영상 분석 시 토큰 한도에 근접합니다.\n"
                        "• Gemini 전용 **\"Gemini 우선 + Whisper 폴백\"** 모드를 선택하면 \n"
                        "  Whisper가 자동 대체 분석합니다.\n"
                        "• 또는 해당 영상은 검색에서 제외하세요 (최대 검색 수 조정)."
                    )
                elif "token" in _first.lower() or "input token" in _first.lower() or "maximum" in _first.lower():
                    st.warning("📊 **토큰 한도 초과 (Token Limit Exceeded)**")
                    st.info(
                        "💡 영상이 너무 길어 Gemini가 토큰 한도를 초과했습니다.\n"
                        "• **무료 티어 기준**: gemini-2.5-flash 최대 입력 토큰 1M (25분 영상 ≈ 450K 토큰)\n"
                        "• **해결책**: 영상 길이 20분 이내로 필터링하거나 Whisper 병행 사용 권장"
                    )
                elif "직접 분석 미지원" in _first or "not support" in _first.lower():
                    st.warning("📹 **영상 URL 직접 분석 미지원** — Gemini가 접근 불가한 유형의 영상")
                    st.info("ℹ️ 일부 영상은 지역제한/비공개/라이브 제한으로 Gemini URL 분석이 블록됩니다.")
                elif "미설치" in _first or "pip install" in _first or "ImportError" in _first:
                    st.error("📦 **Gemini SDK 미설치**")
                    st.code("pip install google-genai google-generativeai", language="bash")
                elif "API 키" in _first or "인증" in _first or "not valid" in _first or "API_KEY_INVALID" in _first:
                    st.error("🔑 **GEMINI_API_KEY 인증 실패**")
                    st.info(
                        "**확인 사항:**\n"
                        "1. secrets.toml의 `GEMINI_API_KEY`가 **Google AI Studio** 키인지 확인\n"
                        "2. YouTube API 키와 혼동하지 않도록 주의 (둘 다 AIzaSy로 시작함)\n"
                        "3. 키 발급: https://aistudio.google.com/app/apikey"
                    )
                elif "할당량" in _first or "quota" in _first.lower() or "429" in _first:
                    st.warning("⏳ **Gemini API 할당량 초과**\n잠시 후 재시도하거나 Gemini Pro 플랜 업그레이드가 필요합니다.")
                else:
                    st.warning("⚠️ Gemini 오류 발생 — 상세 내용:")
                for _e in _gemini_errs:
                    st.caption(_e)

        # ── Whisper 오류 패널 ──
        if _whisper_errs:
            import os as _os_r
            _is_cloud_r = (
                _os_r.environ.get("STREAMLIT_SHARING_MODE") == "1"
                or "streamlit.app" in _os_r.environ.get("HOSTNAME", "")
                or _os_r.path.exists("/mount/src")
            )
            _all_403 = all("403" in e or "Forbidden" in e for e in _whisper_errs)
            if _is_cloud_r or _all_403:
                st.warning(
                    f"🎙️ **Whisper 변환 불가 ({len(_whisper_errs)}개 영상)**  \n"
                    "Streamlit Cloud에서는 YouTube IP 차단으로 오디오 다운로드가 불가합니다.  \n"
                    "✅ **로컬 PC**에서 실행하면 정상 작동합니다."
                )
            else:
                with st.expander(f"⚠️ Whisper 변환 실패 {len(_whisper_errs)}개 (클릭하여 원인 확인)", expanded=True):
                    for _e in _whisper_errs:
                        st.caption(_e)

    # ── 요약 통계 ─────────────────────────────────────────────
    total_videos   = sum(len(v) for v in all_results.values())
    total_channels = len(channel_stats)
    total_views    = sum(v["viewCount"] for vs in all_results.values() for v in vs)
    has_transcript = sum(1 for vs in all_results.values()
                         for v in vs if is_valid_transcript(v.get("transcript", "")))

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="metric-card"><div class="value">{total_videos}</div><div class="label">🎬 분석 영상</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="metric-card"><div class="value">{total_channels}</div><div class="label">📺 채널 수</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="metric-card"><div class="value">{fmt(total_views)}</div><div class="label">👁️ 총 조회수</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="metric-card"><div class="value">{has_transcript}</div><div class="label">📜 자막 있음</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 다운로드 & 구글 시트 버튼 ────────────────────────────
    st.markdown("### 💾 결과 내보내기")
    btn_col1, btn_col2, btn_col3, btn_col4 = st.columns(4)

    now_str  = datetime.now().strftime("%Y%m%d_%H%M")
    slug_kw  = list(all_results.keys())[0][:12].replace(" ","_")
    base_name = f"youtube_{slug_kw}_{now_str}"

    txt_data  = build_txt(all_results, channel_stats, sort_label).encode("utf-8")
    json_data = build_json(all_results, channel_stats).encode("utf-8")
    xlsx_data = save_xlsx_bytes(all_results, channel_stats) if HAS_XLSX else None

    with btn_col1:
        st.download_button(
            label="📄 TXT 다운로드",
            data=txt_data,
            file_name=f"{base_name}.txt",
            mime="text/plain",
            use_container_width=True,
            key=f"dl_txt_{now_str}"
        )
    with btn_col2:
        st.download_button(
            label="🔢 JSON 다운로드",
            data=json_data,
            file_name=f"{base_name}.json",
            mime="application/json",
            use_container_width=True,
            key=f"dl_json_{now_str}"
        )
    with btn_col3:
        if xlsx_data:
            st.download_button(
                label="📊 Excel 다운로드",
                data=xlsx_data,
                file_name=f"{base_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_xlsx_{now_str}"
            )
        else:
            st.button("📊 Excel (openpyxl 미설치)", disabled=True, use_container_width=True)
    with btn_col4:
        # creds_dict 가 세션에 있거나, 자동탐색 파일이 있으면 버튼 활성화
        import os as _os2
        _cred_paths2 = [
            _os2.path.join(_os2.path.dirname(_os2.path.abspath(__file__)), "credentials.json"),
            _os2.path.join(_os2.path.dirname(_os2.path.abspath(__file__)), ".streamlit", "credentials.json"),
            _os2.path.join(_os2.getcwd(), "credentials.json"),
            _os2.path.join(_os2.getcwd(), ".streamlit", "credentials.json"),
        ]
        _found_cred2 = next((p for p in _cred_paths2 if _os2.path.exists(p)), None)
        # ✅ Secrets gcp_service_account 도 업로드 가능 조건에 포함
        _can_upload = HAS_GSHEET and (
            st.session_state.get("creds_dict") or _found_cred2 or _s_gcp_creds
        )

        if _can_upload:
            if st.button("☁️ Google Sheets 업로드", use_container_width=True, type="primary"):
                # creds_dict: 세션 우선, 없으면 파일에서 로드
                _cd = st.session_state.get("creds_dict")
                # ✅ Secrets gcp_service_account 우선 사용
                if not _cd and _s_gcp_creds:
                    _cd = _s_gcp_creds
                if not _cd and _found_cred2:
                    try:
                        with open(_found_cred2, "r", encoding="utf-8") as _ff:
                            _cd = json.load(_ff)
                    except Exception as _e:
                        st.error(f"❌ credentials.json 로드 실패: {_e}")
                        st.stop()

                _email2   = st.session_state.get("share_email") or _s_email or None
                _eid2     = st.session_state.get("existing_id") or _s_existing or None

                if not _cd:
                    st.error("❌ credentials.json 파일이 없습니다. 앱 폴더에 넣어주세요.")
                else:
                    # 업로드 전 데이터 확인
                    _total_rows = sum(len(v) for v in all_results.values())
                    st.info(f"📊 업로드할 영상 수: {_total_rows}개, 키워드: {list(all_results.keys())}")
                    with st.spinner("📊 구글 스프레드시트 업로드 중..."):
                        ok, result = upload_to_gsheet(
                            all_results, channel_stats, sort_label,
                            credentials_dict=_cd,
                            share_email=_email2,
                            existing_id=_eid2
                        )
                    if ok:
                        st.success("✅ 업로드 완료!")
                        st.markdown(f"🔗 [스프레드시트 열기]({result})")
                        st.session_state["gsheet_url"] = result
                    else:
                        st.error(result)
        elif not HAS_GSHEET:
            st.button("☁️ Google Sheets (라이브러리 미설치)", disabled=True, use_container_width=True)
            st.caption("pip install gspread google-auth")
        else:
            st.button("☁️ Google Sheets (credentials 없음)", disabled=True, use_container_width=True)
            st.caption("앱 폴더에 credentials.json을 넣어주세요")

    if st.session_state.get("gsheet_url"):
        st.info(f"📊 마지막 업로드된 시트: {st.session_state['gsheet_url']}")

    st.markdown("---")

    # ── 검색 조건 요약 패널 ───────────────────────────────────
    _fs = st.session_state.get("filter_summary", {})
    if _fs:
        _SORT_COLOR = {"조회수순":"#2563eb","최신순":"#059669","관련성순":"#d97706","평점순":"#7c3aed"}
        _DATE_COLOR = {"전체":"#64748b","오늘":"#dc2626","1주일":"#ea580c",
                       "1개월":"#d97706","3개월":"#059669","6개월":"#0891b2","1년":"#7c3aed"}
        _DUR_COLOR  = {"전체":"#64748b","단편 (4분 미만)":"#2563eb",
                       "중편 (4-20분)":"#059669","장편 (20분 초과)":"#7c3aed"}

        def _tag(label, color, prefix=""):
            return (
                f"<span style='display:inline-flex;align-items:center;gap:3px;"
                f"background:{color}18;border:1px solid {color}55;border-radius:12px;"
                f"padding:3px 10px;font-size:.78rem;font-weight:600;color:{color};"
                f"margin:2px 3px'>{prefix}{label}</span>"
            )

        _sort_tags  = "".join(_tag(f"{i+1}순위 {s}", _SORT_COLOR.get(s,'#64748b'), "")
                              for i, s in enumerate(_fs.get("sort",[])))
        _date_tags  = "".join(_tag(d, _DATE_COLOR.get(d,'#64748b'), "📅 ")
                              for d in _fs.get("date", ["전체"]))
        _dur_tags   = "".join(_tag(d, _DUR_COLOR.get(d,'#64748b'), "⏱ ")
                              for d in _fs.get("dur", ["전체"]))
        _kw_tags    = "".join(_tag(k, "#1e40af", "🔍 ") for k in _fs.get("kws", []))

        _vtype = _fs.get("vtype","전체")
        _vtype_c = {"전체":"#64748b","동영상":"#2563eb","쇼츠":"#dc2626"}.get(_vtype,"#64748b")
        _vtype_tag = _tag(_vtype, _vtype_c, "📹 ")
        _max_tag   = _tag(f"최대 {_fs.get('max',20)}개", "#475569", "📦 ")

        _dm = st.session_state.get("dark_mode", False)
        _panel_bg  = "#1a1c30" if _dm else "#f8faff"
        _panel_bd  = "#2e3157" if _dm else "#dde5f5"
        _title_c   = "#8a90c8" if _dm else "#64748b"

        st.markdown(f"""
<div style='background:{_panel_bg};border:1px solid {_panel_bd};border-radius:14px;
padding:14px 18px;margin-bottom:18px'>
  <div style='font-size:.72rem;font-weight:700;color:{_title_c};
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px'>
  🔎 현재 적용된 검색 조건</div>
  <div style='display:flex;flex-wrap:wrap;gap:4px;align-items:center;margin-bottom:6px'>
    <span style='font-size:.72rem;color:{_title_c};min-width:50px'>키워드</span>
    {_kw_tags}
  </div>
  <div style='display:flex;flex-wrap:wrap;gap:4px;align-items:center;margin-bottom:6px'>
    <span style='font-size:.72rem;color:{_title_c};min-width:50px'>정렬</span>
    {_sort_tags}
  </div>
  <div style='display:flex;flex-wrap:wrap;gap:4px;align-items:center;margin-bottom:6px'>
    <span style='font-size:.72rem;color:{_title_c};min-width:50px'>기간</span>
    {_date_tags}
  </div>
  <div style='display:flex;flex-wrap:wrap;gap:4px;align-items:center'>
    <span style='font-size:.72rem;color:{_title_c};min-width:50px'>길이</span>
    {_dur_tags}
    {_vtype_tag}
    {_max_tag}
  </div>
</div>""", unsafe_allow_html=True)

    # ── 탭 구성 ───────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "🎬 영상 목록", "📊 채널 통계", "🔑 키워드 분석", "📜 대본 전문"
    ])

    # ── 탭1: 영상 목록 ───────────────────────────────────────
    with tab1:
        for kw, videos in all_results.items():
            st.markdown(f"### 🔍 검색어: `{kw}`  &nbsp; ({len(videos)}개 영상 · {sort_label})")
            if not videos:
                st.warning("검색 결과가 없습니다.")
                continue

            for v in videos:
                _shorts_badge = "<span class='badge-shorts'>📱 Shorts</span>" if is_shorts(v) else ""
                badge_html = {
                    "🥇":"<span class='badge-hot'>🥇 1위</span>",
                    "🥈":"<span class='badge-good'>🥈 2위</span>",
                    "🥉":"<span class='badge-good'>🥉 3위</span>",
                    "🔥":"<span class='badge-hot'>🔥 인기</span>",
                    "⭐":"<span class='badge-new'>⭐ 우수</span>",
                }.get(v.get("badge","▶"), f"<span class='badge-norm'>#{v.get('rank',0)}위</span>")
                badge_html = badge_html + (" " + _shorts_badge if _shorts_badge else "")

                # ── 정렬 출처 태그 ──
                _STAG_C = {"조회수순":"#2563eb","최신순":"#059669","관련성순":"#d97706","평점순":"#7c3aed"}
                _STAG_ICON = {"조회수순":"👁","최신순":"🕐","관련성순":"🎯","평점순":"⭐"}
                _src_tags_html = "".join(
                    f"<span style='display:inline-flex;align-items:center;gap:2px;"
                    f"background:{_STAG_C.get(s,'#64748b')}18;"
                    f"border:1px solid {_STAG_C.get(s,'#64748b')}55;"
                    f"border-radius:10px;padding:2px 8px;"
                    f"font-size:.72rem;font-weight:600;color:{_STAG_C.get(s,'#64748b')};margin:1px 2px'>"
                    f"{_STAG_ICON.get(s,'')} {s}</span>"
                    for s in v.get("sort_sources", [])
                )
                badge_html = badge_html + (" " + _src_tags_html if _src_tags_html else "")

                with st.expander(f"{v.get('badge','▶')} #{v.get('rank',0)}위  {v['title']}", expanded=(v.get('rank',0)<=3)):
                    col_img, col_info = st.columns([1, 3])

                    with col_img:
                        if v.get("thumbnail"):
                            st.image(v["thumbnail"], use_container_width=True)
                        st.markdown(f"[▶ YouTube에서 보기]({v['url']})")

                        # ── 📋 전체 정보 복사 (st.code 방식 — Streamlit Cloud 호환) ──
                        _sep  = "=" * 60
                        _sep2 = "-" * 60
                        _transcript_text = v.get('transcript', '')
                        _has_transcript  = is_valid_transcript(_transcript_text)
                        _copy_text = (
f"""{_sep}
■ 제목
{v['title']}

■ 채널명
{v['channelTitle']}

■ URL
{v['url']}
{_sep2}
■ 핵심 키워드
""" +
(' '.join(f'#{k}' for k in v.get('keywords', [])) if v.get('keywords') else '(키워드 없음)') +
f"""
{_sep2}
■ 요약
""" +
(v.get('summary') or '(요약 없음)') +
f"""
{_sep2}
■ 태그 ({len(v.get('tags', []))}개)
""" +
(','.join(f'#{t}' for t in v.get('tags', [])) if v.get('tags') else '(태그 없음)') +
f"""
{_sep2}
■ 영상 설명
""" +
(v.get('description') or '(설명 없음)') +
f"""
{_sep}
■ 대본 전문 {'✅' if _has_transcript else '❌ (자막 없음)'}
{_sep}
""" +
(_transcript_text if _has_transcript else '이 영상에는 자막이 없습니다.') +
f"""
{_sep}
"""
                        )
                        # 토글 버튼으로 펼치기/접기
                        with st.expander("📋 전체 정보 복사 (대본 포함) — 우측 상단 □ 아이콘으로 복사"):
                            st.code(_copy_text, language=None)

                    with col_info:
                        st.markdown(f"""
<div>
{badge_html}
<span class='video-title'> {v['title']}</span><br>
<span class='video-meta'>
📺 {v['channelTitle']} &nbsp;|&nbsp; 구독자 {v['subscriberLabel']} &nbsp;|&nbsp;
⏱ {v['duration']} &nbsp;|&nbsp; 📅 {v['publishedAt']}
</span>
</div>
<div class='stat-row'>
  <span class='stat-item'>👁️ 조회수 {v['viewLabel']}</span>
  <span class='stat-item'>👍 좋아요 {v['likeLabel']}</span>
  <span class='stat-item'>💬 댓글 {v['commentLabel']}</span>
</div>
""", unsafe_allow_html=True)

                        if v.get("keywords"):
                            st.markdown("<div class='section-title'>🔑 핵심 키워드</div>", unsafe_allow_html=True)
                            kw_html = "".join(f"<span class='keyword-tag'>{k}</span>" for k in v["keywords"][:12])
                            st.markdown(kw_html, unsafe_allow_html=True)

                        if v.get("summary") and v["summary"] != "(요약 없음)":
                            st.markdown("<div class='section-title'>📋 요약</div>", unsafe_allow_html=True)
                            st.markdown(
                                f"<div class='summary-text'>{v['summary']}</div>",
                                unsafe_allow_html=True
                            )

                        if v.get("tags"):
                            with st.expander(f"🏷️ 태그 ({len(v['tags'])}개)"):
                                st.write("  |  ".join(v["tags"][:20]))

                        if v.get("description"):
                            with st.expander("📝 영상 설명"):
                                st.text(v["description"][:800] + ("..." if len(v["description"])>800 else ""))

    # ── 탭2: 채널 통계 ──────────────────────────────────────
    with tab2:
        st.markdown(f"### 📊 채널별 통계 (총 {len(channel_stats)}개 채널)")
        import pandas as pd
        df_ch = pd.DataFrame([{
            "채널명":     cs["channel"],
            "구독자":     cs["subscriber"],
            "영상수":     cs["videoCount"],
            "총조회수":   fmt(cs["totalView"])+"회",
            "평균조회수": fmt(cs["avgView"])+"회",
            "총좋아요":   fmt(cs["totalLike"]),
            "평균좋아요": fmt(cs["avgLike"]),
            "총댓글":     fmt(cs["totalComment"]),
        } for cs in channel_stats])
        st.dataframe(df_ch, use_container_width=True, hide_index=True)

        st.markdown("### 🏆 채널별 상세")
        for cs in channel_stats[:10]:
            with st.expander(f"📺 {cs['channel']} | 구독자 {cs['subscriber']} | 영상 {cs['videoCount']}개"):
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("총조회수",   fmt(cs["totalView"])+"회")
                c2.metric("평균조회수", fmt(cs["avgView"])+"회")
                c3.metric("총좋아요",   fmt(cs["totalLike"]))
                c4.metric("평균댓글",   fmt(cs["avgComment"]))

                st.markdown("**소속 영상:**")
                for v in cs["videos"][:5]:
                    st.markdown(f"- [{v['title']}]({v['url']}) — 조회수 {v['viewLabel']}")

    # ── 탭3: 키워드 분석 ────────────────────────────────────
    with tab3:
        st.markdown("### 🔑 검색어별 키워드 분석")
        for kw, videos in all_results.items():
            with st.expander(f"🔍 [{kw}]  ({len(videos)}개 영상)", expanded=True):
                all_kws = []
                for v in videos:
                    all_kws.extend(v.get("keywords", []))
                if all_kws:
                    counter = Counter(all_kws)
                    top20   = counter.most_common(20)

                    c1, c2 = st.columns([2,1])
                    with c1:
                        st.markdown("**🔝 상위 20개 키워드**")
                        kw_html = "".join(
                            f"<span class='keyword-tag' style='font-size:{0.8+count*0.03:.2f}rem'>{word}({count})</span>"
                            for word, count in top20
                        )
                        st.markdown(kw_html, unsafe_allow_html=True)
                    with c2:
                        st.markdown("**📈 키워드 빈도 TOP 10**")
                        import pandas as pd
                        df_kw = pd.DataFrame(top20[:10], columns=["키워드","빈도"])
                        st.dataframe(df_kw, hide_index=True, use_container_width=True)
                else:
                    st.info("키워드를 추출하려면 자막 가져오기를 체크하고 재검색하세요.")

                avg_v = sum(v["viewCount"] for v in videos) // len(videos) if videos else 0
                max_v = max((v["viewCount"] for v in videos), default=0)
                c1, c2, c3 = st.columns(3)
                c1.metric("영상 수", f"{len(videos)}개")
                c2.metric("평균 조회수", fmt(avg_v)+"회")
                c3.metric("최고 조회수", fmt(max_v)+"회")

    # ── 탭4: 대본 전문 ──────────────────────────────────────
    with tab4:
        st.markdown("### 📜 영상 대본 전문")
        # ✅ 중복 videoId 제거 (같은 영상이 여러 키워드에서 나올 때 key 충돌 방지)
        _seen_vids = set()
        transcript_videos = []
        for vs in all_results.values():
            for v in vs:
                if is_valid_transcript(v.get("transcript", "")) and v["videoId"] not in _seen_vids:
                    _seen_vids.add(v["videoId"])
                    transcript_videos.append(v)
        if not transcript_videos:
            st.info("자막이 있는 영상이 없습니다. '자막(대본) 가져오기'를 체크하고 재검색하세요.")
        else:
            st.caption(f"자막 있는 영상: {len(transcript_videos)}개")
            for _tv_i, v in enumerate(transcript_videos):
                with st.expander(f"📺 {v['title']} — {v['channelTitle']}"):
                    st.markdown(f"🔗 [{v['url']}]({v['url']})")
                    st.markdown(f"**길이:** {v['duration']} | **조회수:** {v['viewLabel']}")
                    st.markdown("---")
                    st.markdown(f'<div class="transcript-box">{v["transcript"]}</div>', unsafe_allow_html=True)
                    # ✅ key에 루프 인덱스 추가 → 완전한 유일성 보장
                    st.download_button(
                        label="📄 이 영상 대본 TXT 다운로드",
                        data=v["transcript"].encode("utf-8"),
                        file_name=f"transcript_{v['videoId']}.txt",
                        mime="text/plain",
                        key=f"dl_{v['videoId']}_{_tv_i}"
                    )


if __name__ == "__main__":
    main()